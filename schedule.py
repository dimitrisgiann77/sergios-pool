# -*- coding: utf-8 -*-
"""
Εστία — Module «Πρόγραμμα Εργασίας» (Βάρδιες) — Φάση 1 (v12.40)
==============================================================
Plug-in: `import schedule` από το ΤΕΛΟΣ του app.py (αφού οριστούν app/db/helpers,
ΠΡΙΝ το init_db() ώστε το create_all να πιάσει τους νέους πίνακες).
Spec: 02_MODULES_ESTIA/ΠΡΟΓΡΑΜΜΑ_ΕΡΓΑΣΙΑΣ/00_SPEC.md

Αρχές:
 - Κάθε εργαζόμενος = User (login_enabled=false για το μαζικό προσωπικό).
 - Τμήματα οργανισμού-wide· ξενοδοχείο = ξεχωριστή διάσταση (home_hotel_id).
 - Κωδικολόγιο μισθοδοσίας (8 κωδικοί) + μηχανή ωρών (8.5 normal, έξτρα, σπαστή, νυχτερινή).
 - Workflow υποβολής: WeekPlan(τμήμα) -> ScheduleSubmission(ενοποιημένη/ξενοδοχείο),
   κανόνες (R1 ≥1 ρεπό/εβδ.), χρονικά κλειδώματα (Πέμπτη), versions + diff.
"""
import os, re, json, unicodedata, threading
from datetime import datetime, date, timedelta, time as dtime
from flask import request, redirect, url_for, render_template, session, jsonify, Response
from app import (app, db, current_user, is_admin, allowed_hotels, notify, notify_admins,
                 log_activity, Hotel, User, Setting, ROLE_RANK, role_rank, BASE_DIR,
                 send_email, EMAIL_TO_LIST, active_hotel_id)
from werkzeug.security import generate_password_hash

# ── Σταθερές μηχανής ──────────────────────────────────────────────────────────
NORMAL_HOURS = 8.5          # κανονικό ωράριο/μέρα (περιλαμβάνει 30' διάλειμμα)
HOTEL_CODES  = {'AST', 'CNT', 'SRG', 'PSV', 'PLM', 'IRO', 'CND', 'ΗΡΩ'}
HOTEL_NORM   = {'ΗΡΩ': 'IRO'}

# Κωδικολόγιο βαρδιών (από το σύστημα μισθοδοσίας — seed στο ShiftType)
SHIFT_CODES = [
    # code,  label,                 color,     counts_as_work, note, ergani
    ('ΕΡΓ',   'Εργασία',             '#16a34a', True,  'έξτρα = διάρκεια − 8,5', 'WORK'),
    ('ΑΝ',    'Ρεπό',                '#185FA5', False, 'ρεπό/εβδομάδα',          'OFF'),
    ('ΑΔ',    'Κανονική άδεια',      '#64748b', False, 'το βγάζει το λογιστήριο','LEAVE'),
    ('ΑΣΘ',   'Αναρρωτική',          '#0e9f6e', False, 'το βγάζει το λογιστήριο','SICK'),
    ('ΑΠ',    'Αδικαιολόγητος απών', '#dc2626', False, '—',                      'ABSENT'),
    ('ΑΡΓ',   'Αργία (ρεπό αργίας)', '#7e22ce', False, 'αν δουλέψει → ΕΡΓ',      'HOLIDAY'),
    ('ΑΝΕΥ',  'Άνευ αποδοχών',       '#ea580c', False, 'το βγάζει το λογιστήριο','UNPAID'),
    ('Ειδ.Α', 'Ειδική άδεια',        '#0891b2', False, 'το βγάζει το λογιστήριο','SPECIAL'),
]

# Κανονικά τμήματα (οργανισμού-wide) + aliases (καθαρισμός Ελλ/Αγγλ διπλών)
DEPARTMENTS = [
    ('Housekeeping',  'Housekeeping', '#0ea5e9', ['housekeeping', 'hk']),
    ('Reception',     'Reception',    '#6366f1', ['reception', 'ρεσεψιον', 'υποδοχη']),
    ('Service',       'Service',      '#f59e0b', ['service', 'σερβις']),
    ('Kitchen',       'Kitchen',      '#ef4444', ['kitchen', 'κουζινα']),
    ('Maintenance',   'Maintenance',  '#64748b', ['maintenance', 'συντηρηση']),
    ('Management',    'Management',   '#7e22ce', ['management', 'διοικηση', 'operations']),
    ('Pool Bar',      'Pool Bar',     '#06b6d4', ['poolbar', 'pool bar', 'pool']),
    ('Bar',           'Bar',          '#0891b2', ['bar', 'μπαρ', 'barman']),
    ('Restaurant',    'Restaurant',   '#d97706', ['restaurant', 'oliva', 'μπουφε']),
    ("Kid's Club",    "Kid's Club",   '#ec4899', ["kid's club", 'kids club', 'kidsclub']),
    ('Bellboy',       'Bellboy',      '#0d9488', ['bellboy', 'γκρουμ', 'groom']),
    ('Replacement',   'Replacement',  '#94a3b8', ['replacement', 'replacant', 'replacant']),
]

WEEKDAYS_EL = ['Δευτέρα', 'Τρίτη', 'Τετάρτη', 'Πέμπτη', 'Παρασκευή', 'Σάββατο', 'Κυριακή']
DOW_EL = {0: 'Δευτέρα', 1: 'Τρίτη', 2: 'Τετάρτη', 3: 'Πέμπτη', 4: 'Παρασκευή', 5: 'Σάββατο', 6: 'Κυριακή'}
MONTHS_EL = ['', 'Ιανουάριος', 'Φεβρουάριος', 'Μάρτιος', 'Απρίλιος', 'Μάιος', 'Ιούνιος',
             'Ιούλιος', 'Αύγουστος', 'Σεπτέμβριος', 'Οκτώβριος', 'Νοέμβριος', 'Δεκέμβριος']

WP_STATUS = ('draft', 'ready', 'submitted', 'locked')
WP_LABELS  = {'draft': 'Πρόχειρο', 'ready': 'Έτοιμο', 'submitted': 'Υποβλήθηκε', 'locked': 'Κλειδωμένο'}


# ── Normalization helpers ─────────────────────────────────────────────────────
def _acc(s):
    return ''.join(c for c in unicodedata.normalize('NFD', str(s)) if unicodedata.category(c) != 'Mn')

def _norm(s):
    return re.sub(r'[^a-zα-ω0-9]', '', _acc(s).strip().lower()) if s else ''

def monday_of(d):
    """Δευτέρα της εβδομάδας που περιέχει το d."""
    return d - timedelta(days=d.weekday())


# ── ΜΟΝΤΕΛΑ ───────────────────────────────────────────────────────────────────
class Department(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    name      = db.Column(db.String(60), unique=True, nullable=False)
    name_en   = db.Column(db.String(60))
    color     = db.Column(db.String(9), default='#64748b')
    aliases   = db.Column(db.Text)            # JSON list κανονικοποιημένων aliases
    active    = db.Column(db.Boolean, default=True)
    sort      = db.Column(db.Integer, default=0)

class ShiftType(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    code          = db.Column(db.String(12), unique=True, nullable=False)
    label         = db.Column(db.String(60))
    color         = db.Column(db.String(9), default='#64748b')
    default_start = db.Column(db.String(5))   # 'HH:MM'
    default_end   = db.Column(db.String(5))
    counts_as_work= db.Column(db.Boolean, default=False)
    payroll_note  = db.Column(db.String(120))
    ergani_type   = db.Column(db.String(16))
    active        = db.Column(db.Boolean, default=True)
    sort          = db.Column(db.Integer, default=0)

class EmploymentProfile(db.Model):
    id              = db.Column(db.Integer, primary_key=True)
    user_id         = db.Column(db.Integer, db.ForeignKey('user.id'), unique=True, nullable=False)
    agreement_amount= db.Column(db.Float)               # συμφωνημένο μηνιαίο ποσό
    days_per_month  = db.Column(db.Integer, default=26)
    hours_per_day   = db.Column(db.Float, default=8.0)
    agreement_type  = db.Column(db.String(20), default='Μηνιαίος')   # Μηνιαίος/Management/Ωρομίσθιος
    position        = db.Column(db.String(80))          # θέση/ειδικότητα
    hired_at        = db.Column(db.Date)
    left_at         = db.Column(db.Date)
    status          = db.Column(db.String(20), default='Ενεργός')

    @property
    def day_wage(self):
        try:
            return round(self.agreement_amount / (self.days_per_month or 26), 4) if self.agreement_amount else 0.0
        except Exception:
            return 0.0
    @property
    def hour_wage(self):
        dw = self.day_wage
        return round(dw / 8.0, 4) if dw else 0.0

class ShiftAssignment(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    user_id       = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, index=True)
    work_date     = db.Column(db.Date, nullable=False, index=True)
    shift_code    = db.Column(db.String(12), default='ΕΡΓ')     # κωδικός ShiftType
    segments      = db.Column(db.Text)            # JSON [{'start':'07:00','end':'15:30'}, ...]
    work_hotel_id = db.Column(db.Integer, db.ForeignKey('hotel.id'))   # != home αν δανεικός
    note          = db.Column(db.String(200))
    created_by    = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at    = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    import_hash   = db.Column(db.String(40), index=True)
    __table_args__ = (db.UniqueConstraint('user_id', 'work_date', name='uq_user_date'),)

class Holiday(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    hol_date    = db.Column(db.Date, unique=True, nullable=False)
    description = db.Column(db.String(120))
    year        = db.Column(db.Integer, index=True)

class ScheduleRule(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    code        = db.Column(db.String(20), unique=True, nullable=False)
    description = db.Column(db.String(200))
    severity    = db.Column(db.String(8), default='block')   # block | warn
    params      = db.Column(db.Text)            # JSON
    active      = db.Column(db.Boolean, default=True)

class WeekPlan(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    hotel_id      = db.Column(db.Integer, db.ForeignKey('hotel.id'), nullable=False, index=True)
    department_id = db.Column(db.Integer, db.ForeignKey('department.id'), nullable=False, index=True)
    week_start    = db.Column(db.Date, nullable=False, index=True)   # Δευτέρα
    status        = db.Column(db.String(12), default='draft')
    updated_at    = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    updated_by    = db.Column(db.Integer, db.ForeignKey('user.id'))
    __table_args__ = (db.UniqueConstraint('hotel_id', 'department_id', 'week_start', name='uq_weekplan'),)

class ScheduleSubmission(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    hotel_id      = db.Column(db.Integer, db.ForeignKey('hotel.id'), nullable=False, index=True)
    week_start    = db.Column(db.Date, nullable=False, index=True)
    version       = db.Column(db.Integer, default=1)
    parent_version= db.Column(db.Integer)
    status        = db.Column(db.String(12), default='submitted')
    snapshot      = db.Column(db.Text)          # JSON: τι ακριβώς δηλώθηκε
    changes       = db.Column(db.Text)          # JSON diff από parent_version
    submitted_at  = db.Column(db.DateTime, default=datetime.utcnow)
    submitted_by  = db.Column(db.Integer, db.ForeignKey('user.id'))


# ── MIGRATION (στήλες User) ───────────────────────────────────────────────────
def ensure_schedule_columns():
    """Auto-migration: νέες στήλες στον User (μη καταστροφικό, race-safe μέσω app._add_col)."""
    try:
        from app import _add_col
    except Exception:
        return
    with app.app_context():
        _add_col('user', 'department_id',     'department_id INTEGER')
        _add_col('user', 'employer',          'employer VARCHAR(120)')
        _add_col('user', 'subunit',           'subunit VARCHAR(20)')
        _add_col('user', 'home_hotel_id',     'home_hotel_id INTEGER')
        _add_col('user', 'login_enabled',     'login_enabled BOOLEAN')
        _add_col('user', 'employment_active', 'employment_active BOOLEAN')


# ── ΜΗΧΑΝΗ ΩΡΩΝ ───────────────────────────────────────────────────────────────
_RANGE_RE = re.compile(r'(\d{1,2}):(\d{2})\s*[-–—]\s*(\d{1,2}):(\d{2})')

def segments_hours(segments):
    """segments: list[{'start':'HH:MM','end':'HH:MM'}] -> συνολικές ώρες (νυχτερινή->ίδια μέρα)."""
    tot = 0.0
    for seg in (segments or []):
        try:
            sh, sm = [int(x) for x in str(seg['start']).split(':')]
            eh, em = [int(x) for x in str(seg['end']).split(':')]
        except Exception:
            continue
        a = sh * 60 + sm
        b = eh * 60 + em
        if b <= a:
            b += 1440           # περνά μεσάνυχτα -> μετρά στη μέρα έναρξης
        tot += (b - a) / 60.0
    return round(tot, 4)

def extra_hours(total):
    return round(max(0.0, total - NORMAL_HOURS), 4) if total and total > 0 else 0.0

def parse_cell(v):
    """Κελί Excel -> (shift_code, segments[list], work_hotel_tag).
    Πιστή μεταφορά λογικής ENGINE_v2: τμήματα ΕΡΓ, κωδικοί, cross-hotel tag."""
    if v is None:
        return (None, [], None)
    s = str(v).strip()
    if not s:
        return (None, [], None)
    tag = None
    m = re.match(r'^\s*(AST|CNT|SRG|PSV|PLM|IRO|CND|ΗΡΩ)\b\s*[-: ]*\s*(.*)$', s, re.I)
    if m and (m.group(2).strip() == '' or 'ΕΡΓ' in m.group(2).upper() or _RANGE_RE.search(m.group(2))):
        tag = HOTEL_NORM.get(m.group(1).upper(), m.group(1).upper())
        s = m.group(2).strip()
    up = _acc(s).upper()
    ranges = _RANGE_RE.findall(s)
    if ranges or 'ΕΡΓ' in up or (tag is not None and s):
        segs = [{'start': f'{int(h1):02d}:{m1}', 'end': f'{int(h2):02d}:{m2}'} for h1, m1, h2, m2 in ranges]
        return ('ΕΡΓ', segs, tag)
    # κωδικοί χωρίς ώρες (σειρά: ειδικοί πριν τα γενικά)
    for c in ['ΑΣΘ', 'ΑΝΕΥ', 'ΑΡΓ', 'ΕΙΔ', 'ΑΔ', 'ΑΠ', 'ΑΑ', 'ΑΝ']:
        if up.startswith(_acc(c).upper()):
            return ({'ΕΙΔ': 'Ειδ.Α', 'ΑΑ': 'ΑΠ'}.get(c, c), [], None)
    return (None, [], None)   # άγνωστο -> αγνοείται

def assignment_hours(a):
    try:
        segs = json.loads(a.segments) if a.segments else []
    except Exception:
        segs = []
    return segments_hours(segs)

def is_work_code(code):
    st = ShiftType.query.filter_by(code=code).first()
    if st:
        return bool(st.counts_as_work)
    return code == 'ΕΡΓ'

def aggregate(assignments, home_hotel_id=None):
    """Σύνολα από λίστα ShiftAssignment: work_days, repo, sundays, holidays_worked, extra, elsewhere."""
    hol = {h.hol_date for h in Holiday.query.all()}
    work = repo = sundays = hol_worked = elsewhere = 0
    extra = 0.0
    for a in assignments:
        code = a.shift_code
        if is_work_code(code):
            work += 1
            extra += extra_hours(assignment_hours(a))
            if a.work_date.weekday() == 6:
                sundays += 1
            if a.work_date in hol:
                hol_worked += 1
            if a.work_hotel_id and home_hotel_id and a.work_hotel_id != home_hotel_id:
                elsewhere += 1
        elif code == 'ΑΝ':
            repo += 1
    return {'work_days': work, 'repo': repo, 'sundays': sundays,
            'holidays_worked': hol_worked, 'extra_hours': round(extra, 2),
            'elsewhere_days': elsewhere, 'total_days': work}

def monthly_settlement(year, month, hotel_id=None):
    """Λίστα «ΠΡΟΣ ΛΟΓΙΣΤΗΡΙΟ» ανά εργαζόμενο για μήνα."""
    start = date(year, month, 1)
    end = date(year + (month // 12), (month % 12) + 1, 1)
    q = (db.session.query(ShiftAssignment)
         .filter(ShiftAssignment.work_date >= start, ShiftAssignment.work_date < end))
    rows = q.all()
    by_user = {}
    for a in rows:
        by_user.setdefault(a.user_id, []).append(a)
    out = []
    for uid, alist in by_user.items():
        u = User.query.get(uid)
        if not u:
            continue
        if hotel_id and getattr(u, 'home_hotel_id', None) != hotel_id:
            continue
        agg = aggregate(alist, getattr(u, 'home_hotel_id', None))
        prof = EmploymentProfile.query.filter_by(user_id=uid).first()
        payable = 0.0
        if prof and prof.agreement_amount:
            if prof.agreement_type == 'Management':
                payable = round(prof.agreement_amount, 2)
            else:
                payable = round(prof.day_wage * agg['total_days'] + extra_wage(prof, agg['extra_hours']), 2)
        out.append({'user': u, 'agg': agg, 'payable': payable, 'profile': prof})
    out.sort(key=lambda r: r['user'].full_name or '')
    return out

def extra_wage(prof, hours):
    try:
        return round((prof.hour_wage or 0) * (hours or 0), 2)
    except Exception:
        return 0.0


# ── POLICY / RULES helpers ────────────────────────────────────────────────────
POLICY_DEFAULTS = {
    'planning_horizon_weeks': 8,
    'cutoff_dow': 3,          # 0=Δευτ ... 3=Πέμπτη
    'cutoff_time': '18:00',
    'lead_days': 4,           # προθεσμία = lead_days πριν τη Δευτέρα της W (4 = Πέμπτη προηγ.)
    'allow_admin_override': 1,
}

def get_policy():
    pol = dict(POLICY_DEFAULTS)
    for row in Setting.query.filter(Setting.key.like('sched_%')).all():
        k = row.key[6:]
        if k in pol:
            try:
                pol[k] = type(POLICY_DEFAULTS[k])(row.value) if not isinstance(POLICY_DEFAULTS[k], str) else row.value
            except Exception:
                pass
    return pol

def set_policy(d):
    for k, v in d.items():
        if k in POLICY_DEFAULTS:
            row = Setting.query.get('sched_' + k)
            if not row:
                row = Setting(key='sched_' + k); db.session.add(row)
            row.value = str(v)
    db.session.commit()

def week_deadline(week_start):
    """datetime προθεσμίας οριστικοποίησης για εβδομάδα που ξεκινά (Δευτέρα) week_start."""
    pol = get_policy()
    try:
        hh, mm = [int(x) for x in str(pol['cutoff_time']).split(':')]
    except Exception:
        hh, mm = 18, 0
    d = week_start - timedelta(days=int(pol['lead_days']))
    return datetime.combine(d, dtime(hh, mm))

def week_editable(week_start, user=None):
    """True αν η εβδομάδα είναι ακόμη επεξεργάσιμη (πριν την προθεσμία) ή admin override."""
    if datetime.now() < week_deadline(week_start):
        return True
    if user is not None and role_rank(user.role) >= ROLE_RANK['admin'] and int(get_policy()['allow_admin_override']):
        return True
    return False


# ── ROLE helpers ──────────────────────────────────────────────────────────────
def is_accountant():
    u = current_user()
    return u is not None and (u.role == 'accountant' or role_rank(u.role) >= ROLE_RANK['admin'])

def can_edit_schedule():
    return is_admin() or (current_user() and role_rank(current_user().role) >= ROLE_RANK['manager'])

def resolve_department(name):
    """Ταίριασμα ονόματος τμήματος -> Department (exact -> alias normalized)."""
    if not name:
        return None
    n = _norm(name)
    for d in Department.query.all():
        if _norm(d.name) == n or _norm(d.name_en or '') == n:
            return d
        try:
            for al in json.loads(d.aliases or '[]'):
                if _norm(al) == n:
                    return d
        except Exception:
            pass
    return None


# ── SEED (idempotent) ─────────────────────────────────────────────────────────
def seed_schedule():
  with app.app_context():
    try:
        # ShiftTypes
        if not ShiftType.query.first():
            for i, (code, label, color, cw, note, erg) in enumerate(SHIFT_CODES):
                db.session.add(ShiftType(code=code, label=label, color=color, counts_as_work=cw,
                                         payroll_note=note, ergani_type=erg, sort=i,
                                         default_start='07:00' if code == 'ΕΡΓ' else None,
                                         default_end='15:30' if code == 'ΕΡΓ' else None))
        # Departments
        if not Department.query.first():
            for i, (name, en, color, aliases) in enumerate(DEPARTMENTS):
                db.session.add(Department(name=name, name_en=en, color=color,
                                          aliases=json.dumps(aliases, ensure_ascii=False), sort=i))
        # Rules
        if not ScheduleRule.query.first():
            db.session.add(ScheduleRule(code='R1_repo', severity='block',
                description='Κάθε εργαζόμενος ≥1 ρεπό (ΑΝ) ανά εβδομάδα', params='{}'))
            db.session.add(ScheduleRule(code='R2_complete', severity='block',
                description='Καμία κενή ημέρα για assigned εργαζόμενο (πληρότητα 7/7)', params='{}', active=False))
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f'seed_schedule skipped: {e}')


# ── IMPORT (workbook προγράμματος -> χρήστες/τμήματα/αναθέσεις) ────────────────
CODE_HOTELNAME = {
    'AST': 'Asterias', 'CNT': 'Central', 'SRG': 'Sergios',
    'PSV': 'Piskopiano', 'IRO': 'Iro',  # PLM: σκόπιμα ΧΩΡΙΣ αντιστοίχιση (Giannis)
}
_LABELS = {'τμημα': 'dept', 'ειδικοτητα': 'spec', 'εταιρεια': 'comp', 'υποκ': 'upok',
           'επωνυμο': 'epon', 'ονομα': 'onoma'}

def _resolve_hotel_by_code(code):
    if not code:
        return None
    code = HOTEL_NORM.get(str(code).strip().upper(), str(code).strip().upper())
    target = CODE_HOTELNAME.get(code)
    if not target:
        return None
    tn = _norm(target)
    for h in Hotel.query.all():
        if tn in _norm(h.name):
            return h
    return None

def _find_or_create_user(epon, onoma, hotel, dept, employer, upok):
    full = (str(epon).strip() + ' ' + (str(onoma).strip() if onoma else '')).strip()
    fn = _norm(full)
    # ταίριασμα σε υπάρχοντα χρήστη (ονομαστικά)
    for u in User.query.all():
        if _norm(u.full_name) == fn:
            user = u
            break
    else:
        base = _acc(full).lower().replace(' ', '.')
        uname = re.sub(r'[^a-z0-9.]', '', base) or ('emp' + str(ShiftAssignment.query.count() + 1))
        if User.query.filter_by(username=uname).first():
            uname = uname + '.' + str(User.query.count() + 1)
        user = User(username=uname[:50], password=generate_password_hash(os.urandom(8).hex()),
                    full_name=full[:100], role='staff', approved=True, is_active=True)
        db.session.add(user)
        db.session.flush()
    # μετα-στοιχεία (συμπληρώνουμε όσα λείπουν)
    if hotel and not getattr(user, 'home_hotel_id', None):
        user.home_hotel_id = hotel.id
    if dept and not getattr(user, 'department_id', None):
        user.department_id = dept.id
    if employer and not getattr(user, 'employer', None):
        user.employer = str(employer)[:120]
    if upok and not getattr(user, 'subunit', None):
        user.subunit = str(upok)[:20]
    if getattr(user, 'login_enabled', None) is None:
        user.login_enabled = False
    if getattr(user, 'employment_active', None) is None:
        user.employment_active = True
    return user

def import_schedule_workbook(source, only_year=None, created_by=None):
    """source = path ή bytes. Επιστρέφει στατιστικά. Idempotent (upsert ανά user/μέρα)."""
    import openpyxl, io
    if isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True)
    else:
        wb = openpyxl.load_workbook(source, data_only=True)
    stats = {'users_new': 0, 'users_seen': 0, 'assign_new': 0, 'assign_upd': 0,
             'no_hotel': 0, 'cells': 0}
    before_users = User.query.count()
    seen_users = set()
    for ws in wb.worksheets:
        mr, mc = ws.max_row, min(ws.max_column, 40)
        r = 1
        while r <= mr:
            a = ws.cell(r, 1).value
            if a and _norm(a) == 'τμημα':
                colmap = {}; datecols = []
                for c in range(1, mc + 1):
                    v = ws.cell(r, c).value
                    nv = _norm(v)
                    if nv in _LABELS:
                        colmap[_LABELS[nv]] = c
                    d = None
                    if isinstance(v, datetime):
                        d = v.date()
                    elif isinstance(v, str):
                        mm = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', v)
                        if mm:
                            d = date(int(mm.group(3)), int(mm.group(2)), int(mm.group(1)))
                    if d:
                        datecols.append((c, d))
                ce, co = colmap.get('epon'), colmap.get('onoma')
                rr = r + 1
                while rr <= mr and not (ws.cell(rr, 1).value and _norm(ws.cell(rr, 1).value) == 'τμημα'):
                    epon = ws.cell(rr, ce).value if ce else None
                    if epon and str(epon).strip():
                        onoma = ws.cell(rr, co).value if co else ''
                        upok = ws.cell(rr, colmap['upok']).value if colmap.get('upok') else None
                        dept_v = ws.cell(rr, colmap['dept']).value if colmap.get('dept') else None
                        comp_v = ws.cell(rr, colmap['comp']).value if colmap.get('comp') else None
                        hotel = _resolve_hotel_by_code(upok)
                        dept = resolve_department(dept_v)
                        user = _find_or_create_user(epon, onoma, hotel, dept, comp_v, upok)
                        if user.id not in seen_users:
                            seen_users.add(user.id)
                        if not hotel:
                            stats['no_hotel'] += 1
                        for c, dt in datecols:
                            if only_year and dt.year != only_year:
                                continue
                            cell = ws.cell(rr, c).value
                            code, segs, tag = parse_cell(cell)
                            if code is None:
                                continue
                            stats['cells'] += 1
                            whid = None
                            if tag:
                                th = _resolve_hotel_by_code(tag)
                                whid = th.id if th else None
                            elif hotel:
                                whid = hotel.id
                            ex = ShiftAssignment.query.filter_by(user_id=user.id, work_date=dt).first()
                            if ex:
                                ex.shift_code = code
                                ex.segments = json.dumps(segs, ensure_ascii=False)
                                ex.work_hotel_id = whid
                                stats['assign_upd'] += 1
                            else:
                                db.session.add(ShiftAssignment(
                                    user_id=user.id, work_date=dt, shift_code=code,
                                    segments=json.dumps(segs, ensure_ascii=False),
                                    work_hotel_id=whid, created_by=created_by))
                                stats['assign_new'] += 1
                    rr += 1
                db.session.commit()
                r = rr
            else:
                r += 1
    wb.close()
    stats['users_new'] = User.query.count() - before_users
    stats['users_seen'] = len(seen_users)
    return stats


# ── ROUTES helpers ────────────────────────────────────────────────────────────
def _auth():
    return 'user_id' in session

def _week_arg():
    w = request.args.get('week')
    if w:
        try:
            return monday_of(datetime.strptime(w, '%Y-%m-%d').date())
        except Exception:
            pass
    return monday_of(date.today())

def _dept_users(hotel_id, dept_id):
    q = User.query.filter(User.is_active == True)
    if hotel_id:
        q = q.filter(User.home_hotel_id == hotel_id)
    if dept_id:
        q = q.filter(User.department_id == dept_id)
    return q.order_by(User.full_name).all()

def week_grid(hotel_id, dept_id, week_start):
    days = [week_start + timedelta(days=i) for i in range(7)]
    users = _dept_users(hotel_id, dept_id)
    uids = [u.id for u in users]
    amap = {}
    if uids:
        for a in (ShiftAssignment.query
                  .filter(ShiftAssignment.user_id.in_(uids))
                  .filter(ShiftAssignment.work_date >= days[0], ShiftAssignment.work_date <= days[6]).all()):
            amap[(a.user_id, a.work_date.isoformat())] = a
    rows = []
    for u in users:
        cells = []
        wk_hours = 0.0; repo = 0; work_days = 0
        for d in days:
            a = amap.get((u.id, d.isoformat()))
            if a:
                hrs = assignment_hours(a)
                if is_work_code(a.shift_code):
                    wk_hours += hrs; work_days += 1
                elif a.shift_code == 'ΑΝ':
                    repo += 1
                try:
                    segs = json.loads(a.segments) if a.segments else []
                except Exception:
                    segs = []
                label = a.shift_code
                if segs:
                    label = ' & '.join(f"{s['start']}-{s['end']}" for s in segs)
                cells.append({'date': d.isoformat(), 'code': a.shift_code, 'segs': segs,
                              'label': label, 'hours': round(hrs, 1),
                              'elsewhere': bool(a.work_hotel_id and a.work_hotel_id != hotel_id),
                              'note': a.note or ''})
            else:
                cells.append({'date': d.isoformat(), 'code': '', 'segs': [], 'label': '', 'hours': 0,
                              'elsewhere': False, 'note': ''})
        rows.append({'user': u, 'cells': cells, 'wk_hours': round(wk_hours, 1), 'repo': repo, 'work_days': work_days})
    return days, rows

def validate_hotel_week(hotel_id, week_start):
    """Εφαρμογή ScheduleRules σε όλο το ξενοδοχείο/εβδομάδα. Επιστρέφει issues."""
    days = [week_start + timedelta(days=i) for i in range(7)]
    issues = []
    rules = {r.code: r for r in ScheduleRule.query.filter_by(active=True).all()}
    users = User.query.filter(User.is_active == True, User.home_hotel_id == hotel_id).all()
    for u in users:
        alist = (ShiftAssignment.query.filter(ShiftAssignment.user_id == u.id)
                 .filter(ShiftAssignment.work_date >= days[0], ShiftAssignment.work_date <= days[6]).all())
        if not alist:
            continue
        codes = [a.shift_code for a in alist]
        if 'R1_repo' in rules and codes.count('ΑΝ') < 1:
            issues.append({'user': u, 'rule': 'R1_repo', 'severity': rules['R1_repo'].severity,
                           'msg': f'{u.full_name}: κανένα ρεπό αυτή την εβδομάδα'})
        if 'R2_complete' in rules and len(alist) < 7:
            issues.append({'user': u, 'rule': 'R2_complete', 'severity': rules['R2_complete'].severity,
                           'msg': f'{u.full_name}: {7 - len(alist)} κενές ημέρες'})
    return issues


# ── ROUTE: Board (multi-week: επιλογή ξενοδοχείου/τμήματος + Ν εβδομάδες) ──────
def _build_block(hotel_id, dept_id, week_start, user):
    days, rows = week_grid(hotel_id, dept_id, week_start)
    sub = None
    if hotel_id:
        sub = (ScheduleSubmission.query.filter_by(hotel_id=hotel_id, week_start=week_start)
               .order_by(ScheduleSubmission.version.desc()).first())
    return {
        'week_start': week_start, 'days': days, 'rows': rows,
        'editable': week_editable(week_start, user) and can_edit_schedule(),
        'issues': validate_hotel_week(hotel_id, week_start) if hotel_id else [],
        'deadline': week_deadline(week_start), 'sub': sub,
        'iso': week_start.isoformat(),
        'label': f"{week_start.strftime('%d/%m')} – {(week_start + timedelta(days=6)).strftime('%d/%m/%Y')}",
    }

@app.route('/dashboard/schedule')
def schedule_board():
    if not _auth():
        return redirect(url_for('login'))
    user = current_user()
    hotels = allowed_hotels(user)
    hotel_id = request.args.get('hotel_id', type=int) or active_hotel_id() or (hotels[0].id if hotels else None)
    depts = Department.query.filter_by(active=True).order_by(Department.sort, Department.name).all()
    dept_id = request.args.get('department_id', type=int) or (depts[0].id if depts else None)
    week_start = _week_arg()
    pol = get_policy()
    horizon = max(1, int(pol.get('planning_horizon_weeks', 8)))
    weeks = request.args.get('weeks', type=int) or 1
    weeks = max(1, min(weeks, horizon))
    blocks = [_build_block(hotel_id, dept_id, week_start + timedelta(days=7 * i), user) for i in range(weeks)]
    shift_types = ShiftType.query.filter_by(active=True).order_by(ShiftType.sort).all()
    shift_lookup = {st.code: st for st in shift_types}
    shift_types_json = json.dumps([{'code': st.code, 'color': st.color} for st in shift_types], ensure_ascii=False)
    cur_dept = Department.query.get(dept_id) if dept_id else None
    cur_hotel = Hotel.query.get(hotel_id) if hotel_id else None
    return render_template('schedule_board.html',
        shift_lookup=shift_lookup, shift_types_json=shift_types_json,
        hotels=hotels, depts=depts, hotel_id=hotel_id, dept_id=dept_id,
        cur_dept=cur_dept, cur_hotel=cur_hotel, weekdays=WEEKDAYS_EL,
        shift_types=shift_types, blocks=blocks, weeks=weeks, horizon=horizon,
        week_start=week_start, week_start_iso=week_start.isoformat(),
        prev_week=(week_start - timedelta(days=7)).isoformat(),
        next_week=(week_start + timedelta(days=7)).isoformat(),
        month_el=MONTHS_EL, is_admin=is_admin())


# ── API: autosave κελιού ──────────────────────────────────────────────────────
@app.route('/dashboard/schedule/cell', methods=['POST'])
def schedule_cell():
    if not _auth():
        return ('', 401)
    if not can_edit_schedule():
        return jsonify(ok=False, err='forbidden'), 403
    user = current_user()
    d = request.json or {}
    try:
        uid = int(d['user_id'])
        wd = datetime.strptime(d['date'], '%Y-%m-%d').date()
    except Exception:
        return jsonify(ok=False, err='bad'), 400
    if not week_editable(monday_of(wd), user):
        return jsonify(ok=False, err='locked'), 423
    code = (d.get('code') or '').strip()
    segs = d.get('segments') or []
    note = (d.get('note') or '')[:200]
    whid = d.get('work_hotel_id')
    a = ShiftAssignment.query.filter_by(user_id=uid, work_date=wd).first()
    if not code:
        if a:
            db.session.delete(a); db.session.commit()
        return jsonify(ok=True, deleted=True)
    # προεπιλεγμένες ώρες αν ΕΡΓ χωρίς segments
    if code == 'ΕΡΓ' and not segs:
        st = ShiftType.query.filter_by(code='ΕΡΓ').first()
        if st and st.default_start and st.default_end:
            segs = [{'start': st.default_start, 'end': st.default_end}]
    if not a:
        a = ShiftAssignment(user_id=uid, work_date=wd, created_by=user.id)
        db.session.add(a)
    a.shift_code = code
    a.segments = json.dumps(segs, ensure_ascii=False)
    a.work_hotel_id = whid
    a.note = note
    # WeekPlan -> draft (αν δεν υπάρχει)
    u = User.query.get(uid)
    if u and getattr(u, 'home_hotel_id', None) and getattr(u, 'department_id', None):
        wp = WeekPlan.query.filter_by(hotel_id=u.home_hotel_id, department_id=u.department_id,
                                      week_start=monday_of(wd)).first()
        if not wp:
            wp = WeekPlan(hotel_id=u.home_hotel_id, department_id=u.department_id,
                          week_start=monday_of(wd), status='draft')
            db.session.add(wp)
        if wp.status in ('submitted', 'locked'):
            wp.status = 'draft'
        wp.updated_by = user.id
    db.session.commit()
    hrs = assignment_hours(a)
    return jsonify(ok=True, hours=round(hrs, 1), work=is_work_code(code))


# ── Αντιγραφή προηγούμενης εβδομάδας ──────────────────────────────────────────
@app.route('/dashboard/schedule/copyprev', methods=['POST'])
def schedule_copyprev():
    if not _auth() or not can_edit_schedule():
        return redirect(url_for('login'))
    user = current_user()
    hotel_id = request.form.get('hotel_id', type=int)
    dept_id = request.form.get('department_id', type=int)
    week_start = monday_of(datetime.strptime(request.form['week'], '%Y-%m-%d').date())
    if not week_editable(week_start, user):
        return redirect(f'/dashboard/schedule?hotel_id={hotel_id}&department_id={dept_id}&week={week_start}&embed=1&err=locked')
    prev = week_start - timedelta(days=7)
    users = _dept_users(hotel_id, dept_id)
    n = 0
    for u in users:
        for i in range(7):
            sd, dd = prev + timedelta(days=i), week_start + timedelta(days=i)
            src = ShiftAssignment.query.filter_by(user_id=u.id, work_date=sd).first()
            if not src:
                continue
            dst = ShiftAssignment.query.filter_by(user_id=u.id, work_date=dd).first()
            if not dst:
                dst = ShiftAssignment(user_id=u.id, work_date=dd, created_by=user.id)
                db.session.add(dst)
            dst.shift_code = src.shift_code
            dst.segments = src.segments
            dst.work_hotel_id = src.work_hotel_id
            n += 1
    db.session.commit()
    log_activity('schedule_copyprev', f'{n} κελιά', hotel_id=hotel_id)
    return redirect(f'/dashboard/schedule?hotel_id={hotel_id}&department_id={dept_id}&week={week_start}&embed=1&ok=copied')


# ── ΥΠΟΒΟΛΗ (ενοποιημένη ανά ξενοδοχείο-εβδομάδα) ─────────────────────────────
def _hotel_week_snapshot(hotel_id, week_start):
    days = [week_start + timedelta(days=i) for i in range(7)]
    users = User.query.filter(User.is_active == True, User.home_hotel_id == hotel_id).all()
    snap = {}
    for u in users:
        alist = (ShiftAssignment.query.filter(ShiftAssignment.user_id == u.id)
                 .filter(ShiftAssignment.work_date >= days[0], ShiftAssignment.work_date <= days[6]).all())
        if not alist:
            continue
        dd = {}
        for a in alist:
            dd[a.work_date.isoformat()] = {'code': a.shift_code, 'segs': a.segments or '[]',
                                           'wh': a.work_hotel_id}
        snap[str(u.id)] = {'name': u.full_name, 'dept': getattr(u, 'department_id', None), 'days': dd}
    return snap

def _diff_snapshots(old, new):
    changes = []
    old = old or {}
    keys = set(old.keys()) | set(new.keys())
    for uid in keys:
        oname = (old.get(uid) or new.get(uid) or {}).get('name', uid)
        od = (old.get(uid) or {}).get('days', {})
        nd = (new.get(uid) or {}).get('days', {})
        for d in sorted(set(od.keys()) | set(nd.keys())):
            ov = od.get(d); nv = nd.get(d)
            if (ov or {}).get('code') != (nv or {}).get('code') or (ov or {}).get('segs') != (nv or {}).get('segs'):
                changes.append({'user_id': uid, 'name': oname, 'date': d,
                                'old': (ov or {}).get('code', '—'), 'new': (nv or {}).get('code', '—')})
    return changes

@app.route('/dashboard/schedule/submit', methods=['POST'])
def schedule_submit():
    if not _auth() or not can_edit_schedule():
        return redirect(url_for('login'))
    user = current_user()
    hotel_id = request.form.get('hotel_id', type=int)
    week_start = monday_of(datetime.strptime(request.form['week'], '%Y-%m-%d').date())
    issues = validate_hotel_week(hotel_id, week_start)
    blockers = [i for i in issues if i['severity'] == 'block']
    if blockers:
        return redirect(f'/dashboard/schedule?hotel_id={hotel_id}&week={week_start}&embed=1&err=rules')
    last = (ScheduleSubmission.query.filter_by(hotel_id=hotel_id, week_start=week_start)
            .order_by(ScheduleSubmission.version.desc()).first())
    snap = _hotel_week_snapshot(hotel_id, week_start)
    version = (last.version + 1) if last else 1
    changes = _diff_snapshots(json.loads(last.snapshot) if last and last.snapshot else None, snap) if last else []
    sub = ScheduleSubmission(hotel_id=hotel_id, week_start=week_start, version=version,
                             parent_version=(last.version if last else None),
                             status='submitted', snapshot=json.dumps(snap, ensure_ascii=False),
                             changes=json.dumps(changes, ensure_ascii=False),
                             submitted_by=user.id)
    db.session.add(sub)
    for wp in WeekPlan.query.filter_by(hotel_id=hotel_id, week_start=week_start).all():
        wp.status = 'submitted'
    db.session.commit()
    log_activity('schedule_submit', f'v{version}', hotel_id=hotel_id)
    # ειδοποίηση + email λογιστηρίου (background)
    hotel = Hotel.query.get(hotel_id)
    hn = hotel.name if hotel else ''
    label = 'ΤΡΟΠΟΠΟΙΗΣΗ' if version > 1 else 'Νέα υποβολή'
    _notify_accountants(hotel_id, week_start, version, label, changes)
    return redirect(f'/dashboard/schedule?hotel_id={hotel_id}&week={week_start}&embed=1&ok=submitted')

def _notify_accountants(hotel_id, week_start, version, label, changes):
    try:
        hotel = Hotel.query.get(hotel_id)
        hn = hotel.name if hotel else ''
        wk = week_start.strftime('%d/%m/%Y')
        accts = User.query.filter_by(role='accountant', is_active=True).all()
        for a in accts:
            notify(a.id, f'Πρόγραμμα {hn} — εβδ. {wk} ({label} v{version})', '/dashboard/schedule/submissions?embed=1')
        db.session.commit()
        recips = [a.email for a in accts if a.email] + list(EMAIL_TO_LIST)
        rows = ''.join(f"<tr><td>{c['name']}</td><td>{c['date']}</td><td>{c['old']}</td><td>{c['new']}</td></tr>" for c in changes)
        chtml = (f"<p><b>Αλλαγές (v{version}):</b></p><table border=1 cellpadding=4 style='border-collapse:collapse'>"
                 f"<tr><th>Εργαζόμενος</th><th>Ημ/νία</th><th>Πριν</th><th>Μετά</th></tr>{rows}</table>") if changes else ''
        html = (f"<h3>Πρόγραμμα Εργασίας — {hn}</h3>"
                f"<p>Εβδομάδα <b>{wk}</b> — <b>{label}</b> (έκδοση v{version}).</p>{chtml}"
                f"<p>Δες στο σύστημα: Εστία → Πρόγραμμα Εργασίας → Υποβολές.</p>")
        threading.Thread(target=lambda: send_email(f'[Εστία] Πρόγραμμα {hn} — εβδ. {wk} ({label})', html, recips)).start()
    except Exception as e:
        db.session.rollback()
        print('notify_accountants:', e)

@app.route('/dashboard/schedule/submissions')
def schedule_submissions():
    if not _auth():
        return redirect(url_for('login'))
    if not (is_admin() or is_accountant()):
        return ('Δεν έχετε πρόσβαση', 403)
    user = current_user()
    hids = {h.id for h in allowed_hotels(user)}
    q = ScheduleSubmission.query.order_by(ScheduleSubmission.submitted_at.desc())
    aid = active_hotel_id()
    if aid:
        q = q.filter(ScheduleSubmission.hotel_id == aid)
    subs = [s for s in q.limit(300).all() if (not hids) or s.hotel_id in hids]
    hotels = {h.id: h.name for h in Hotel.query.all()}
    return render_template('schedule_submissions.html', subs=subs, hotels=hotels,
                           month_el=MONTHS_EL)

@app.route('/dashboard/schedule/submission/<int:sid>')
def schedule_submission_view(sid):
    if not _auth():
        return redirect(url_for('login'))
    if not (is_admin() or is_accountant()):
        return ('Δεν έχετε πρόσβαση', 403)
    sub = ScheduleSubmission.query.get_or_404(sid)
    try:
        snap = json.loads(sub.snapshot or '{}')
        changes = json.loads(sub.changes or '[]')
    except Exception:
        snap, changes = {}, []
    days = [sub.week_start + timedelta(days=i) for i in range(7)]
    changed = {(c['user_id'], c['date']) for c in changes}
    hotel = Hotel.query.get(sub.hotel_id)
    return render_template('schedule_submission_view.html', sub=sub, snap=snap, changes=changes,
                           days=days, weekdays=WEEKDAYS_EL, changed=changed,
                           hotel=hotel, month_el=MONTHS_EL)


# ── EXPORT «ΠΡΟΣ ΛΟΓΙΣΤΗΡΙΟ» (μηνιαίο) ────────────────────────────────────────
@app.route('/dashboard/schedule/export.xlsx')
def schedule_export():
    if not _auth() or not (is_admin() or is_accountant()):
        return redirect(url_for('login'))
    import openpyxl, io
    year = request.args.get('year', type=int) or date.today().year
    month = request.args.get('month', type=int) or date.today().month
    hotel_id = request.args.get('hotel_id', type=int) or active_hotel_id()
    data = monthly_settlement(year, month, hotel_id)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'ΠΡΟΣ ΛΟΓΙΣΤΗΡΙΟ'
    ws.append(['Ονοματεπώνυμο', 'Μήνας', 'Καθημερινές εργάσιμες', 'Κυριακές', 'Αργίες',
               'Έξτρα ώρες', 'Ρεπό', 'Μέρες αλλού', 'Σύνολο ημερών', 'Πληρωτέο συμφωνίας'])
    mname = MONTHS_EL[month]
    for r in data:
        a = r['agg']
        ws.append([r['user'].full_name, mname, a['work_days'], a['sundays'], a['holidays_worked'],
                   a['extra_hours'], a['repo'], a['elsewhere_days'], a['total_days'], r['payable']])
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    fn = f'PROS_LOGISTIRIO_{year}_{month:02d}.xlsx'
    return Response(bio.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={fn}'})


# ── IMPORT page (upload workbook) ─────────────────────────────────────────────
@app.route('/dashboard/schedule/import', methods=['GET', 'POST'])
def schedule_import():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    res = None
    if request.method == 'POST':
        f = request.files.get('file')
        if f:
            try:
                res = import_schedule_workbook(f.read(), only_year=request.form.get('year', type=int),
                                               created_by=session.get('user_id'))
                seed_schedule()
                log_activity('schedule_import', str(res))
            except Exception as e:
                res = {'error': str(e)}
    return render_template('schedule_import.html', res=res, year=date.today().year)


# ── ADMIN settings (κωδικοί / τμήματα / αργίες / πολιτική / κανόνες) ───────────
@app.route('/dashboard/schedule/settings', methods=['GET', 'POST'])
def schedule_settings():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    if request.method == 'POST':
        act = request.form.get('action')
        if act == 'policy':
            set_policy({k: request.form.get(k) for k in POLICY_DEFAULTS if request.form.get(k) is not None})
        elif act == 'holiday':
            try:
                hd = datetime.strptime(request.form['hol_date'], '%Y-%m-%d').date()
                if not Holiday.query.filter_by(hol_date=hd).first():
                    db.session.add(Holiday(hol_date=hd, description=request.form.get('description', '')[:120], year=hd.year))
                    db.session.commit()
            except Exception:
                db.session.rollback()
        elif act == 'holiday_del':
            h = Holiday.query.get(request.form.get('id', type=int))
            if h:
                db.session.delete(h); db.session.commit()
        elif act == 'rule_toggle':
            r = ScheduleRule.query.get(request.form.get('id', type=int))
            if r:
                r.active = not r.active; db.session.commit()
        elif act == 'shift_toggle':
            s = ShiftType.query.get(request.form.get('id', type=int))
            if s:
                s.active = not s.active; db.session.commit()
        return redirect('/dashboard/schedule/settings?embed=1&ok=1')
    return render_template('schedule_settings.html',
        policy=get_policy(), shift_types=ShiftType.query.order_by(ShiftType.sort).all(),
        depts=Department.query.order_by(Department.sort).all(),
        holidays=Holiday.query.order_by(Holiday.hol_date).all(),
        rules=ScheduleRule.query.all(), dow_el=DOW_EL)


# ── ROUTE: Διαχείριση Προσωπικού (οργανόγραμμα: τμήμα/ξενοδοχείο/login) ────────
@app.route('/dashboard/schedule/staff', methods=['GET', 'POST'])
def schedule_staff():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    if request.method == 'POST':
        act = request.form.get('action')
        if act == 'add':
            full = (request.form.get('full_name') or '').strip()
            if full:
                base = re.sub(r'[^a-z0-9.]', '', _acc(full).lower().replace(' ', '.')) or 'emp'
                uname = base
                i = 1
                while User.query.filter_by(username=uname).first():
                    i += 1; uname = f'{base}.{i}'
                u = User(username=uname[:50], password=generate_password_hash(os.urandom(8).hex()),
                         full_name=full[:100], role='staff', approved=True, is_active=True,
                         department_id=request.form.get('department_id', type=int) or None,
                         home_hotel_id=request.form.get('home_hotel_id', type=int) or None,
                         employer=(request.form.get('employer') or '')[:120] or None,
                         login_enabled=bool(request.form.get('login_enabled')),
                         employment_active=True)
                db.session.add(u); db.session.commit()
                log_activity('staff_add', full)
        elif act == 'edit':
            u = User.query.get(request.form.get('user_id', type=int))
            if u:
                u.department_id = request.form.get('department_id', type=int) or None
                u.home_hotel_id = request.form.get('home_hotel_id', type=int) or None
                u.employer = (request.form.get('employer') or '')[:120] or None
                u.login_enabled = bool(request.form.get('login_enabled'))
                db.session.commit()
        elif act == 'toggle_active':
            u = User.query.get(request.form.get('user_id', type=int))
            if u:
                u.employment_active = not bool(getattr(u, 'employment_active', True))
                db.session.commit()
        qs = f"?embed=1&hotel_id={request.form.get('f_hotel','')}&department_id={request.form.get('f_dept','')}"
        return redirect('/dashboard/schedule/staff' + qs)
    f_hotel = request.args.get('hotel_id', type=int)
    f_dept = request.args.get('department_id', type=int)
    q = User.query.filter(User.is_active == True)
    if f_hotel:
        q = q.filter(User.home_hotel_id == f_hotel)
    if f_dept:
        q = q.filter(User.department_id == f_dept)
    users = q.order_by(User.full_name).limit(800).all()
    dup_groups = find_dup_groups()
    return render_template('schedule_staff.html',
        users=users, depts=Department.query.order_by(Department.sort).all(),
        hotels=Hotel.query.filter_by(is_active=True).order_by(Hotel.name).all(),
        f_hotel=f_hotel, f_dept=f_dept, dup_groups=dup_groups,
        dept_map={d.id: d.name for d in Department.query.all()},
        hotel_map={h.id: h.name for h in Hotel.query.all()})


# ── ΕΚΚΑΘΑΡΙΣΗ / ΣΥΓΧΩΝΕΥΣΗ ΔΙΠΛΩΝ ΕΡΓΑΖΟΜΕΝΩΝ ──────────────────────────────
def _lev(a, b):
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + (ca != cb)))
        prev = cur
    return prev[-1]

def _name_parts(full):
    toks = _acc(full or '').lower().split()
    sur = toks[0] if toks else ''
    first = ' '.join(toks[1:]) if len(toks) > 1 else ''
    return _norm(sur), _norm(first)

def _likely_dup(u1, u2):
    s1, f1 = _name_parts(u1.full_name)
    s2, f2 = _name_parts(u2.full_name)
    if not s1 or s1 != s2:
        return False                       # διαφορετικό επώνυμο -> όχι
    if f1 == f2:
        return True                        # ίδιο πλήρες όνομα
    if not f1 or not f2:
        return True                        # ο ένας χωρίς μικρό (π.χ. "FARO" vs "FARO ANNA")
    if f1.startswith(f2) or f2.startswith(f1):
        return True
    if _lev(f1, f2) <= 2 and min(len(f1), len(f2)) >= 3:
        return True                        # JOEY vs JOY, typos
    return False

def find_dup_groups():
    users = User.query.filter(User.is_active == True).order_by(User.full_name).all()
    parent = {u.id: u.id for u in users}
    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]; x = parent[x]
        return x
    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra
    by_sur = {}
    for u in users:
        s, _ = _name_parts(u.full_name)
        by_sur.setdefault(s, []).append(u)
    for s, lst in by_sur.items():
        if not s or len(lst) < 2:
            continue
        for i in range(len(lst)):
            for j in range(i + 1, len(lst)):
                if _likely_dup(lst[i], lst[j]):
                    union(lst[i].id, lst[j].id)
    groups = {}
    umap = {u.id: u for u in users}
    for u in users:
        groups.setdefault(find(u.id), []).append(u)
    out = []
    for gid, members in groups.items():
        if len(members) > 1:
            # εμπλουτισμός: πλήθος αναθέσεων ανά μέλος (για επιλογή canonical)
            for m in members:
                m._assign_n = ShiftAssignment.query.filter_by(user_id=m.id).count()
            members.sort(key=lambda m: m._assign_n, reverse=True)
            out.append(members)
    out.sort(key=lambda g: g[0].full_name or '')
    return out

def merge_users(keep_id, drop_ids):
    keep = User.query.get(keep_id)
    if not keep:
        return 0
    moved = 0
    keep_dates = {a.work_date for a in ShiftAssignment.query.filter_by(user_id=keep_id).all()}
    # home_hotel heuristic: όπου έχει τις περισσότερες αναθέσεις συνολικά
    hotel_count = {}
    for d in drop_ids:
        u = User.query.get(d)
        if not u or u.id == keep_id:
            continue
        # μέτρα αναθέσεις ανά work_hotel
        for a in ShiftAssignment.query.filter_by(user_id=d).all():
            if a.work_date in keep_dates:
                db.session.delete(a)               # διπλή μέρα -> κράτα του keep
            else:
                a.user_id = keep_id; keep_dates.add(a.work_date); moved += 1
                if a.work_hotel_id:
                    hotel_count[a.work_hotel_id] = hotel_count.get(a.work_hotel_id, 0) + 1
        # EmploymentProfile: μετάφερε αν λείπει στον keep
        kp = EmploymentProfile.query.filter_by(user_id=keep_id).first()
        dp = EmploymentProfile.query.filter_by(user_id=d).first()
        if dp and not kp:
            dp.user_id = keep_id
        elif dp:
            db.session.delete(dp)
        # specialties (faults)
        try:
            import faults as _flt
            for us in _flt.UserSpecialty.query.filter_by(user_id=d).all():
                db.session.delete(us)
        except Exception:
            pass
        # fill blanks στον keep
        if not keep.department_id and u.department_id:
            keep.department_id = u.department_id
        if not keep.home_hotel_id and u.home_hotel_id:
            keep.home_hotel_id = u.home_hotel_id
        if not keep.employer and u.employer:
            keep.employer = u.employer
        if not keep.subunit and u.subunit:
            keep.subunit = u.subunit
        db.session.delete(u)
    db.session.commit()
    return moved

@app.route('/dashboard/schedule/staff/merge', methods=['POST'])
def schedule_staff_merge():
    if not _auth() or not is_admin():
        return redirect(url_for('login'))
    keep_id = request.form.get('keep_id', type=int)
    drop_ids = [int(x) for x in request.form.getlist('drop_ids') if x.isdigit() and int(x) != keep_id]
    if keep_id and drop_ids:
        n = merge_users(keep_id, drop_ids)
        log_activity('staff_merge', f'keep={keep_id} drop={drop_ids} moved={n}')
    return redirect('/dashboard/schedule/staff?embed=1&ok=merged')
