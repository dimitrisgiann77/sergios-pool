"""
Εστία (Estia) — CONDIAN HOTELS · Κεντρική πλατφόρμα προσωπικού (v12.25)
Ασφάλεια (v12.25): SECRET_KEY enforced σε production, session cookies (HttpOnly/SameSite/Secure),
rate-limit στο login, bootstrap admin & default κωδικοί ομάδας μέσω env (όχι σταθεροί στον κώδικα).
Backend: Flask + PostgreSQL + SMTP + AI Assistant

Modules:
  - Water Log (νερά χρήσης) — multi-hotel / multi-δίκτυο (v12.3)
  - Pool Log (πισίνες) — multi-hotel / multi-pool
  - SpithaAI (chat bubble) — provider-agnostic (Anthropic/OpenAI)
  - Records feed (v12.4) — ενιαία λίστα υποβολών πισινών & νερών
  - v12.5 — Πίνακας Πισινών: κεντρικό πάνελ (σημερινές μετρήσεις + charts + φίλτρα)·
            Records: εξαγωγή PDF & XLSX (έτοιμα προς εκτύπωση)
  - v12.6 — Shell header: ζωντανό ρολόι + καιρός/location, κουμπιά Guest/Staff/Admin (placeholder)·
            Footer (Εστία · All Rights Reserved 2026 · version/build · CONDIAN Hotels)
  - v12.7 — Βοηθός μετονομάστηκε «SpithaAI»· Διαχείριση Ξενοδοχείων & Πισινών (/dashboard/hotels) —
            μεταφέρθηκε από τον Πίνακα Πισινών στο μενού Διαχείρισης
  - v12.8 — Header: 4ο κουμπί Operations · ειδοποιήσεις στο καμπανάκι · dropdown προφίλ
            (προφίλ/μηνύματα/γλώσσα/ρυθμίσεις/λογαριασμός/θέμα/έξοδος) · Day/Dark/System (cookie) · γλώσσα el/en/uk
  - v12.9 — Login: φουτουριστικό AI background (neural-net canvas)· Μηνύματα μεταξύ χρηστών (Message)·
            σελίδες Ρυθμίσεις & Λογαριασμός μου· badge αδιάβαστων μηνυμάτων στο header
  - v12.10 — Pre-login redesign (login + εγγραφή): κοινό static/auth.css + auth.js· μεγαλύτερο logo με glow·
             tagline «we reinvent the way we work» (Space Grotesk)· language selector (globe popover)·
             footer (© 2026 CONDIAN Hotels + facebook/linkedin)· εγγραφή με Google/Apple (UI placeholders)
  - v12.11 — Εμφάνιση: ξεχωριστές ρυθμίσεις login/register (login_logo, μέγεθος, γραμματοσειρά, tagline)
             + social links (facebook/linkedin)· αφαιρέθηκε ο κύκλος πίσω από το logo
  - v12.12 — Dark mode ΠΑΝΤΟΥ μέσω after_request (auto-inject theme snippet σε κάθε σελίδα,
             εκτός login/register)· Χρήστες: πλήρες edit (username+γλώσσα) + reveal νέου κωδικού
  - v12.13 — Χρήστες: χρωματικά badges για όλους τους ρόλους (masteradmin/admin/manager/staff/viewer)·
             η διαχείριση χρήστη σε pop-up modal (αντί για ανάπτυξη)· delete μέσα στο modal
  - v12.14 — Module Βλαβοληψία Φάση 1 (faults.py): δέντρο κατηγοριών (240), 9 καταστάσεις/state machine,
             αυτόματη/χειροκίνητη ανάθεση, σχόλια 2 καναλιών, audit log, SLA seed, μαζική «Ολοκλήρωση»
  - v12.15 — Βλαβοληψία Φάση 2: admin ρυθμίσεις (ειδικότητες/SLA/tags/χάρτης editable)· δέντρο κατηγοριών·
             SLA badges· export xlsx/csv/pdf· ειδικότητες ανά χρήστη· overview KPI σε νέο Fault
  - v12.16 — Βλαβοληψία Φάση 2β: φωτό/συνημμένα (cover + γκαλερί/upload)· email ειδοποίηση νέας βλάβης
  - v12.17 — Μενού: «Πίνακας Βλαβών» στην Παρακολούθηση· «Αναφορά βλάβης»→«Δήλωση βλάβης» στη Συντήρηση
"""

from flask import Flask, render_template, request, redirect, url_for, session, jsonify, Response
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import text
from datetime import datetime, date, timedelta
import os, smtplib, threading, json, time, secrets, urllib.request, urllib.error, urllib.parse
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

app = Flask(__name__)

# Railway/Heroku δίνουν 'postgres://' — η SQLAlchemy 2.x θέλει 'postgresql://'
_db_url = os.environ.get('DATABASE_URL', 'sqlite:///water.db')
if _db_url.startswith('postgres://'):
    _db_url = _db_url.replace('postgres://', 'postgresql://', 1)
_IS_PRODUCTION = not _db_url.startswith('sqlite')

# ── v12.25 Ασφάλεια: SECRET_KEY ──
# Σε production (Postgres) ΑΠΑΙΤΕΙΤΑΙ SECRET_KEY από env — αλλιώς η εφαρμογή δεν ξεκινά.
# Το dev fallback επιτρέπεται ΜΟΝΟ τοπικά (SQLite).
_secret = os.environ.get('SECRET_KEY')
if not _secret:
    if _IS_PRODUCTION:
        raise RuntimeError('SECRET_KEY δεν εχει οριστει. Ορισε SECRET_KEY στις μεταβλητες περιβαλλοντος (Railway) πριν το deploy.')
    _secret = 'estia-dev-secret-change-me'   # μονο για τοπικο dev (SQLite)
app.secret_key = _secret

app.config['SQLALCHEMY_DATABASE_URI'] = _db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {'pool_pre_ping': True}
app.config['MAX_CONTENT_LENGTH'] = 2 * 1024 * 1024   # 2MB (avatar upload)

# ── v12.25 Ασφάλεια: session cookies ──
app.config['SESSION_COOKIE_HTTPONLY']    = True             # οχι προσβαση απο JavaScript (XSS)
app.config['SESSION_COOKIE_SAMESITE']    = 'Lax'            # μετριαζει CSRF
app.config['SESSION_COOKIE_SECURE']      = _IS_PRODUCTION   # cookies μονο μεσω HTTPS σε production
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)

# ── v12.25 Ασφαλεια: απλο rate-limit στο login (in-memory ανα worker) ──
_LOGIN_MAX_FAILS = 5     # μετα απο τοσες αποτυχιες...
_LOGIN_WINDOW    = 300   # ...μεσα σε τοσα δευτερολεπτα → προσωρινο κλειδωμα
_login_fails     = {}    # key(IP) -> [timestamps αποτυχιων]

def _login_key():
    xff = request.headers.get('X-Forwarded-For', '')
    return xff.split(',')[0].strip() if xff else (request.remote_addr or 'unknown')

def _login_blocked(key):
    now = time.time()
    fails = [t for t in _login_fails.get(key, []) if now - t < _LOGIN_WINDOW]
    _login_fails[key] = fails
    return len(fails) >= _LOGIN_MAX_FAILS

def _login_record_fail(key):
    _login_fails.setdefault(key, []).append(time.time())

def _login_reset(key):
    _login_fails.pop(key, None)

SMTP_SERVER    = 'condian.gr'
SMTP_PORT      = 465
# ΣΗΜ.: όρισε EMAIL_FROM/GRAPH_SENDER σε πραγματικό mailbox @condianhotels.gr μέσω env στο Railway
EMAIL_FROM     = os.environ.get('EMAIL_FROM', 'report@condian.gr')
EMAIL_PASSWORD = os.environ.get('EMAIL_PASSWORD', '')
EMAIL_TO_LIST  = ['dimitris@condianhotels.gr', 'm.xypakis@condianhotels.gr', 'g.giakoumakis@condianhotels.gr']
APP_URL        = os.environ.get('APP_URL', 'https://estia.condianhotels.gr')   # βάση για absolute links

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REG_CODE = os.environ.get('REG_CODE', 'condian2026')          # κωδικός εγγραφής προσωπικού
ENABLE_SCHEDULER = os.environ.get('ENABLE_SCHEDULER', 'true').lower() in ('1','true','yes','on')
REMINDER_HOUR = int(os.environ.get('REMINDER_HOUR', '18'))    # ώρα υπενθύμισης (Europe/Athens)

# Microsoft Graph (Office 365) — ενεργό αν οριστεί GRAPH_CLIENT_ID
GRAPH_TENANT_ID     = os.environ.get('GRAPH_TENANT_ID', '')
GRAPH_CLIENT_ID     = os.environ.get('GRAPH_CLIENT_ID', '')
GRAPH_CLIENT_SECRET = os.environ.get('GRAPH_CLIENT_SECRET', '')
GRAPH_SENDER        = os.environ.get('GRAPH_SENDER', EMAIL_FROM)

# ── AI Assistant config (provider-agnostic) ──
AI_PROVIDER       = os.environ.get('AI_PROVIDER', 'auto')   # auto | anthropic | openai
ANTHROPIC_API_KEY = os.environ.get('ANTHROPIC_API_KEY', '')
OPENAI_API_KEY    = os.environ.get('OPENAI_API_KEY', '')
AI_MODEL          = os.environ.get('AI_MODEL', '')          # override; αλλιώς default ανά πάροχο

POOL_ASSISTANT_PROMPT = """Είσαι το «SpithaAI», ο ψηφιακός βοηθός της πλατφόρμας Εστία (CONDIAN HOTELS).
Βοηθάς το προσωπικό συντήρησης και τους υπεύθυνους βάρδιας στην ασφαλή, καθαρή και
νόμιμη καθημερινή λειτουργία των πισινών και των δικτύων νερού/ΖΝΧ (νερά χρήσης, legionella).
# ΑΡΜΟΔΙΟΤΗΤΕΣ
- Χημεία νερού: ερμηνεία μετρήσεων (pH, ελεύθερο/ολικό χλώριο, αλκαλικότητα,
  κυανουρικό οξύ, σκληρότητα ασβεστίου, θερμοκρασία) και προτάσεις διόρθωσης.
- Δοσολογία χημικών: υπολογισμός ποσότητας με βάση τον όγκο κάθε πισίνας και την
  απόκλιση από τις επιθυμητές τιμές.
- Πρόγραμμα καθαρισμού: skimming, βούρτσισμα, σκούπισμα πυθμένα, καθαρισμός
  καλαθιών skimmer/προφίλτρου, backwash φίλτρων.
- Έλεγχοι εξοπλισμού: αντλίες, κυκλοφορία, στάθμη νερού, πίεση φίλτρων, διαρροές.
- Ασφάλεια & συμμόρφωση: διαύγεια νερού, σήμανση, βάθη, ναυαγοσώστης, τήρηση
  αρχείου μετρήσεων.
- Ημερολόγιο: καταγραφή μετρήσεων και ενεργειών ανά βάρδια και ανά πισίνα.
# ΛΕΙΤΟΥΡΓΙΑ
1. Ρώτα πρώτα ποια πισίνα αφορά και ζήτα τις τρέχουσες μετρήσεις αν δεν τις έχεις.
2. Δίνε σαφή, πρακτικά βήματα — «τι να κάνω τώρα», όχι θεωρία.
3. Στη δοσολογία δείξε όγκο, απόκλιση και τελική ποσότητα προϊόντος. Πάντα
   επιβεβαίωνε τον όγκο της πισίνας.
4. Σήμανε ως ΕΠΕΙΓΟΝ ό,τι είναι εκτός ορίων (χλώριο, θολό νερό, βλάβη αντλίας).
5. Αν μια τιμή ήταν προβληματική νωρίτερα, υπενθύμισε επανέλεγχο.
# ΟΡΙΑ ΑΣΦΑΛΕΙΑΣ
- Δεν αντικαθιστάς πιστοποιημένο τεχνικό ή τη νομοθεσία· τα όρια τιμών διαφέρουν
  ανά περιοχή — σύστησε επαλήθευση με τους τοπικούς υγειονομικούς κανονισμούς.
- Σε σοβαρά περιστατικά (υπερδοσολογία, ηλεκτρικό πρόβλημα, ατύχημα) δώσε άμεση
  οδηγία ασφαλείας, σύστησε κλείσιμο πισίνας και ειδοποίηση υπευθύνου.
- ΠΟΤΕ μην προτείνεις ανάμειξη χλωρίου με οξέα ή χημικών μεταξύ τους.
# ΥΦΟΣ
- Σύντομος, πρακτικός, απλή γλώσσα για προσωπικό βάρδιας.
- Checklists για ελέγχους και προγράμματα.
- Στο τέλος κάθε αναφοράς, πρότεινε τι να καταγραφεί στο log."""

db = SQLAlchemy(app)

# ──────────────────────────────────────────────────────────────────────────
#  POOL LIMITS  (min, max) — None means no limit on that side.
# ──────────────────────────────────────────────────────────────────────────
POOL_LIMITS = {
    'free_chlorine':     (0.4, 1.5),
    'combined_chlorine': (None, 0.5),
    'ph':                (7.2, 7.8),
    'temp':              (None, 32.0),
    'turbidity':         (None, 1.0),
    'cyanuric_acid':     (None, 75.0),
    'total_alkalinity':  (80.0, 120.0),
    'orp':               (650.0, None),
}

PARAM_LABELS = {
    'free_chlorine': 'Ελεύθερο χλώριο', 'combined_chlorine': 'Συνδεδεμένο χλώριο',
    'ph': 'pH', 'temp': 'Θερμοκρασία', 'turbidity': 'Θολότητα',
    'cyanuric_acid': 'Κυανουρικό οξύ', 'total_alkalinity': 'Ολική αλκαλικότητα', 'orp': 'ORP',
}

# Κανόνας ενεργειών όταν τιμή εκτός ορίων (low = κάτω από min, high = πάνω από max)
ACTION_RULES = {
    'free_chlorine':     {'low': 'Χαμηλό χλώριο — κάνε χλωρίωση και επανέλεγξε σε 30΄.',
                          'high': 'Υψηλό χλώριο — σταμάτα τη δοσομέτρηση/άσε να πέσει· απόφυγε χρήση μέχρι <1.5 mg/L.'},
    'combined_chlorine': {'high': 'Υψηλό δεσμευμένο χλώριο — υπερχλωρίωση (shock) + αερισμός· έλεγξε ανανέωση νερού.'},
    'ph':                {'low': 'Χαμηλό pH — πρόσθεσε pH plus (ανθρακική σόδα).',
                          'high': 'Υψηλό pH — πρόσθεσε pH minus (οξύ), σταδιακά.'},
    'temp':              {'high': 'Υψηλή θερμοκρασία — έλεγξε/μείωσε θέρμανση· παρακολούθησε χλώριο.'},
    'turbidity':         {'high': 'Θολό νερό — backwash φίλτρου, έλεγξε διήθηση/κυκλοφορία, εξέτασε κροκίδωση.'},
    'cyanuric_acid':     {'high': 'Υψηλό κυανουρικό οξύ — μερική ανανέωση νερού (αραίωση)· μείωσε σταθεροποιητή.'},
    'total_alkalinity':  {'low': 'Χαμηλή αλκαλικότητα — πρόσθεσε alkalinity up (ανθρακική σόδα).',
                          'high': 'Υψηλή αλκαλικότητα — πρόσθεσε οξύ σταδιακά.'},
    'orp':               {'low': 'Χαμηλό ORP — ανέβασε ελεύθερο χλώριο και ρύθμισε pH στο 7.2–7.6.'},
}
SAFETY_NOTE = 'Ποτέ μην αναμειγνύεις χημικά μεταξύ τους (ειδικά χλώριο με οξύ). Πρόσθεσε ένα-ένα, με την κυκλοφορία ανοιχτή.'

AREA_LABELS = {'pump': 'Αντλία', 'filter': 'Φίλτρο', 'chemicals': 'Χημικά / Δοσομέτρηση',
               'water': 'Νερό / Στάθμη', 'electrical': 'Ηλεκτρικό', 'app': 'Εφαρμογή', 'other': 'Άλλο'}

def _urgent(param, val):
    return ((param == 'free_chlorine' and val < 0.2) or
            (param == 'combined_chlorine' and val > 1.0) or
            (param == 'turbidity' and val > 2.0))

def compute_pool_actions(rec):
    out = []
    for key, rule in ACTION_RULES.items():
        val = getattr(rec, key, None)
        if val is None:
            continue
        mn, mx = POOL_LIMITS.get(key, (None, None))
        action = None
        if mn is not None and val < mn:
            action = rule.get('low')
        elif mx is not None and val > mx:
            action = rule.get('high')
        if action:
            out.append({'label': PARAM_LABELS.get(key, key), 'action': action, 'urgent': _urgent(key, val)})
    return out


# ── ΖΝΧ / Δίκτυο νερού: κανόνες προτεινόμενων ενεργειών (legionella) ─────────
def _clo2_act(name):
    return (f'ClO2 {name} <1 ppm: αύξησε τη δοσομέτρηση ClO2· έλεγξε δοσομετρική αντλία/απόθεμα.',
            f'ClO2 {name} >2 ppm: μείωσε τη δοσομέτρηση ClO2.')

# key WaterRecord -> (min, max, ενέργεια_low, ενέργεια_high). Όρια ίδια με το report νερού.
WATER_ACTION_RULES = {
    'temp_dhw_out':     (60.0, None, 'Κολεκτέρ ΖΝΧ <60°C: ανέβασε τη θερμοκρασία αποθήκευσης ≥60°C (κίνδυνος legionella)· έλεγξε λέβητα/εναλλάκτη/θερμοστάτη.', None),
    'temp_dhw_return':  (50.0, None, 'Επιστροφή ανακυκλοφορίας <50°C: ανεπαρκής ανακυκλοφορία· έλεγξε αντλία & βάνες ανακυκλοφορίας· εξέτασε θερμική απολύμανση/flushing.', None),
    'temp_kitchen_hot': (50.0, None, 'Ζεστό Κουζίνας <50°C: flushing του σημείου· έλεγξε ανακυκλοφορία/μόνωση γραμμής.', None),
    'temp_remote_hot':  (50.0, None, 'Ζεστό Απομακρυσμένου <50°C: flushing· έλεγξε ανακυκλοφορία (κρίσιμο τελευταίο σημείο δικτύου).', None),
    'temp_tank':        (None, 20.0, None, 'Δεξαμενή (κρύο) >20°C: εξέτασε ψύξη/μόνωση/ανανέωση νερού· κίνδυνος ανάπτυξης μικροβίων.'),
    'clo2_dhw_out':     (1.0, 2.0, *_clo2_act('Αναχώρηση ΖΝΧ')),
    'clo2_dhw_return':  (1.0, 2.0, *_clo2_act('Επιστροφή ΖΝΧ')),
    'clo2_tank':        (1.0, 2.0, *_clo2_act('Δεξαμενή')),
    'clo2_kitchen':     (1.0, 2.0, *_clo2_act('Κουζίνα')),
    'clo2_remote':      (1.0, 2.0, *_clo2_act('Απομακρυσμένο')),
    'clo2_ro':          (1.0, 2.0, *_clo2_act('Αντ. Όσμωση')),
}
WATER_ACTION_LABELS = {
    'temp_dhw_out': 'Κολεκτέρ ΖΝΧ (Αναχ.)', 'temp_dhw_return': 'Κολεκτέρ Ανακυκλ. (Επιστρ.)',
    'temp_kitchen_hot': 'Κουζίνα Ζεστό', 'temp_remote_hot': 'Απομακρυσμένο Ζεστό', 'temp_tank': 'Δεξαμενή',
    'clo2_dhw_out': 'ClO2 Αναχώρηση ΖΝΧ', 'clo2_dhw_return': 'ClO2 Επιστροφή ΖΝΧ',
    'clo2_tank': 'ClO2 Δεξαμενή', 'clo2_kitchen': 'ClO2 Κουζίνα', 'clo2_remote': 'ClO2 Απομακρυσμένο', 'clo2_ro': 'ClO2 Αντ. Όσμωση',
}

def _water_urgent(key, val):
    if key == 'temp_dhw_out' and val < 50: return True
    if key in ('temp_dhw_return', 'temp_kitchen_hot', 'temp_remote_hot') and val < 45: return True
    if key in ('clo2_dhw_out', 'clo2_dhw_return') and val < 0.3: return True
    return False

def compute_water_actions(rec):
    """Προτεινόμενες ενέργειες ΖΝΧ/δικτύου νερού όταν μέτρηση εκτός ορίων. Δεν αγγίζει πισίνες."""
    out = []
    for key, (mn, mx, low, high) in WATER_ACTION_RULES.items():
        val = getattr(rec, key, None)
        if val is None:
            continue
        action = None
        if mn is not None and val < mn:
            action = low
        elif mx is not None and val > mx:
            action = high
        if action:
            out.append({'label': WATER_ACTION_LABELS.get(key, key), 'action': action, 'urgent': _water_urgent(key, val)})
    return out


def notify(user_id, text, link=None):
    if not user_id:
        return
    try:
        db.session.add(Notification(user_id=user_id, text=(text or '')[:300], link=link))
        db.session.commit()
    except Exception:
        db.session.rollback()

def notify_admins(text, link=None):
    for u in User.query.filter(User.is_active == True, User.role.in_(['admin', 'masteradmin'])).all():
        notify(u.id, text, link)

def send_fault_email(fr):
    if not EMAIL_PASSWORD and not GRAPH_CLIENT_ID:
        return False
    label = AREA_LABELS.get(fr.area, fr.area)
    who = fr.user.full_name if fr.user else '-'
    where = (fr.pool.hotel.name + ' / ' + fr.pool.name) if (fr.pool and fr.pool.hotel) else (fr.pool.name if fr.pool else '-')
    html = ('<div style="font-family:Arial,sans-serif;max-width:600px;">'
            '<div style="background:#b91c1c;color:white;padding:16px;border-radius:8px 8px 0 0;"><h2 style="margin:0;">Αναφορα βλαβης</h2></div>'
            '<div style="background:#f9f9f9;padding:16px;border:1px solid #eee;">'
            + '<p><b>Τομεας:</b> ' + label + '</p><p><b>Σημειο:</b> ' + where + '</p><p><b>Απο:</b> ' + who + '</p>'
            + '<p><b>Περιγραφη:</b><br>' + (fr.message or '') + '</p></div></div>')
    return send_email('Αναφορα βλαβης - ' + label, html, EMAIL_TO_LIST)


# ── ROLES & PERMISSIONS ──
ROLE_RANK = {'viewer': 0, 'staff': 1, 'manager': 2, 'admin': 3, 'masteradmin': 4}
ROLE_LABELS = {'staff': 'Staff (καταγραφή)', 'manager': 'Manager (επόπτης)', 'admin': 'Admin', 'masteradmin': 'Master Admin'}

def role_rank(role):
    return ROLE_RANK.get(role, 0)

def current_user():
    uid = session.get('user_id')
    return User.query.get(uid) if uid else None

@app.context_processor
def inject_nav():
    uid = session.get('user_id')
    unread = Notification.query.filter_by(user_id=uid, is_read=False).count() if uid else 0
    msgs   = Message.query.filter_by(recipient_id=uid, is_read=False).count() if uid else 0
    return {'nav_unread': unread, 'msg_unread': msgs}

@app.context_processor
def inject_theme():
    return {'theme': get_theme()}

# έκδοση/build για το footer του shell
APP_VERSION = '12.52'
APP_BUILD   = '332'

# ── v12.36 — Ιστορικό εκδόσεων («Τι νέο»). Newest first. ──────────────────────
CHANGELOG = [
    {'v': '12.52', 'b': '332', 'date': '14/06/2026', 'time': '18:20', 'title': 'Διόρθωση: Palm Island Suites (PLM)',
     'items': ['Σωστή αντιστοίχιση του ξενοδοχείου PLM = Palm Island Suites (πρόγραμμα & μισθοδοσία → εταιρεία ΦΥΤΩΡΙΑ ΚΡΗΤΗΣ).']},
    {'v': '12.51', 'b': '331', 'date': '14/06/2026', 'time': '18:05', 'title': 'Μισθοδοσία — Έλεγχος & Έγκριση (πίνακας Γεν. Διευθυντή)',
     'items': ['Νέα σελίδα «Έλεγχος & Έγκριση»: όλοι οι εργαζόμενοι/μήνα με εργάσιμες, ρεπό, έξτρα ώρες, Management (στο χέρι), Λογιστήριο (καθαρά), διαφορά, εκτός λογιστηρίου, πληρωτέο, υπόλοιπο + σημαίες κανόνων.',
               'Φίλτρα εταιρεία/ξενοδοχείο/τμήμα + σύνολα.',
               'Έγκριση μισθοδοσίας ανά εταιρεία: κλείδωμα + ίχνος (ποιος/πότε) + ειδοποίηση λογιστηρίου (πλατφόρμα + email) + εξαγωγή πίνακα ανά ξενοδοχείο (Excel).']},
    {'v': '12.50', 'b': '330', 'date': '14/06/2026', 'time': '17:35', 'title': 'Μισθοδοσία — Ανθεκτική εισαγωγή Epsilon',
     'items': ['Διορθώθηκε άνοιγμα αρχείων Epsilon που είχαν ασυμβατότητα μορφοποίησης (δεν φόρτωναν).',
               'Σαφές μήνυμα για παλιά αρχεία .xls (να αποθηκευτούν ως .xlsx).']},
    {'v': '12.49', 'b': '329', 'date': '14/06/2026', 'time': '17:10', 'title': 'Μισθοδοσία — Δώρα & άδεια στις δύο όψεις (Φ2.1)',
     'items': ['Η εισαγωγή Epsilon πιάνει πλέον και τα Δώρα Πάσχα/Χριστουγέννων + επίδομα/αποζημίωση αδείας (ξεχωριστά ανά εργαζόμενο).',
               'Η όψη εκτέλεσης δείχνει «Δώρα/άδεια» σπαστά και τα αθροίζει σε σύνολο (στο χέρι & κόστος εργοδότη) — σε Management και Λογιστήριο.',
               'Νέα σύνολα: Στο χέρι (σύνολο) και Κόστος εργοδότη (σύνολο).']},
    {'v': '12.48', 'b': '328', 'date': '14/06/2026', 'time': '16:30', 'title': 'Μισθοδοσία — Μηχανή & δύο όψεις (Φάση 2)',
     'items': ['Νέα «Εκτελέσεις»: υπολογισμός μισθοδοσίας ανά εταιρεία/μήνα από το Πρόγραμμα Εργασίας.',
               'Δύο όψεις: «Management» (τι θα πληρωθεί ο εργαζόμενος: μικτά συμφωνίας + έξτρα) και «Λογιστήριο» (νόμιμα καθαρά/εισφορές/κόστος).',
               'Εισαγωγή Epsilon (.xlsx): φορτώνει τα πραγματικά νόμιμα καθαρά και συμπληρώνει αυτόματα ΑΦΜ/ΑΜΚΑ/ειδικότητα/τράπεζα.',
               'Μενού «Οικονομικά» → «Μισθοδοσία» (Μητρώο/Εκτελέσεις).']},
    {'v': '12.47', 'b': '327', 'date': '14/06/2026', 'time': '15:45', 'title': 'Μισθοδοσία — Θεμέλιο (Φάση 1)',
     'items': ['Νέα ομάδα μενού «Οικονομικά» με σελίδα «Μισθοδοσία» (μόνο διαχειριστές).',
               'Μητρώο εργαζομένων ανά εταιρεία + καρτέλα με ευαίσθητα στοιχεία (ΑΦΜ/ΑΜΚΑ/ΙΒΑΝ) και συμφωνία.',
               'Νομικές οντότητες (εργοδότες) με αυτόματη αντιστοίχιση ξενοδοχείου→εταιρεία + συντελεστές 2026 (επεξεργάσιμοι).',
               'Η μηχανή υπολογισμού, τα εκκαθαριστικά και η πρόβλεψη έρχονται στη Φάση 2.']},
    {'v': '12.46', 'b': '326', 'date': '14/06/2026', 'time': '14:45', 'title': 'Βοήθεια: FAQ & Επίλυση προβλημάτων',
     'items': ['Νέα σελίδα «Βοήθεια (FAQ)» στο μενού Ενημέρωση — συχνές ερωτήσεις + επίλυση προβλημάτων με αναπτυσσόμενες απαντήσεις.',
               'Οι οδηγίες διαχείρισης εμφανίζονται μόνο σε διαχειριστές.']},
    {'v': '12.45', 'b': '325', 'date': '14/06/2026', 'time': '14:25', 'title': 'Roadmap · Multi-file importer ιστορικών',
     'items': ['Νέα σελίδα «Roadmap»: όλη η πορεία της Εστίας ανά τομέα (ολοκληρωμένα/σε εξέλιξη/σχεδιασμένα).',
               'Εισαγωγή πολλών αρχείων μαζί, με αυτόματη αναγνώριση: πρόγραμμα εργασίας ή μητρώο μισθοδοσίας.',
               'Νέα πηγή «Μητρώο Εργαζομένων» (καθαρή): τμήμα/εταιρεία/πρόσληψη/αποχώρηση + προφίλ· υποστήριξη παλιών ετών.']},
    {'v': '12.44', 'b': '324', 'date': '14/06/2026', 'time': '13:40', 'title': 'Τι νέο με ώρα · καθαρή διαχείριση εισαγμένου προσωπικού',
     'items': ['Το «Τι νέο» δείχνει πλέον ώρα μαζί με την ημερομηνία.',
               'Το εισαγμένο προσωπικό (χωρίς σύνδεση) δεν εμφανίζεται στη σελίδα «Χρήστες» — μένει καθαρή για τους πραγματικούς λογαριασμούς.',
               'Νέο κουμπί «Καθαρισμός εισαγμένου προσωπικού» για ασφαλή επανεκκαθάριση/επανεισαγωγή.']},
    {'v': '12.43', 'b': '323', 'date': '14/06/2026', 'time': '13:23', 'title': 'Πρόσβαση ανά ρόλο · Feedback · δυναμικά φίλτρα τμημάτων',
     'items': ['Νέο «Μενού ανά ρόλο» (Admin): ρυθμίζεις τι βλέπει Manager/Staff στο μενού.',
               'Οι managers (υποδοχή) βλέπουν πλέον Πρόγραμμα Εργασίας/Βλάβες/Records/Τι νέο.',
               'Νέο «Στείλε feedback» για όλους + σελίδα Feedback χρηστών (Admin).',
               'Πρόγραμμα Εργασίας: επιλογή πολλών τμημάτων με chips, δυναμικά χωρίς refresh (+ στήλη Εργάσιμες).',
               'Το «Το προφίλ μου» αφαιρέθηκε από το πλαϊνό μενού (υπάρχει στο avatar πάνω δεξιά).']},
    {'v': '12.42', 'b': '322', 'date': '14/06/2026', 'time': '12:55', 'title': 'Πρόγραμμα Εργασίας — εκκαθάριση διπλών + πρόσβαση ρόλων',
     'items': ['Εργαλείο εντοπισμού & συγχώνευσης διπλοεγγραφών προσωπικού (π.χ. ίδιο άτομο σε 2 ξενοδοχεία).',
               'Οι «Υποβολές (Λογιστήριο)» κρύβονται από μη-διαχειριστές.',
               'Οι managers βλέπουν Πρόγραμμα/Βλάβες/Logbook χωρίς πρόσβαση σε admin/εισαγωγές.']},
    {'v': '12.41', 'b': '321', 'date': '14/06/2026', 'time': '12:24', 'title': 'Πρόγραμμα Εργασίας — multi-week, εργάσιμες, διαχείριση προσωπικού',
     'items': ['Προβολή πολλών εβδομάδων μαζί (επιλογή 1–12 εβδ. μπροστά).',
               'Νέα στήλη «Εργάσιμες» (πλήθος ημερών εργασίας) ανά εργαζόμενο/εβδομάδα.',
               'Νέα σελίδα «Πρόγραμμα · Προσωπικό»: προσθήκη/επεξεργασία εργαζομένων (τμήμα/ξενοδοχείο/εταιρεία/σύνδεση).',
               'Ο κωδικός PLM δεν αντιστοιχίζεται πλέον σε ξενοδοχείο.']},
    {'v': '12.40', 'b': '320', 'date': '14/06/2026', 'time': '12:08', 'title': 'Module Πρόγραμμα Εργασίας (Βάρδιες) — Φάση 1',
     'items': ['Νέα ομάδα μενού «Προσωπικό»: Πρόγραμμα Εργασίας + Υποβολές (Λογιστήριο).',
               'Κάθε εργαζόμενος = χρήστης σε τμήμα (οργανόγραμμα)· import workbooks (χρήστες/τμήματα/βάρδιες).',
               'Φόρμα εβδομάδας Δ–Κ με κωδικούς βαρδιών, αυτόματες ώρες/ρεπό.',
               'Κανόνες (≥1 ρεπό/εβδ.), κλειδώματα προθεσμίας (Πέμπτη), ενοποιημένη αποστολή στο λογιστήριο.',
               'Τροποποίηση με νέα έκδοση + σήμανση αλλαγών· export ΠΡΟΣ ΛΟΓΙΣΤΗΡΙΟ.']},
    {'v': '12.39', 'b': '319', 'date': '14/06/2026', 'title': 'Dashboard: πλαϊνό κουμπί widgets + διόρθωση drag & drop',
     'items': ['Η μπάρα widgets ανοίγει από κομψό πλαϊνό κουμπί στη δεξιά άκρη της οθόνης.',
               'Διορθώθηκε το drag & drop — σύρε ένα widget από τη μπάρα στο πλέγμα.',
               'Εναλλακτικά, κάνε απλό κλικ σε ένα widget για να προστεθεί.']},
    {'v': '12.38', 'b': '318', 'date': '14/06/2026', 'title': 'Dashboard: πλαϊνή μπάρα widgets με drag & drop',
     'items': ['Η προσθήκη widget γίνεται από κομψή πλαϊνή μπάρα δεξιά — σύρε & άφησε στο πλέγμα.',
               'Σύρε ένα widget στη ζώνη απόρριψης για να το αφαιρέσεις.',
               'Τα προκαθορισμένα widgets μένουν πάντα διαθέσιμα στη μπάρα.',
               'Θεμέλιο για «Pin to dashboard»: ενότητα «Καρφιτσωμένα» (κάρφιτσωμα reports έρχεται σύντομα).',
               'Διόρθωση: το κλικ δεν επιλέγει πια κείμενο στα tiles.',
               'Άνοιξε το Staff HUB (προς το παρόν δείχνει τις Καταγραφές).']},
    {'v': '12.37', 'b': '317', 'date': '14/06/2026', 'title': 'Νέο Dashboard: ελεύθερο πλέγμα με resizable widgets',
     'items': ['Η «Σύνοψη» μετονομάστηκε σε Dashboard.',
               'Τα widgets αλλάζουν μέγεθος & θέση με το ποντίκι (με ελάχιστα όρια ώστε να μένουν ευανάγνωστα).',
               'Κουμπί «+ Προσθήκη widget» με μπάρα κατηγοριών — διάλεξε & πρόσθεσε εύκολα.',
               'Διακριτικό glow περίγραμμα στα tiles (έμπνευση από την οθόνη σύνδεσης).',
               'Πλήρης υποστήριξη σκούρου θέματος σε όλα τα tiles.',
               'Διόρθωση: το μενού προφίλ εμφανίζει σωστά όλες τις επιλογές.']},
    {'v': '12.36', 'b': '316', 'date': '13/06/2026', 'title': 'Ειδοποιήσεις popup · «Τι νέο» · γυάλισμα Σύνοψης & παλετών',
     'items': ['Το καμπανάκι ανοίγει popup με λίστα ειδοποιήσεων — διαβάζονται & φεύγουν επιτόπου.',
               'Νέα επιλογή «Ειδοποιήσεις» στο μενού προφίλ.',
               'Αυτή η σελίδα «Τι νέο»: ιστορικό εκδόσεων με τι άλλαξε.',
               'Σύνοψη: ξαναδουλεμένη premium εμφάνιση (καθαρές κάρτες, accent, εικονίδια) — ίδια λειτουργία tiles.',
               '4 παλέτες: το accent οδηγεί πλέον κουμπιά/links/καταστάσεις, βελτιωμένη αντίθεση.']},
    {'v': '12.34', 'b': '315', 'date': '13/06/2026', 'title': 'Σύνοψη = προσωπικό dashboard με tiles',
     'items': ['Η Σύνοψη έγινε προσωπικός πίνακας ανά χρήστη: διάλεξε & τακτοποίησε KPI/widget tiles (drag&drop).',
               'Νέα tiles: Δικές μου βλάβες, Μετρήσεις σήμερα, Απαντήσεις ερωτηματολογίων, Σημεία ελέγχου, Πρόσφατη δραστηριότητα.',
               'Σταθεροποίηση migration βάσης (race-safe).']},
    {'v': '12.33', 'b': '313', 'date': '13/06/2026', 'title': 'Αντίγραφα ασφαλείας: ρυθμίσεις & χρονοδιάγραμμα',
     'items': ['Ρυθμίσεις αυτόματου backup από το UI (ώρες/διατήρηση) χωρίς redeploy.',
               'Ημερήσιο backup 2 φορές/μέρα προς SharePoint, με όνομα ανά έκδοση.']},
    {'v': '12.31', 'b': '311', 'date': '13/06/2026', 'title': 'Module Αντιγράφων Ασφαλείας → SharePoint',
     'items': ['Πλήρη αντίγραφα δεδομένων (μετρήσεις/βλάβες/ερωτηματολόγια + εικόνες) στο εταιρικό SharePoint.',
               'Χειροκίνητο «Backup τώρα», τοπικό κατέβασμα, επαναφορά.']},
    {'v': '12.29', 'b': '309', 'date': '13/06/2026', 'title': 'Κέντρο Εισαγωγής Δεδομένων',
     'items': ['Ανέβασμα Excel/CSV με αυτόματη αντιστοίχιση στηλών & προεπισκόπηση.',
               'Εισαγωγή ιστορικών βλαβών & ερωτηματολογίων.']},
    {'v': '12.23', 'b': '—', 'date': '13/06/2026', 'title': 'Ερωτηματολόγια (Υποδοχή)',
     'items': ['Builder ερωτηματολογίων, δημόσιος σύνδεσμος για πελάτες (QR-ready), αποτελέσματα/NPS.']},
    {'v': '12.14', 'b': '—', 'date': '13/06/2026', 'title': 'Module Βλαβοληψία',
     'items': ['Δήλωση βλάβης με κατηγορίες/φωτογραφίες, Πίνακας Βλαβών (KPIs/Kanban), αυτόματη ανάθεση, SLA.']},
]

@app.context_processor
def inject_version():
    return {'app_version': APP_VERSION, 'app_build': APP_BUILD}

# v12.12 — Dark mode παντού: αυτόματη ένεση του theme snippet σε ΚΑΘΕ HTML σελίδα
# (εξαιρούνται login/register που έχουν δικό τους σκούρο design, και όσες το έχουν ήδη).
_DARK_SNIPPET = ('<script>(function(){try{var c=document.cookie;'
                 'var m=(c.match(/(?:^|; )estia_theme=([^;]+)/)||[])[1]||"system";'
                 'var d=m==="dark"||(m==="system"&&window.matchMedia&&matchMedia("(prefers-color-scheme:dark)").matches);'
                 'if(d)document.documentElement.setAttribute("data-theme","dark");'
                 'var p=(c.match(/(?:^|; )estia_palette=([^;]+)/)||[])[1]||"aurora";'
                 'document.documentElement.setAttribute("data-palette",p);}catch(e){}})();</script>'
                 '<link rel="stylesheet" href="/static/estia-theme.css">'
                 '<link rel="stylesheet" href="/static/estia-palette.css">')

@app.after_request
def _inject_dark(resp):
    try:
        if (resp.content_type or '').startswith('text/html'):
            body = resp.get_data(as_text=True)
            if '</head>' in body and '/static/estia-theme.css' not in body and '/static/auth.css' not in body:
                resp.set_data(body.replace('</head>', _DARK_SNIPPET + '</head>', 1))
    except Exception:
        pass
    return resp

def has_rank(min_rank):
    u = current_user()
    return u is not None and role_rank(u.role) >= min_rank

def is_admin():
    return has_rank(ROLE_RANK['admin'])      # admin ή masteradmin

def can_log():
    return has_rank(ROLE_RANK['staff'])      # ό,τι πάνω από viewer

def allowed_hotels(u):
    if u is None:
        return []
    if role_rank(u.role) >= ROLE_RANK['admin']:
        return Hotel.query.filter_by(is_active=True).order_by(Hotel.name).all()
    assigned = [h for h in u.hotels if h.is_active]
    if assigned:
        return sorted(assigned, key=lambda h: h.name)
    return Hotel.query.filter_by(is_active=True).order_by(Hotel.name).all()   # χωρίς ανάθεση = όλα

# ── v12.19 — Global εναλλάκτης ξενοδοχείου (session scope) ────────────────────
def active_hotel_id():
    """Το ενεργό ξενοδοχείο από τον global εναλλάκτη (None = όλα)."""
    v = session.get('active_hotel_id')
    try:
        return int(v) if v not in (None, '', 'all') else None
    except Exception:
        return None

def scoped_hotels(u):
    """allowed_hotels περιορισμένα στο ενεργό ξενοδοχείο (αν έχει επιλεγεί & επιτρέπεται)."""
    hs = allowed_hotels(u)
    aid = active_hotel_id()
    if aid:
        sub = [h for h in hs if h.id == aid]
        if sub:
            return sub
    return hs

def scoped_hotel_ids(u):
    return {h.id for h in scoped_hotels(u)}

@app.route('/set-hotel/<hid>')
def set_hotel(hid):
    if 'user_id' not in session:
        return ('', 401)
    session['active_hotel_id'] = '' if hid in ('all', '0', '') else hid
    return ('', 204)

@app.context_processor
def inject_hotelscope():
    u = current_user()
    return {'nav_hotels': (allowed_hotels(u) if u else []), 'active_hid': active_hotel_id()}

# ── v12.28 — Global φίλτρο περιόδου (session scope) ───────────────────────────
PERIOD_LABELS = [
    ('all', 'Όλα'), ('today', 'Σήμερα'), ('yesterday', 'Χθες'),
    ('l7', 'Τελευταίες 7 ημέρες'), ('week', 'Τρέχουσα εβδομάδα'),
    ('month', 'Τρέχων μήνας'), ('l30', 'Τελευταίες 30 ημέρες'),
    ('ytd', 'Από αρχή έτους (YTD)'), ('year', 'Τρέχον έτος'),
    ('ly', 'Πέρσι (LY)'), ('custom', 'Προσαρμογή…'),
]

def _period_bounds(key, today=None):
    t = today or datetime.utcnow().date()
    if key == 'today': return t, t
    if key == 'yesterday': d = t - timedelta(days=1); return d, d
    if key == 'l7': return t - timedelta(days=6), t
    if key == 'week': sft = t - timedelta(days=t.weekday()); return sft, sft + timedelta(days=6)
    if key == 'month':
        sft = t.replace(day=1)
        nm = (sft.replace(day=28) + timedelta(days=4)).replace(day=1)
        return sft, nm - timedelta(days=1)
    if key == 'l30': return t - timedelta(days=29), t
    if key == 'ytd': return t.replace(month=1, day=1), t
    if key == 'year': return t.replace(month=1, day=1), t.replace(month=12, day=31)
    if key == 'ly': return date(t.year - 1, 1, 1), date(t.year - 1, 12, 31)
    return None, None

def active_period():
    key = session.get('active_period', 'all')
    label = dict(PERIOD_LABELS).get(key, 'Όλα')
    start = end = None
    if key == 'custom':
        try: start = datetime.strptime(session.get('period_from', ''), '%Y-%m-%d').date()
        except Exception: start = None
        try: end = datetime.strptime(session.get('period_to', ''), '%Y-%m-%d').date()
        except Exception: end = None
        label = ('%s → %s' % (start.strftime('%d/%m/%Y'), end.strftime('%d/%m/%Y'))) if (start and end) else 'Προσαρμογή'
    elif key != 'all':
        start, end = _period_bounds(key)
    return {'key': key, 'label': label, 'start': start, 'end': end,
            'from': session.get('period_from', '') if key == 'custom' else '',
            'to':   session.get('period_to', '')   if key == 'custom' else ''}

def apply_period(q, col, dt=False):
    """Φιλτράρει query με την ενεργή περίοδο. col=date column (ή datetime αν dt=True). 'Όλα' = no-op."""
    p = active_period()
    sft, e = p['start'], p['end']
    if sft is None and e is None:
        return q
    if dt:
        if sft: q = q.filter(col >= datetime(sft.year, sft.month, sft.day))
        if e:   q = q.filter(col < datetime(e.year, e.month, e.day) + timedelta(days=1))
    else:
        if sft: q = q.filter(col >= sft)
        if e:   q = q.filter(col <= e)
    return q

@app.route('/set-period/<key>')
def set_period(key):
    if 'user_id' not in session:
        return ('', 401)
    keys = [k for k, _ in PERIOD_LABELS]
    session['active_period'] = key if key in keys else 'all'
    if key == 'custom':
        session['period_from'] = (request.args.get('from') or '').strip()
        session['period_to'] = (request.args.get('to') or '').strip()
    return ('', 204)

@app.context_processor
def inject_period():
    return {'active_period': active_period(), 'PERIOD_LABELS': PERIOD_LABELS}

def can_access_pool(u, pool):
    if u is None or pool is None:
        return False
    if role_rank(u.role) >= ROLE_RANK['admin']:
        return True
    assigned = {h.id for h in u.hotels if h.is_active}
    if not assigned:
        return True   # χωρίς ανάθεση = όλα
    return pool.hotel_id in assigned

def can_access_water_system(u, ws):
    """v12.3 — scoping δικτύου νερού (mirror can_access_pool)."""
    if u is None or ws is None:
        return False
    if role_rank(u.role) >= ROLE_RANK['admin']:
        return True
    assigned = {h.id for h in u.hotels if h.is_active}
    if not assigned:
        return True   # χωρίς ανάθεση = όλα
    return ws.hotel_id in assigned

def log_activity(action, detail='', hotel_id=None):
    try:
        if hotel_id is None:
            try: hotel_id = active_hotel_id()
            except Exception: hotel_id = None
        db.session.add(ActivityLog(user_id=session.get('user_id'), hotel_id=hotel_id,
                                   action=(action or '')[:60], detail=(detail or '')[:300]))
        db.session.commit()
    except Exception:
        db.session.rollback()

# ──────────────────────────────────────────────────────────────────────────
#  MODELS
# ──────────────────────────────────────────────────────────────────────────
user_hotels = db.Table('user_hotels',
    db.Column('user_id', db.Integer, db.ForeignKey('user.id'), primary_key=True),
    db.Column('hotel_id', db.Integer, db.ForeignKey('hotel.id'), primary_key=True),
)

class ActivityLog(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    user_id    = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    hotel_id   = db.Column(db.Integer, db.ForeignKey('hotel.id'), nullable=True)   # v12.22
    action     = db.Column(db.String(60))
    detail     = db.Column(db.String(300))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    user       = db.relationship('User')
    hotel      = db.relationship('Hotel')

class User(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    username   = db.Column(db.String(50), unique=True, nullable=False)
    password   = db.Column(db.String(200), nullable=False)
    full_name  = db.Column(db.String(100), nullable=False)
    role       = db.Column(db.String(20), default='staff')
    language   = db.Column(db.String(5), default='el')
    email      = db.Column(db.String(120))
    phone      = db.Column(db.String(40))
    approved   = db.Column(db.Boolean, default=True)
    avatar     = db.Column(db.Text)                       # data URL (base64) εικόνας προφίλ
    dashboard  = db.Column(db.Text)                       # v12.34 — προσωπική διάταξη tiles (JSON)
    # v12.40 — Module Πρόγραμμα Εργασίας (οργανόγραμμα + βάρδιες)
    department_id     = db.Column(db.Integer)
    employer          = db.Column(db.String(120))
    subunit           = db.Column(db.String(20))
    home_hotel_id     = db.Column(db.Integer)
    login_enabled     = db.Column(db.Boolean)
    employment_active = db.Column(db.Boolean)
    is_active  = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    hotels     = db.relationship('Hotel', secondary=user_hotels, backref='members')

class WaterRecord(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    user_id     = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    record_date = db.Column(db.Date, default=date.today, nullable=False)
    period      = db.Column(db.String(10), nullable=False)
    recorded_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at  = db.Column(db.DateTime, nullable=True)
    updated_by  = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    clo2_tank        = db.Column(db.Float)
    clo2_kitchen     = db.Column(db.Float)
    clo2_remote      = db.Column(db.Float)
    clo2_dhw_out     = db.Column(db.Float)
    clo2_dhw_return  = db.Column(db.Float)
    clo2_ro          = db.Column(db.Float)
    location_kitchen = db.Column(db.String(100))
    location_remote  = db.Column(db.String(100))
    temp_tank         = db.Column(db.Float)
    temp_dhw_out      = db.Column(db.Float)
    temp_dhw_return   = db.Column(db.Float)
    temp_ro           = db.Column(db.Float)
    temp_kitchen_cold = db.Column(db.Float)
    temp_kitchen_hot  = db.Column(db.Float)
    temp_remote_cold  = db.Column(db.Float)
    temp_remote_hot   = db.Column(db.Float)
    ph_tank = db.Column(db.Float)
    notes   = db.Column(db.Text)
    water_system_id = db.Column(db.Integer, db.ForeignKey('water_system.id'), nullable=True)  # v12.3
    user         = db.relationship('User', foreign_keys=[user_id], backref='water_records')
    updated_user = db.relationship('User', foreign_keys=[updated_by])
    water_system = db.relationship('WaterSystem')

class Hotel(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(120), unique=True, nullable=False)
    is_active  = db.Column(db.Boolean, default=True)
    company_id = db.Column(db.Integer)   # v12.47 — Μισθοδοσία: νομικός εργοδότης (payroll.Company)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    pools = db.relationship('Pool', backref='hotel', order_by='Pool.name')

class Pool(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    hotel_id   = db.Column(db.Integer, db.ForeignKey('hotel.id'), nullable=False)
    name       = db.Column(db.String(120), nullable=False)
    location   = db.Column(db.String(120))
    pool_type  = db.Column(db.String(20), default='pool')
    volume_m3  = db.Column(db.Float)          # όγκος σε m³ (για δοσολογία)
    is_active  = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class PoolRecord(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    pool_id     = db.Column(db.Integer, db.ForeignKey('pool.id'), nullable=False)
    user_id     = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    record_date = db.Column(db.Date, default=date.today, nullable=False)
    period      = db.Column(db.String(10), nullable=False)
    recorded_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at  = db.Column(db.DateTime, nullable=True)
    updated_by  = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    free_chlorine     = db.Column(db.Float)
    combined_chlorine = db.Column(db.Float)
    ph                = db.Column(db.Float)
    temp              = db.Column(db.Float)
    turbidity         = db.Column(db.Float)
    cyanuric_acid     = db.Column(db.Float)
    total_alkalinity  = db.Column(db.Float)
    orp               = db.Column(db.Float)
    backwash_done     = db.Column(db.Boolean, default=False)
    notes             = db.Column(db.Text)
    pool         = db.relationship('Pool')
    user         = db.relationship('User', foreign_keys=[user_id])
    updated_user = db.relationship('User', foreign_keys=[updated_by])


# v12.3 — Δίκτυο Νερού ανά ξενοδοχείο (mirror Pool). Ένα ξενοδοχείο μπορεί να έχει >1.
class WaterSystem(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    hotel_id   = db.Column(db.Integer, db.ForeignKey('hotel.id'), nullable=False)
    name       = db.Column(db.String(120), nullable=False)
    location   = db.Column(db.String(120))
    is_active  = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    hotel      = db.relationship('Hotel')


class ReminderSent(db.Model):
    day = db.Column(db.String(10), primary_key=True)        # 'YYYY-MM-DD' lock ανά ημέρα
    sent_at = db.Column(db.DateTime, default=datetime.utcnow)

class FaultReport(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    user_id     = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    hotel_id    = db.Column(db.Integer, db.ForeignKey('hotel.id'), nullable=True)
    pool_id     = db.Column(db.Integer, db.ForeignKey('pool.id'), nullable=True)
    area        = db.Column(db.String(20))
    message     = db.Column(db.Text)
    status      = db.Column(db.String(12), default='open')   # open / resolved
    answer      = db.Column(db.Text)
    answered_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)
    resolved_at = db.Column(db.DateTime, nullable=True)
    user = db.relationship('User', foreign_keys=[user_id])
    pool = db.relationship('Pool')

class Notification(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    user_id    = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    text       = db.Column(db.String(300))
    link       = db.Column(db.String(120))
    is_read    = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# v12.38 — Καρφιτσωμένα widgets (pin to dashboard). Νέος πίνακας· create_all το δημιουργεί.
class PinnedWidget(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    user_id    = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    title      = db.Column(db.String(80), nullable=False)
    icon       = db.Column(db.String(40))
    link       = db.Column(db.String(300))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# v12.9 — Μηνύματα μεταξύ χρηστών (νέος πίνακας· create_all το δημιουργεί αυτόματα)
class Message(db.Model):
    id           = db.Column(db.Integer, primary_key=True)
    sender_id    = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    recipient_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    body         = db.Column(db.Text, nullable=False)
    is_read      = db.Column(db.Boolean, default=False)
    created_at   = db.Column(db.DateTime, default=datetime.utcnow)
    sender    = db.relationship('User', foreign_keys=[sender_id])
    recipient = db.relationship('User', foreign_keys=[recipient_id])

class EmailLog(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    subject    = db.Column(db.String(200))
    recipients = db.Column(db.String(300))
    ok         = db.Column(db.Boolean, default=False)
    via        = db.Column(db.String(10))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Setting(db.Model):
    key   = db.Column(db.String(40), primary_key=True)
    value = db.Column(db.Text)

# ── Generic monitoring engine (Πλατφόρμα Συντήρησης) ──
class MonitorTemplate(db.Model):
    key       = db.Column(db.String(30), primary_key=True)   # 'tank','energy',...
    name      = db.Column(db.String(80))
    icon      = db.Column(db.String(30), default='ti-checklist')
    frequency = db.Column(db.String(10), default='daily')    # twice|daily|weekly|monthly
    sort      = db.Column(db.Integer, default=0)
    is_active = db.Column(db.Boolean, default=True)
    params    = db.relationship('MonitorParam', backref='template', order_by='MonitorParam.sort',
                                cascade='all, delete-orphan')

class MonitorParam(db.Model):
    id           = db.Column(db.Integer, primary_key=True)
    template_key = db.Column(db.String(30), db.ForeignKey('monitor_template.key'))
    pkey         = db.Column(db.String(40))      # κλειδί τιμής
    label        = db.Column(db.String(80))
    unit         = db.Column(db.String(20), default='')
    min_v        = db.Column(db.Float)
    max_v        = db.Column(db.Float)
    action_low   = db.Column(db.String(200))
    action_high  = db.Column(db.String(200))
    periodic     = db.Column(db.Boolean, default=False)   # μόνο πρωί/πρώτη βάρδια
    sort         = db.Column(db.Integer, default=0)

class Area(db.Model):
    id           = db.Column(db.Integer, primary_key=True)
    hotel_id     = db.Column(db.Integer, db.ForeignKey('hotel.id'), nullable=False)
    template_key = db.Column(db.String(30), db.ForeignKey('monitor_template.key'), nullable=False)
    name         = db.Column(db.String(120))
    location     = db.Column(db.String(120))
    is_active    = db.Column(db.Boolean, default=True)
    created_at   = db.Column(db.DateTime, default=datetime.utcnow)
    hotel        = db.relationship('Hotel')
    template     = db.relationship('MonitorTemplate')

class Reading(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    area_id     = db.Column(db.Integer, db.ForeignKey('area.id'), nullable=False)
    template_key= db.Column(db.String(30))
    user_id     = db.Column(db.Integer, db.ForeignKey('user.id'))
    record_date = db.Column(db.Date, default=date.today)
    period      = db.Column(db.String(10), default='day')
    recorded_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at  = db.Column(db.DateTime)
    updated_by  = db.Column(db.Integer, db.ForeignKey('user.id'))
    values      = db.Column(db.Text)     # JSON {pkey: value}
    notes       = db.Column(db.Text)
    area        = db.relationship('Area')
    user        = db.relationship('User', foreign_keys=[user_id])

FREQ_PERIODS = {'twice': ['morning', 'afternoon'], 'daily': ['day'], 'weekly': ['day'], 'monthly': ['day']}
FREQ_LABEL   = {'twice': '2×/ημέρα (πρωί/απόγευμα)', 'daily': 'Καθημερινά', 'weekly': 'Εβδομαδιαία', 'monthly': 'Μηνιαία'}

def periods_for(freq):
    return FREQ_PERIODS.get(freq, ['day'])

def area_actions(reading):
    """Generic έλεγχος ορίων -> ενέργειες, με βάση τις παραμέτρους του template."""
    out = []
    try:
        vals = json.loads(reading.values or '{}')
    except Exception:
        vals = {}
    tpl = MonitorTemplate.query.get(reading.template_key)
    if not tpl:
        return out
    for p in tpl.params:
        v = vals.get(p.pkey)
        if v is None:
            continue
        try:
            v = float(v)
        except (ValueError, TypeError):
            continue
        act = None
        if p.min_v is not None and v < p.min_v:
            act = p.action_low or ('Χαμηλό: ' + p.label)
        elif p.max_v is not None and v > p.max_v:
            act = p.action_high or ('Υψηλό: ' + p.label)
        if act:
            out.append({'label': p.label, 'value': v, 'unit': p.unit, 'action': act})
    return out


def flt(data, key):
    try:
        return float(data[key]) if data.get(key) not in (None, '') else None
    except (ValueError, TypeError):
        return None


# ──────────────────────────────────────────────────────────────────────────
#  EMAIL (Microsoft Graph ή SMTP) + log
# ──────────────────────────────────────────────────────────────────────────
def _graph_token():
    data = urllib.parse.urlencode({
        'client_id': GRAPH_CLIENT_ID, 'client_secret': GRAPH_CLIENT_SECRET,
        'scope': 'https://graph.microsoft.com/.default', 'grant_type': 'client_credentials',
    }).encode()
    req = urllib.request.Request('https://login.microsoftonline.com/' + GRAPH_TENANT_ID + '/oauth2/v2.0/token', data=data)
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.loads(r.read().decode())['access_token']

def _send_graph(subject, html, recips):
    token = _graph_token()
    payload = {'message': {'subject': subject, 'body': {'contentType': 'HTML', 'content': html},
               'toRecipients': [{'emailAddress': {'address': a}} for a in recips]}, 'saveToSentItems': True}
    req = urllib.request.Request('https://graph.microsoft.com/v1.0/users/' + GRAPH_SENDER + '/sendMail',
        data=json.dumps(payload).encode(),
        headers={'Authorization': 'Bearer ' + token, 'Content-Type': 'application/json'})
    with urllib.request.urlopen(req, timeout=30) as r:
        return r.status in (200, 202)

def _send_smtp(subject, html, recips):
    msg = MIMEMultipart(); msg['From'] = EMAIL_FROM; msg['To'] = ', '.join(recips); msg['Subject'] = subject
    msg.attach(MIMEText(html, 'html', 'utf-8'))
    sv = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT); sv.login(EMAIL_FROM, EMAIL_PASSWORD)
    sv.sendmail(EMAIL_FROM, recips, msg.as_string()); sv.quit()
    return True

def send_email(subject, html, recips=None):
    recips = recips or EMAIL_TO_LIST
    via = 'graph' if GRAPH_CLIENT_ID else ('smtp' if EMAIL_PASSWORD else 'none')
    ok = False
    if via != 'none':
        try:
            ok = _send_graph(subject, html, recips) if via == 'graph' else _send_smtp(subject, html, recips)
        except Exception as e:
            print('email error (' + via + '):', e); ok = False
    try:
        db.session.add(EmailLog(subject=(subject or '')[:200], recipients=', '.join(recips)[:300], ok=ok, via=via))
        db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
    return ok


THEME_DEFAULTS = {'primary': '#193847', 'accent': '#BB9549', 'app_title': 'Εστία', 'logo': '',
                  # v12.11 — εξατομίκευση σελίδας login/register
                  'login_logo': '', 'login_logo_size': '88', 'login_font': 'Space Grotesk',
                  'login_tagline': 'we reinvent the way we work',
                  'social_facebook': '', 'social_linkedin': ''}
LOGIN_FONTS = ['Space Grotesk', 'Poppins', 'Montserrat', 'Inter', 'Sora', 'System']

def get_theme():
    t = dict(THEME_DEFAULTS)
    try:
        for row in Setting.query.filter(Setting.key.like('theme_%')).all():
            k = row.key[6:]
            if k in t and row.value is not None:
                t[k] = row.value
    except Exception:
        pass
    return t

def get_ai_config():
    cfg = {'provider': AI_PROVIDER, 'anthropic': ANTHROPIC_API_KEY, 'openai': OPENAI_API_KEY, 'model': AI_MODEL}
    try:
        for row in Setting.query.filter(Setting.key.like('ai_%')).all():
            k = row.key[3:]
            if k in cfg and row.value:
                cfg[k] = row.value
    except Exception:
        pass
    return cfg


# ──────────────────────────────────────────────────────────────────────────
#  WATER LOG
# ──────────────────────────────────────────────────────────────────────────
def apply_record(record, data, period):
    record.clo2_tank        = flt(data, 'clo2_tank')
    record.clo2_kitchen     = flt(data, 'clo2_kitchen')
    record.clo2_remote      = flt(data, 'clo2_remote')
    record.location_kitchen = data.get('location_kitchen', '')
    record.location_remote  = data.get('location_remote', '')
    record.temp_tank        = flt(data, 'temp_tank')
    record.temp_dhw_out     = flt(data, 'temp_dhw_out')
    record.temp_dhw_return  = flt(data, 'temp_dhw_return')
    record.temp_ro          = flt(data, 'temp_ro')
    record.temp_kitchen_cold = flt(data, 'temp_kitchen_cold')
    record.temp_kitchen_hot  = flt(data, 'temp_kitchen_hot')
    record.temp_remote_cold  = flt(data, 'temp_remote_cold')
    record.temp_remote_hot   = flt(data, 'temp_remote_hot')
    record.notes = data.get('notes', '')
    if period == 'morning':
        record.clo2_dhw_out    = flt(data, 'clo2_dhw_out')
        record.clo2_dhw_return = flt(data, 'clo2_dhw_return')
        record.clo2_ro         = flt(data, 'clo2_ro')
        record.ph_tank         = flt(data, 'ph_tank')

def send_report_email(record, user):
    if not EMAIL_PASSWORD and not GRAPH_CLIENT_ID:
        return False
    period_gr = 'Πρωι' if record.period == 'morning' else 'Απογευμα'
    # v12.3 — δυναμικό ξενοδοχείο + δίκτυο (αντί hardcoded «Sergios»)
    _ws = record.water_system
    place = ''
    if _ws:
        place = (_ws.hotel.name + ' · ' + _ws.name) if _ws.hotel else _ws.name
    place_sfx = f' ({place})' if place else ''
    def row(label, val, unit='', min_v=None, max_v=None):
        if val is None:
            return f'<tr><td style="padding:7px 8px;border:1px solid #eee;">{label}</td><td style="padding:7px 8px;border:1px solid #eee;">-</td><td style="padding:7px 8px;border:1px solid #eee;color:#888;">{unit}</td></tr>'
        ok = (min_v is None or val >= min_v) and (max_v is None or val <= max_v)
        color = '#16a34a' if ok else '#dc2626'
        icon  = 'OK' if ok else 'ΠΡΟΣΟΧΗ'
        limit = f'min {min_v}{unit}' if min_v else (f'max {max_v}{unit}' if max_v else '')
        return f'<tr><td style="padding:7px 8px;border:1px solid #eee;">{label}</td><td style="padding:7px 8px;border:1px solid #eee;color:{color};font-weight:500;">{icon}: {val} {unit}</td><td style="padding:7px 8px;border:1px solid #eee;color:#888;">{limit}</td></tr>'
    loc_kit = f' ({record.location_kitchen})' if record.location_kitchen else ''
    loc_rem = f' ({record.location_remote})' if record.location_remote else ''
    clo2_rows  = row('Δεξαμενη', record.clo2_tank, 'ppm', 1.0, 2.0)
    clo2_rows += row(f'Κουζινα{loc_kit}', record.clo2_kitchen, 'ppm', 1.0, 2.0)
    clo2_rows += row(f'Απομακρυσμενο{loc_rem}', record.clo2_remote, 'ppm', 1.0, 2.0)
    if record.period == 'morning':
        clo2_rows += row('Αναχωρηση ΖΝΧ', record.clo2_dhw_out, 'ppm', 1.0, 2.0)
        clo2_rows += row('Επιστροφη ΖΝΧ', record.clo2_dhw_return, 'ppm', 1.0, 2.0)
        clo2_rows += row('Αντ. Οσμωση', record.clo2_ro, 'ppm', 1.0, 2.0)
    temp_rows  = row('Δεξαμενη', record.temp_tank, 'C', None, 20.0)
    temp_rows += row('Κολεκτερ ΖΝΧ (Αναχ.)', record.temp_dhw_out, 'C', 60.0, None)
    temp_rows += row('Κολεκτερ Ανακυκλ. (Επιστρ.)', record.temp_dhw_return, 'C', 50.0, None)
    temp_rows += row('Αντ. Οσμωση', record.temp_ro, 'C')
    temp_rows += row(f'Κουζινα Κρυο{loc_kit}', record.temp_kitchen_cold, 'C')
    temp_rows += row(f'Κουζινα Ζεστο{loc_kit}', record.temp_kitchen_hot, 'C', 50.0, None)
    temp_rows += row(f'Απομακρυσμενο Κρυο{loc_rem}', record.temp_remote_cold, 'C')
    temp_rows += row(f'Απομακρυσμενο Ζεστο{loc_rem}', record.temp_remote_hot, 'C', 50.0, None)
    _wa = compute_water_actions(record)
    if _wa:
        _li = ''.join('<li style="margin-bottom:4px;' + ('color:#b91c1c;font-weight:600;' if a['urgent'] else '') + '">' + ('ΕΠΕΙΓΟΝ — ' if a['urgent'] else '') + a['label'] + ': ' + a['action'] + '</li>' for a in _wa)
        actions_html = '<h2 style="font-size:15px;color:#b45309;margin-top:18px;">Προτεινομενες ενεργειες (ΖΝΧ / νερο)</h2><ul style="background:#fffbeb;border:1px solid #fde68a;border-radius:6px;padding:12px 12px 12px 28px;color:#333;">' + _li + '</ul>'
    else:
        actions_html = ''
    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:620px;margin:0 auto;">
      <div style="background:#0369a1;color:white;padding:20px;border-radius:8px 8px 0 0;">
        <h1 style="margin:0;font-size:20px;">Εστία — Νερά Χρήσης{place_sfx} {period_gr}</h1>
        <p style="margin:5px 0 0;opacity:0.8;">{record.record_date.strftime('%d/%m/%Y')} | Υπευθυνος: {user.full_name}</p>
      </div>
      <div style="background:#f9f9f9;padding:20px;border:1px solid #eee;">
        <h2 style="font-size:15px;color:#333;border-bottom:2px solid #0369a1;padding-bottom:6px;">CLO2 (ppm) - Στοχος: 1.0-2.0 ppm</h2>
        <table style="width:100%;border-collapse:collapse;">
          <tr style="background:#0369a1;color:white;"><th style="padding:8px;text-align:left;">Σημειο</th><th style="padding:8px;text-align:left;">Μετρηση</th><th style="padding:8px;text-align:left;">Ορια</th></tr>
          {clo2_rows}
        </table>
        <h2 style="font-size:15px;color:#333;border-bottom:2px solid #0369a1;padding-bottom:6px;margin-top:20px;">Θερμοκρασια (C)</h2>
        <table style="width:100%;border-collapse:collapse;">
          <tr style="background:#0369a1;color:white;"><th style="padding:8px;text-align:left;">Σημειο</th><th style="padding:8px;text-align:left;">Μετρηση</th><th style="padding:8px;text-align:left;">Ορια</th></tr>
          {temp_rows}
        </table>
        {actions_html}
        {f'<h2 style="font-size:15px;color:#333;margin-top:20px;">pH</h2><p style="background:#fff;padding:10px;border:1px solid #eee;">Δεξαμενη: {record.ph_tank}</p>' if record.period == 'morning' and record.ph_tank else ''}
        {f'<h2 style="font-size:15px;color:#333;margin-top:20px;">Παρατηρησεις</h2><p style="background:#fff;padding:10px;border:1px solid #eee;">{record.notes}</p>' if record.notes else ''}
      </div>
      <div style="background:#f0f0f0;padding:12px;text-align:center;font-size:12px;color:#888;border-radius:0 0 8px 8px;">
        Εστία — CONDIAN HOTELS{place_sfx} · {record.record_date.strftime('%d/%m/%Y')} · {period_gr}
      </div>
    </div>"""
    return send_email(f'Εστία — Νερά Χρήσης{place_sfx} {period_gr} {record.record_date.strftime("%d/%m/%Y")}', html, EMAIL_TO_LIST)


# ──────────────────────────────────────────────────────────────────────────
#  POOL LOG
# ──────────────────────────────────────────────────────────────────────────
def apply_pool_record(record, data, period):
    record.free_chlorine     = flt(data, 'free_chlorine')
    record.combined_chlorine = flt(data, 'combined_chlorine')
    record.ph                = flt(data, 'ph')
    record.temp              = flt(data, 'temp')
    record.turbidity         = flt(data, 'turbidity')
    record.backwash_done     = data.get('backwash_done') in ('1', 'on', 'true', 'True')
    record.notes             = data.get('notes', '')
    if period == 'morning':
        record.cyanuric_acid    = flt(data, 'cyanuric_acid')
        record.total_alkalinity = flt(data, 'total_alkalinity')
        record.orp              = flt(data, 'orp')

def send_pool_report_email(record, user):
    if not EMAIL_PASSWORD and not GRAPH_CLIENT_ID:
        return False
    period_gr = 'Πρωι' if record.period == 'morning' else 'Απογευμα'
    pool   = record.pool
    hotel  = pool.hotel.name if pool and pool.hotel else ''
    point  = f' — {pool.location}' if pool and pool.location else ''
    def row(label, val, unit, key):
        min_v, max_v = POOL_LIMITS.get(key, (None, None))
        if val is None:
            return f'<tr><td style="padding:7px 8px;border:1px solid #eee;">{label}</td><td style="padding:7px 8px;border:1px solid #eee;">-</td><td style="padding:7px 8px;border:1px solid #eee;color:#888;">{unit}</td></tr>'
        ok = (min_v is None or val >= min_v) and (max_v is None or val <= max_v)
        color = '#16a34a' if ok else '#dc2626'
        icon  = 'OK' if ok else 'ΠΡΟΣΟΧΗ'
        if min_v is not None and max_v is not None:
            limit = f'{min_v}-{max_v} {unit}'
        elif min_v is not None:
            limit = f'min {min_v} {unit}'
        elif max_v is not None:
            limit = f'max {max_v} {unit}'
        else:
            limit = unit
        return f'<tr><td style="padding:7px 8px;border:1px solid #eee;">{label}</td><td style="padding:7px 8px;border:1px solid #eee;color:{color};font-weight:500;">{icon}: {val} {unit}</td><td style="padding:7px 8px;border:1px solid #eee;color:#888;">{limit}</td></tr>'
    rows  = row('Ελευθερο χλωριο', record.free_chlorine, 'mg/L', 'free_chlorine')
    rows += row('Συνδεδεμενο χλωριο', record.combined_chlorine, 'mg/L', 'combined_chlorine')
    rows += row('pH', record.ph, '', 'ph')
    rows += row('Θερμοκρασια', record.temp, 'C', 'temp')
    rows += row('Θολοτητα', record.turbidity, 'NTU', 'turbidity')
    if record.period == 'morning':
        rows += row('Κυανουρικο οξυ', record.cyanuric_acid, 'mg/L', 'cyanuric_acid')
        rows += row('Ολικη αλκαλικοτητα', record.total_alkalinity, 'mg/L', 'total_alkalinity')
        rows += row('ORP (Redox)', record.orp, 'mV', 'orp')
    backwash = 'Ναι' if record.backwash_done else 'Οχι'
    _acts = compute_pool_actions(record)
    if _acts:
        _li = ''.join('<li style="margin-bottom:4px;' + ('color:#b91c1c;font-weight:600;' if a['urgent'] else '') + '">' + ('ΕΠΕΙΓΟΝ — ' if a['urgent'] else '') + a['label'] + ': ' + a['action'] + '</li>' for a in _acts)
        actions_html = '<h2 style="font-size:15px;color:#b45309;margin-top:18px;">Προτεινομενες ενεργειες</h2><ul style="background:#fffbeb;border:1px solid #fde68a;border-radius:6px;padding:12px 12px 12px 28px;color:#333;">' + _li + '</ul><p style="font-size:11px;color:#888;">' + SAFETY_NOTE + '</p>'
    else:
        actions_html = ''
    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:620px;margin:0 auto;">
      <div style="background:#0e7490;color:white;padding:20px;border-radius:8px 8px 0 0;">
        <h1 style="margin:0;font-size:20px;">{hotel} - Πισινα {period_gr}</h1>
        <p style="margin:5px 0 0;opacity:0.85;">{pool.name if pool else ''}{point}</p>
        <p style="margin:5px 0 0;opacity:0.8;">{record.record_date.strftime('%d/%m/%Y')} | Υπευθυνος: {user.full_name}</p>
      </div>
      <div style="background:#f9f9f9;padding:20px;border:1px solid #eee;">
        <table style="width:100%;border-collapse:collapse;">
          <tr style="background:#0e7490;color:white;"><th style="padding:8px;text-align:left;">Παραμετρος</th><th style="padding:8px;text-align:left;">Μετρηση</th><th style="padding:8px;text-align:left;">Ορια</th></tr>
          {rows}
        </table>
        <p style="margin-top:14px;font-size:13px;color:#555;">Ανταποπλυση φιλτρου (backwash): <b>{backwash}</b></p>
        {actions_html}
        {f'<h2 style="font-size:15px;color:#333;margin-top:16px;">Παρατηρησεις</h2><p style="background:#fff;padding:10px;border:1px solid #eee;">{record.notes}</p>' if record.notes else ''}
      </div>
      <div style="background:#f0f0f0;padding:12px;text-align:center;font-size:12px;color:#888;border-radius:0 0 8px 8px;">
        {hotel} - Πισινα {pool.name if pool else ''} - {record.record_date.strftime('%d/%m/%Y')} - {period_gr}
      </div>
    </div>"""
    return send_email(f'{hotel} - Πισινα {pool.name if pool else ""} {period_gr} {record.record_date.strftime("%d/%m/%Y")}', html, EMAIL_TO_LIST)


# ──────────────────────────────────────────────────────────────────────────
#  AI ASSISTANT helpers
# ──────────────────────────────────────────────────────────────────────────
def resolve_provider():
    c = get_ai_config()
    if c['provider'] == 'anthropic' and c['anthropic']:
        return 'anthropic'
    if c['provider'] == 'openai' and c['openai']:
        return 'openai'
    if c['provider'] == 'auto':
        if c['anthropic']:
            return 'anthropic'
        if c['openai']:
            return 'openai'
    return None

def call_llm(system_prompt, messages):
    """messages: list of {'role':'user'|'assistant','content':str}. Returns (reply, error)."""
    provider = resolve_provider()
    if not provider:
        return None, 'not_configured'
    c = get_ai_config()
    try:
        if provider == 'anthropic':
            model = c['model'] or 'claude-sonnet-4-6'
            payload = {'model': model, 'max_tokens': 1024,
                       'system': system_prompt, 'messages': messages}
            req = urllib.request.Request(
                'https://api.anthropic.com/v1/messages',
                data=json.dumps(payload).encode('utf-8'),
                headers={'content-type': 'application/json',
                         'x-api-key': c['anthropic'],
                         'anthropic-version': '2023-06-01'})
            with urllib.request.urlopen(req, timeout=60) as r:
                data = json.loads(r.read().decode('utf-8'))
            return data['content'][0]['text'], None
        else:  # openai
            model = c['model'] or 'gpt-4o-mini'
            msgs = [{'role': 'system', 'content': system_prompt}] + messages
            payload = {'model': model, 'max_tokens': 1024, 'messages': msgs}
            req = urllib.request.Request(
                'https://api.openai.com/v1/chat/completions',
                data=json.dumps(payload).encode('utf-8'),
                headers={'content-type': 'application/json',
                         'authorization': 'Bearer ' + c['openai']})
            with urllib.request.urlopen(req, timeout=60) as r:
                data = json.loads(r.read().decode('utf-8'))
            return data['choices'][0]['message']['content'], None
    except urllib.error.HTTPError as e:
        body = ''
        try:
            body = e.read().decode('utf-8')[:300]
        except Exception:
            pass
        return None, f'http_{e.code}: {body}'
    except Exception as e:
        return None, str(e)

def build_pool_context(pool):
    if not pool:
        return 'Δεν έχει επιλεγεί συγκεκριμένη πισίνα — ρώτα τον χρήστη ποια πισίνα αφορά.'
    lines = [
        f'Ξενοδοχείο: {pool.hotel.name if pool.hotel else "-"}',
        f'Πισίνα: {pool.name}' + (f' ({pool.location})' if pool.location else ''),
        f'Τύπος: {pool.pool_type}',
        f'Όγκος: {(str(pool.volume_m3) + " m3") if pool.volume_m3 else "ΑΓΝΩΣΤΟΣ — ζήτησε να τον επιβεβαιώσει ο χρήστης"}',
    ]
    recs = (PoolRecord.query.filter_by(pool_id=pool.id)
            .order_by(PoolRecord.record_date.desc(), PoolRecord.recorded_at.desc())
            .limit(2).all())
    if recs:
        lines.append('Τελευταίες μετρήσεις:')
        for r in recs:
            per = 'Πρωί' if r.period == 'morning' else 'Απόγευμα'
            parts = []
            def add(lbl, val, key, unit=''):
                if val is None:
                    return
                mn, mx = POOL_LIMITS.get(key, (None, None))
                ok = (mn is None or val >= mn) and (mx is None or val <= mx)
                parts.append(f'{lbl}={val}{unit}' + ('' if ok else ' [ΕΚΤΟΣ ΟΡΙΩΝ]'))
            add('ελ.χλώριο', r.free_chlorine, 'free_chlorine', ' mg/L')
            add('συνδ.χλώριο', r.combined_chlorine, 'combined_chlorine', ' mg/L')
            add('pH', r.ph, 'ph')
            add('θερμ', r.temp, 'temp', ' C')
            add('θολότητα', r.turbidity, 'turbidity', ' NTU')
            add('κυανουρικό', r.cyanuric_acid, 'cyanuric_acid', ' mg/L')
            add('αλκαλικότητα', r.total_alkalinity, 'total_alkalinity', ' mg/L')
            add('ORP', r.orp, 'orp', ' mV')
            lines.append(f'  {r.record_date.strftime("%d/%m")} {per}: ' + (', '.join(parts) if parts else '—'))
    else:
        lines.append('Δεν υπάρχουν καταγεγραμμένες μετρήσεις για αυτή την πισίνα ακόμα.')
    lines.append('Επιθυμητά όρια: ελ.χλώριο 0.4-1.5 mg/L, συνδ.χλώριο <=0.5, pH 7.2-7.8, θερμ <=32C, θολότητα <=1 NTU, κυανουρικό <=75, αλκαλικότητα 80-120, ORP >=650 mV.')
    return '\n'.join(lines)


def build_pool_report_pdf(rep_date, records):
    """Branded PDF αναφορά πισινών για μια ημέρα (fpdf2 + DejaVuSans)."""
    from fpdf import FPDF
    NAVY=(25,56,71); GOLD=(187,149,73); GREY=(120,120,120); RED=(200,30,30); GREEN=(20,140,60)
    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(True, margin=15)
    pdf.add_font('dv', '', os.path.join(BASE_DIR, 'assets', 'fonts', 'DejaVuSans.ttf'))
    pdf.add_font('dv', 'B', os.path.join(BASE_DIR, 'assets', 'fonts', 'DejaVuSans-Bold.ttf'))
    pdf.add_page()
    try:
        pdf.image(os.path.join(BASE_DIR, 'static', 'img', 'logo.png'), x=12, y=10, h=14)
    except Exception:
        pass
    pdf.set_xy(32, 11); pdf.set_font('dv', 'B', 16); pdf.set_text_color(*NAVY)
    pdf.cell(0, 8, 'Εστία — CONDIAN HOTELS · Αναφορά Πισινών', ln=1)
    pdf.set_x(32); pdf.set_font('dv', '', 11); pdf.set_text_color(*GREY)
    pdf.cell(0, 6, 'Ημερομηνία: ' + rep_date.strftime('%d/%m/%Y'), ln=1)
    pdf.ln(8)
    if not records:
        pdf.set_font('dv', '', 12); pdf.set_text_color(*GREY)
        pdf.cell(0, 8, 'Δεν υπάρχουν καταγραφές για αυτή την ημέρα.', ln=1)
    cur = None
    for r in records:
        pname = (r.pool.hotel.name + ' — ' + r.pool.name) if (r.pool and r.pool.hotel) else (r.pool.name if r.pool else '—')
        if pname != cur:
            cur = pname
            pdf.ln(2); pdf.set_font('dv', 'B', 12); pdf.set_text_color(*NAVY); pdf.set_fill_color(240,243,245)
            pdf.cell(0, 8, '  ' + pname, ln=1, fill=True)
        per = 'Πρωί' if r.period == 'morning' else 'Απόγευμα'
        pdf.set_font('dv', 'B', 10); pdf.set_text_color(*GOLD)
        pdf.cell(0, 6, per + ' — ' + (r.user.full_name if r.user else ''), ln=1)
        def line(lbl, val, key, unit=''):
            if val is None: return
            mn, mx = POOL_LIMITS.get(key, (None, None))
            ok = (mn is None or val >= mn) and (mx is None or val <= mx)
            pdf.set_font('dv', '', 10); pdf.set_text_color(40,40,40); pdf.cell(62, 5, '   ' + lbl)
            pdf.set_text_color(*(GREEN if ok else RED))
            pdf.cell(0, 5, str(val) + unit + ('' if ok else '  (εκτός ορίων)'), ln=1)
        line('Ελεύθερο χλώριο', r.free_chlorine, 'free_chlorine', ' mg/L')
        line('Συνδεδεμένο χλώριο', r.combined_chlorine, 'combined_chlorine', ' mg/L')
        line('pH', r.ph, 'ph')
        line('Θερμοκρασία', r.temp, 'temp', ' C')
        line('Θολότητα', r.turbidity, 'turbidity', ' NTU')
        line('Κυανουρικό', r.cyanuric_acid, 'cyanuric_acid', ' mg/L')
        line('Αλκαλικότητα', r.total_alkalinity, 'total_alkalinity', ' mg/L')
        line('ORP', r.orp, 'orp', ' mV')
        if r.notes:
            pdf.set_x(pdf.l_margin); pdf.set_font('dv', '', 9); pdf.set_text_color(*GREY)
            pdf.multi_cell(pdf.epw, 5, '   Σημ: ' + r.notes)
        _acts = compute_pool_actions(r)
        if _acts:
            pdf.set_x(pdf.l_margin); pdf.set_font('dv', 'B', 9); pdf.set_text_color(180, 83, 9)
            pdf.multi_cell(pdf.epw, 5, 'Προτεινόμενες ενέργειες:')
            for a in _acts:
                pdf.set_x(pdf.l_margin); pdf.set_font('dv', '', 9)
                pdf.set_text_color(185, 28, 28) if a['urgent'] else pdf.set_text_color(80, 80, 80)
                pdf.multi_cell(pdf.epw, 4.6, ('• ΕΠΕΙΓΟΝ — ' if a['urgent'] else '• ') + a['label'] + ': ' + a['action'])
    return bytes(pdf.output())


# v12.5 — Records: print-ready εξαγωγές (PDF + XLSX)
_RECORDS_HEADERS = ['Ημ/νία', 'Ώρα', 'Τύπος', 'Βάρδια', 'Ξενοδοχείο', 'Πισίνα / Δίκτυο', 'Υπεύθυνος', 'Διόρθωση']

def _records_row(it):
    return [
        it['date'].strftime('%d/%m/%Y') if it.get('date') else '',
        it['when'].strftime('%H:%M') if it.get('when') else '',
        'Πισίνα' if it['kind'] == 'pool' else 'Νερά',
        'Πρωί' if it['period'] == 'morning' else 'Απόγευμα',
        it.get('hotel') or '', it.get('place') or '', it.get('user') or '',
        'διορθώθηκε' if it.get('updated') else '',
    ]

def build_records_pdf(items, ftype='all'):
    """Branded, print-ready PDF λίστας υποβολών (Records) — fpdf2 landscape."""
    from fpdf import FPDF
    NAVY=(25,56,71); GREY=(120,120,120)
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(True, margin=12)
    pdf.add_font('dv', '', os.path.join(BASE_DIR, 'assets', 'fonts', 'DejaVuSans.ttf'))
    pdf.add_font('dv', 'B', os.path.join(BASE_DIR, 'assets', 'fonts', 'DejaVuSans-Bold.ttf'))
    pdf.add_page()
    try:
        pdf.image(os.path.join(BASE_DIR, 'static', 'img', 'logo.png'), x=12, y=9, h=13)
    except Exception:
        pass
    pdf.set_xy(30, 10); pdf.set_font('dv', 'B', 15); pdf.set_text_color(*NAVY)
    pdf.cell(0, 8, 'Εστία — CONDIAN HOTELS · Records', ln=1)
    pdf.set_x(30); pdf.set_font('dv', '', 10); pdf.set_text_color(*GREY)
    pdf.cell(0, 6, _RECORDS_TYPE_LABEL.get(ftype, 'Όλες οι υποβολές')
             + ' · Εκτύπωση: ' + date.today().strftime('%d/%m/%Y')
             + ' · Σύνολο: ' + str(len(items)), ln=1)
    pdf.ln(6)
    widths = [24, 16, 22, 24, 62, 62, 50, 26]
    pdf.set_font('dv', 'B', 9); pdf.set_text_color(255, 255, 255); pdf.set_fill_color(*NAVY)
    for h, w in zip(_RECORDS_HEADERS, widths):
        pdf.cell(w, 8, h, border=0, fill=True, align='L')
    pdf.ln(8)
    if not items:
        pdf.set_font('dv', '', 11); pdf.set_text_color(*GREY)
        pdf.cell(0, 10, 'Δεν υπάρχουν υποβολές.', ln=1)
        return bytes(pdf.output())
    fill = False
    for it in items:
        pdf.set_fill_color(243, 247, 250) if fill else pdf.set_fill_color(255, 255, 255)
        pdf.set_font('dv', '', 8.5); pdf.set_text_color(40, 40, 40)
        for val, w in zip(_records_row(it), widths):
            s = str(val)
            while s and pdf.get_string_width(s) > w - 2 and len(s) > 3:
                s = s[:-2]
            pdf.cell(w, 7, s, border=0, fill=True, align='L')
        pdf.ln(7)
        fill = not fill
    return bytes(pdf.output())

def build_records_xlsx(items, ftype='all'):
    """Print-ready Excel (.xlsx) λίστας υποβολών — openpyxl."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.properties import PageSetupProperties
    wb = Workbook(); ws = wb.active; ws.title = 'Records'
    ncol = len(_RECORDS_HEADERS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    t = ws.cell(row=1, column=1,
                value='Εστία — CONDIAN HOTELS · Records (' + _RECORDS_TYPE_LABEL.get(ftype, 'Όλες οι υποβολές') + ')')
    t.font = Font(bold=True, size=14, color='193847')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    ws.cell(row=2, column=1,
            value='Εκτύπωση: ' + date.today().strftime('%d/%m/%Y') + ' · Σύνολο: ' + str(len(items))
            ).font = Font(size=10, color='777777')
    hdr = 4
    navy = PatternFill('solid', fgColor='193847')
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for j, h in enumerate(_RECORDS_HEADERS, 1):
        c = ws.cell(row=hdr, column=j, value=h)
        c.font = Font(bold=True, color='FFFFFF'); c.fill = navy
        c.alignment = Alignment(horizontal='left', vertical='center'); c.border = border
    for i, it in enumerate(items):
        for j, v in enumerate(_records_row(it), 1):
            c = ws.cell(row=hdr + 1 + i, column=j, value=v)
            c.border = border; c.alignment = Alignment(horizontal='left', vertical='center')
    for j, w in enumerate([13, 8, 11, 12, 26, 26, 22, 13], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = 'A%d' % (hdr + 1)
    ws.print_title_rows = '%d:%d' % (hdr, hdr)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ── Background email senders (τρέχουν σε thread, με app context) ──
def _bg_send_water(rid, uid):
    with app.app_context():
        r = WaterRecord.query.get(rid); u = User.query.get(uid)
        if r and u:
            send_report_email(r, u)

def _bg_send_pool(rid, uid):
    with app.app_context():
        r = PoolRecord.query.get(rid); u = User.query.get(uid)
        if r and u:
            send_pool_report_email(r, u)

def _bg_send_fault(fid):
    with app.app_context():
        fr = FaultReport.query.get(fid)
        if fr:
            send_fault_email(fr)


# ──────────────────────────────────────────────────────────────────────────
#  AUTH
# ──────────────────────────────────────────────────────────────────────────
def landing_for(user):
    r = role_rank(user.role)
    if r >= ROLE_RANK['admin']:
        return url_for('dashboard')
    if r >= ROLE_RANK['viewer'] and r != ROLE_RANK['staff']:
        return url_for('pools_dashboard')   # manager / viewer
    return url_for('pools_app')             # staff
@app.route('/')
def index():
    if 'user_id' in session:
        user = User.query.get(session['user_id'])
        if user:
            return render_template('shell.html', user=user, is_admin=is_admin(),
                                   is_master=(user.role == 'masteradmin'), can_log=can_log(),
                                   rank=role_rank(user.role), RANK=ROLE_RANK)
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        _k = _login_key()
        if _login_blocked(_k):
            return render_template('login.html', error='Πολλες αποτυχημενες προσπαθειες. Δοκιμασε ξανα σε λιγα λεπτα.')
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username, is_active=True).first()
        if user and user.approved and getattr(user, 'login_enabled', True) is not False and check_password_hash(user.password, password):
            _login_reset(_k)
            session['user_id']   = user.id
            session['user_name'] = user.full_name
            session['user_role'] = user.role
            session['language']  = user.language
            log_activity('login')
            return redirect(url_for('index'))
        _login_record_fail(_k)
        error = 'Λαθος username η password'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    log_activity('logout')
    session.clear()
    return redirect(url_for('login'))

# ── v12.9 — Μηνύματα / Ρυθμίσεις / Λογαριασμός ──
@app.route('/messages')
def messages():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    uid = session['user_id']
    box = request.args.get('box', 'in')
    if box == 'out':
        items = (Message.query.filter_by(sender_id=uid)
                 .order_by(Message.created_at.desc()).limit(100).all())
    else:
        items = (Message.query.filter_by(recipient_id=uid)
                 .order_by(Message.created_at.desc()).limit(100).all())
        for m in items:
            if not m.is_read:
                m.is_read = True
        db.session.commit()
    people = (User.query.filter(User.id != uid, User.is_active == True, User.approved == True)
              .order_by(User.full_name).all())
    return render_template('messages.html', items=items, box=box, people=people, me=current_user())

@app.route('/messages/send', methods=['POST'])
def messages_send():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    uid = session['user_id']
    rid = request.form.get('recipient_id', type=int)
    body = (request.form.get('body') or '').strip()
    if rid and body and User.query.get(rid):
        db.session.add(Message(sender_id=uid, recipient_id=rid, body=body[:4000]))
        me = current_user()
        db.session.add(Notification(user_id=rid,
            text='Νέο μήνυμα από ' + (me.full_name if me else ''), link='/messages?embed=1'))
        db.session.commit()
    return redirect('/messages?box=out&embed=1')

@app.route('/settings')
def settings_page():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('settings.html', me=current_user())

@app.route('/account')
def account_page():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('account.html', me=current_user())

@app.route('/register', methods=['GET', 'POST'])
def register():
    if not REG_CODE:
        return render_template('register.html', enabled=False, error=None, done=False)
    error = None
    if request.method == 'POST':
        fm = request.form
        if fm.get('reg_code', '').strip() != REG_CODE:
            error = 'Λάθος κωδικός εγγραφής'
        elif not (fm.get('username') and fm.get('password') and fm.get('full_name')):
            error = 'Συμπλήρωσε ονοματεπώνυμο, username και κωδικό'
        elif User.query.filter_by(username=fm['username'].strip()).first():
            error = 'Το username υπάρχει ήδη'
        else:
            db.session.add(User(
                username=fm['username'].strip(),
                password=generate_password_hash(fm['password']),
                full_name=fm['full_name'].strip(),
                email=fm.get('email', '').strip(),
                phone=fm.get('phone', '').strip(),
                role='staff', language=fm.get('language', 'el'),
                approved=False, is_active=True))
            db.session.commit()
            return render_template('register.html', enabled=True, error=None, done=True)
    return render_template('register.html', enabled=True, error=error, done=False)

@app.route('/profile', methods=['GET', 'POST'])
def profile():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = User.query.get(session['user_id'])
    saved = False
    if request.method == 'POST':
        fm = request.form
        user.email = fm.get('email', '').strip()
        user.phone = fm.get('phone', '').strip()
        if fm.get('language') in ('el', 'en'):
            user.language = fm['language']; session['language'] = fm['language']
        if fm.get('password'):
            user.password = generate_password_hash(fm['password'])
        photo = request.files.get('photo')
        if photo and photo.filename and photo.mimetype and photo.mimetype.startswith('image/'):
            raw = photo.read()
            if 0 < len(raw) <= 800 * 1024:
                import base64
                user.avatar = 'data:' + photo.mimetype + ';base64,' + base64.b64encode(raw).decode()
        db.session.commit()
        log_activity('profile_update')
        saved = True
    return render_template('profile.html', user=user, saved=saved)


# ──────────────────────────────────────────────────────────────────────────
#  WATER LOG ROUTES
# ──────────────────────────────────────────────────────────────────────────
def water_systems_payload(user=None):
    """v12.3 — allowed ξενοδοχεία + ενεργά δίκτυα νερού τους (mirror hotels_payload)."""
    if user is not None and role_rank(user.role) < ROLE_RANK['admin']:
        hotels = allowed_hotels(user)
    else:
        hotels = Hotel.query.filter_by(is_active=True).order_by(Hotel.name).all()
    payload = []
    for h in hotels:
        systems = (WaterSystem.query.filter_by(hotel_id=h.id, is_active=True)
                   .order_by(WaterSystem.name).all())
        payload.append({'id': h.id, 'name': h.name,
                        'systems': [{'id': s.id, 'name': s.name, 'location': s.location or ''}
                                    for s in systems]})
    return hotels, payload

@app.route('/app')
def water_app():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if not can_log():
        return redirect(url_for('pools_dashboard'))
    user = User.query.get(session['user_id'])
    hotels, systems_json = water_systems_payload(user)
    # σημερινές καταγραφές ανά δίκτυο: {system_id: {period: record_id}}
    done = {}
    for r in WaterRecord.query.filter_by(record_date=date.today()).all():
        if r.water_system_id:
            done.setdefault(str(r.water_system_id), {})[r.period] = r.id
    return render_template('app.html', user=user, hotels=hotels,
                           systems_json=systems_json, done=done)

@app.route('/submit', methods=['POST'])
def submit():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Μη εξουσιοδοτημενο'}), 401
    user   = User.query.get(session['user_id'])
    data   = request.form
    period = data.get('period', 'morning')
    ws_id  = data.get('water_system_id')
    ws = WaterSystem.query.filter_by(id=ws_id, is_active=True).first() if ws_id else None
    if not ws:
        return jsonify({'success': False, 'message': 'Επιλεξτε δικτυο νερου'}), 400
    if not can_access_water_system(user, ws):
        return jsonify({'success': False, 'message': 'Δεν εχεις προσβαση σε αυτο το δικτυο'}), 403
    # v12.3 — dedup ΑΝΑ δίκτυο (όχι πια κοινό record για όλους)
    record = WaterRecord.query.filter_by(water_system_id=ws.id,
                                         record_date=date.today(), period=period).first()
    if record:
        apply_record(record, data, period)
        record.updated_at = datetime.utcnow()
        record.updated_by = user.id
    else:
        record = WaterRecord(user_id=user.id, water_system_id=ws.id,
                             record_date=date.today(), period=period)
        apply_record(record, data, period)
        db.session.add(record)
    db.session.commit()
    place = (ws.hotel.name + ' · ' + ws.name) if ws.hotel else ws.name
    log_activity('water_submit', f'{place} ({period})', hotel_id=(ws.hotel_id if ws else None))
    _wa = compute_water_actions(record)
    if _wa:
        notify_admins((('ΕΠΕΙΓΟΝ — ' if any(x['urgent'] for x in _wa) else '')
                       + f'Μέτρηση νερού/ΖΝΧ εκτός ορίων — {place}.'), '/dashboard')
    threading.Thread(target=_bg_send_water, args=(record.id, user.id), daemon=True).start()
    period_gr = 'Πρωι' if period == 'morning' else 'Απογευμα'
    return jsonify({'success': True, 'record_id': record.id,
                    'message': f'Καταγραφη {ws.name} ({period_gr}) αποθηκευτηκε!'})

@app.route('/edit/<int:record_id>', methods=['GET', 'POST'])
def edit_record(record_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user   = User.query.get(session['user_id'])
    record = WaterRecord.query.get_or_404(record_id)
    if record.water_system and not can_access_water_system(user, record.water_system):
        return redirect(url_for('water_app'))
    if not is_admin() and record.record_date != date.today():
        return redirect(url_for('water_app'))
    if request.method == 'POST':
        apply_record(record, request.form, record.period)
        record.updated_at = datetime.utcnow()
        record.updated_by = user.id
        db.session.commit()
        if user.role == 'admin':
            return redirect(url_for('dashboard') + '?success=updated')
        return redirect(url_for('water_app'))
    return render_template('edit.html', user=user, record=record)

@app.route('/api/record/<int:record_id>')
def api_record(record_id):
    if 'user_id' not in session:
        return jsonify({}), 401
    r = WaterRecord.query.get_or_404(record_id)
    _ws = r.water_system
    return jsonify({
        'id': r.id, 'period': r.period,
        'water_system_id': r.water_system_id,
        'water_system': (_ws.name if _ws else None),
        'hotel': (_ws.hotel.name if _ws and _ws.hotel else None),
        'record_date': r.record_date.strftime('%d/%m/%Y'),
        'clo2_tank': r.clo2_tank, 'clo2_kitchen': r.clo2_kitchen,
        'clo2_remote': r.clo2_remote, 'clo2_dhw_out': r.clo2_dhw_out,
        'clo2_dhw_return': r.clo2_dhw_return, 'clo2_ro': r.clo2_ro,
        'location_kitchen': r.location_kitchen, 'location_remote': r.location_remote,
        'temp_tank': r.temp_tank, 'temp_dhw_out': r.temp_dhw_out,
        'temp_dhw_return': r.temp_dhw_return, 'temp_ro': r.temp_ro,
        'temp_kitchen_cold': r.temp_kitchen_cold, 'temp_kitchen_hot': r.temp_kitchen_hot,
        'temp_remote_cold': r.temp_remote_cold, 'temp_remote_hot': r.temp_remote_hot,
        'ph_tank': r.ph_tank, 'notes': r.notes or ''
    })

@app.route('/set-language/<lang>')
def set_language(lang):
    if lang in ['el', 'en', 'uk'] and 'user_id' in session:
        session['language'] = lang
        user = User.query.get(session['user_id'])
        if user:
            user.language = lang
            db.session.commit()
    return redirect(request.referrer or url_for('water_app'))


# ──────────────────────────────────────────────────────────────────────────
#  POOL LOG ROUTES
# ──────────────────────────────────────────────────────────────────────────
def hotels_payload(user=None):
    if user is not None and role_rank(user.role) < ROLE_RANK['admin']:
        hotels = allowed_hotels(user)
    else:
        hotels = Hotel.query.filter_by(is_active=True).order_by(Hotel.name).all()
    payload = [{
        'id': h.id, 'name': h.name,
        'pools': [{'id': p.id, 'name': p.name, 'location': p.location or '',
                   'type': p.pool_type, 'volume_m3': p.volume_m3}
                  for p in h.pools if p.is_active]
    } for h in hotels]
    return hotels, payload

@app.route('/pools')
def pools_app():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if not can_log():
        return redirect(url_for('pools_dashboard'))
    user = User.query.get(session['user_id'])
    hotels, hotels_json = hotels_payload(user)
    todays = PoolRecord.query.filter_by(record_date=date.today()).all()
    done = {}
    for r in todays:
        done.setdefault(str(r.pool_id), []).append(r.period)
    return render_template('pools.html', user=user, hotels=hotels,
                           hotels_json=hotels_json, done=done, limits=POOL_LIMITS,
                           action_rules=ACTION_RULES, param_labels=PARAM_LABELS, safety_note=SAFETY_NOTE)

@app.route('/submit-pool', methods=['POST'])
def submit_pool():
    if 'user_id' not in session or not can_log():
        return jsonify({'success': False, 'message': 'Μη εξουσιοδοτημενο'}), 401
    user   = User.query.get(session['user_id'])
    data   = request.form
    period = data.get('period', 'morning')
    pool_id = data.get('pool_id')
    pool = Pool.query.filter_by(id=pool_id, is_active=True).first() if pool_id else None
    if not pool:
        return jsonify({'success': False, 'message': 'Επιλεξτε πισινα'}), 400
    if not can_access_pool(user, pool):
        return jsonify({'success': False, 'message': 'Δεν εχεις προσβαση σε αυτη την πισινα'}), 403
    record = PoolRecord.query.filter_by(pool_id=pool.id, record_date=date.today(), period=period).first()
    if record:
        apply_pool_record(record, data, period)
        record.updated_at = datetime.utcnow()
        record.updated_by = user.id
    else:
        record = PoolRecord(pool_id=pool.id, user_id=user.id, record_date=date.today(), period=period)
        apply_pool_record(record, data, period)
        db.session.add(record)
    db.session.commit()
    log_activity('pool_submit', f'{pool.name} ({period})', hotel_id=pool.hotel_id)
    _a = compute_pool_actions(record)
    if _a:
        notify_admins((('ΕΠΕΙΓΟΝ — ' if any(x['urgent'] for x in _a) else '') + 'Μέτρηση εκτός ορίων: ' + pool.name), '/pools/dashboard')
    threading.Thread(target=_bg_send_pool, args=(record.id, user.id), daemon=True).start()
    period_gr = 'Πρωι' if period == 'morning' else 'Απογευμα'
    return jsonify({'success': True, 'message': f'Καταγραφη {pool.name} ({period_gr}) αποθηκευτηκε!'})

@app.route('/pools/edit/<int:record_id>', methods=['GET', 'POST'])
def edit_pool_record(record_id):
    if 'user_id' not in session or not can_log():
        return redirect(url_for('login'))
    user   = current_user()
    record = PoolRecord.query.get_or_404(record_id)
    if not can_access_pool(user, record.pool):
        return redirect(url_for('pools_app'))
    if not is_admin() and record.record_date != date.today():
        return redirect(url_for('pools_app'))
    if request.method == 'POST':
        apply_pool_record(record, request.form, record.period)
        record.updated_at = datetime.utcnow()
        record.updated_by = user.id
        db.session.commit()
        log_activity('pool_edit', record.pool.name if record.pool else '')
        if user.role == 'admin':
            return redirect(url_for('pools_dashboard') + '?success=updated')
        return redirect(url_for('pools_app'))
    return render_template('pool_edit.html', user=user, record=record, limits=POOL_LIMITS)

@app.route('/api/pool-record/<int:record_id>')
def api_pool_record(record_id):
    if 'user_id' not in session:
        return jsonify({}), 401
    r = PoolRecord.query.get_or_404(record_id)
    return jsonify({
        'id': r.id, 'period': r.period, 'pool_id': r.pool_id,
        'record_date': r.record_date.strftime('%d/%m/%Y'),
        'free_chlorine': r.free_chlorine, 'combined_chlorine': r.combined_chlorine,
        'ph': r.ph, 'temp': r.temp, 'turbidity': r.turbidity,
        'cyanuric_acid': r.cyanuric_acid, 'total_alkalinity': r.total_alkalinity,
        'orp': r.orp, 'backwash_done': r.backwash_done, 'notes': r.notes or ''
    })

@app.route('/pools/report.pdf')
def pool_report_pdf():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    ds = request.args.get('date')
    try:
        rep_date = datetime.strptime(ds, '%Y-%m-%d').date() if ds else date.today()
    except ValueError:
        rep_date = date.today()
    records = (PoolRecord.query.filter_by(record_date=rep_date)
               .order_by(PoolRecord.pool_id, PoolRecord.period).all())
    pdf_bytes = build_pool_report_pdf(rep_date, records)
    fname = 'pool-report-' + rep_date.strftime('%Y-%m-%d') + '.pdf'
    return Response(pdf_bytes, mimetype='application/pdf',
                    headers={'Content-Disposition': 'attachment; filename=' + fname})


# ──────────────────────────────────────────────────────────────────────────
#  AI ASSISTANT ROUTES
# ──────────────────────────────────────────────────────────────────────────
@app.route('/assistant')
def assistant_page():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = User.query.get(session['user_id'])
    hotels, hotels_json = hotels_payload(current_user())
    return render_template('assistant.html', user=user,
                           hotels_json=hotels_json,
                           configured=(resolve_provider() is not None))

@app.route('/api/assistant', methods=['POST'])
def api_assistant():
    if 'user_id' not in session:
        return jsonify({'reply': '', 'error': 'auth'}), 401
    data = request.get_json(force=True, silent=True) or {}
    raw  = data.get('messages', [])
    pool_id = data.get('pool_id')
    messages = []
    for m in raw[-12:]:
        role = m.get('role')
        content = (m.get('content') or '').strip()
        if role in ('user', 'assistant') and content:
            messages.append({'role': role, 'content': content[:4000]})
    if not messages:
        return jsonify({'reply': '', 'error': 'empty'}), 400
    pool = Pool.query.filter_by(id=pool_id, is_active=True).first() if pool_id else None
    context = build_pool_context(pool)
    system = POOL_ASSISTANT_PROMPT + '\n\n# ΔΕΔΟΜΕΝΑ ΕΦΑΡΜΟΓΗΣ (τρέχουσα κατάσταση)\n' + context
    reply, err = call_llm(system, messages)
    if err == 'not_configured':
        return jsonify({'reply': '', 'error': 'not_configured'})
    if err:
        return jsonify({'reply': '', 'error': err}), 502
    return jsonify({'reply': reply, 'error': None})


# ──────────────────────────────────────────────────────────────────────────
#  ADMIN DASHBOARDS
# ──────────────────────────────────────────────────────────────────────────
@app.route('/dashboard')
def dashboard():
    if not is_admin():
        return redirect(url_for('login'))
    hotels  = Hotel.query.filter_by(is_active=True).order_by(Hotel.name).all()
    systems = (WaterSystem.query.filter_by(is_active=True)
               .order_by(WaterSystem.hotel_id, WaterSystem.name).all())
    # v12.3 — φίλτρα ανά ξενοδοχείο / δίκτυο / χρήστη
    f_hotel  = request.args.get('hotel_id', type=int) or active_hotel_id()
    f_system = request.args.get('system_id', type=int)
    f_staff  = request.args.get('staff_id', type=int)
    def _scope(q):
        if f_system:
            return q.filter(WaterRecord.water_system_id == f_system)
        if f_hotel:
            sids = [s.id for s in systems if s.hotel_id == f_hotel]
            return q.filter(WaterRecord.water_system_id.in_(sids or [-1]))
        return q
    q = _scope(WaterRecord.query)
    if f_staff:
        q = q.filter(WaterRecord.user_id == f_staff)
    q = apply_period(q, WaterRecord.record_date)
    records = q.order_by(WaterRecord.record_date.desc(), WaterRecord.period).limit(60).all()
    users   = User.query.filter_by(is_active=True, approved=True).all()
    today_q = _scope(WaterRecord.query.filter_by(record_date=date.today()))
    today_records = today_q.order_by(WaterRecord.water_system_id).all()
    today_m = [r for r in today_records if r.period == 'morning']
    today_a = [r for r in today_records if r.period == 'afternoon']
    return render_template('dashboard.html', records=records, users=users,
                           today_morning_list=today_m, today_afternoon_list=today_a,
                           hotels=hotels, systems=systems,
                           f_hotel=f_hotel, f_system=f_system, f_staff=f_staff)


# v12.4 — Χρήστες: δική τους σελίδα στη Διαχείριση (έφυγαν από τον Πίνακα Νερών)
@app.route('/dashboard/users')
def users_admin():
    if not is_admin():
        return redirect(url_for('login'))
    # v12.44 — το εισαγμένο προσωπικό (login_enabled=False) ΔΕΝ εμφανίζεται εδώ· μένει καθαρή για πραγματικούς λογαριασμούς
    users   = [u for u in User.query.filter_by(is_active=True, approved=True).all() if getattr(u, 'login_enabled', None) is not False]
    inactive = [u for u in User.query.filter_by(is_active=False).all() if getattr(u, 'login_enabled', None) is not False]
    pending = User.query.filter_by(is_active=True, approved=False).all()
    try:                                  # v12.15 — ειδικότητες (Module Βλαβοληψία)
        import faults as _flt
        fault_specs = [s.name for s in _flt.Specialty.query.filter_by(is_active=True).order_by(_flt.Specialty.sort, _flt.Specialty.name).all()]
        user_specs = {}
        for us in _flt.UserSpecialty.query.all():
            user_specs.setdefault(us.user_id, []).append(us.specialty)
    except Exception:
        fault_specs, user_specs = [], {}
    return render_template('users_admin.html', users=users + inactive, pending=pending,
                           all_hotels=Hotel.query.filter_by(is_active=True).order_by(Hotel.name).all(),
                           role_labels=ROLE_LABELS, me=current_user(),
                           fault_specs=fault_specs, user_specs=user_specs)


# v12.4 — Records: ενιαίο feed υποβολών (πισίνες + νερά χρήσης), role-scoped
_RECORDS_TYPE_LABEL = {'all': 'Όλες οι υποβολές', 'pools': 'Πισίνες', 'water': 'Νερά Χρήσης'}

def _records_items(user, ftype='all'):
    """Κοινός builder της λίστας Records (page + εξαγωγές PDF/XLSX)."""
    hids = scoped_hotel_ids(user)
    items = []
    if ftype in ('all', 'pools'):
        pids = [p.id for p in Pool.query.all() if p.hotel_id in hids]
        _pq = apply_period(PoolRecord.query.filter(PoolRecord.pool_id.in_(pids or [-1])), PoolRecord.record_date)
        for r in (_pq.order_by(PoolRecord.recorded_at.desc()).limit(120).all()):
            items.append({
                'kind': 'pool', 'when': r.recorded_at, 'date': r.record_date,
                'period': r.period,
                'hotel': r.pool.hotel.name if r.pool and r.pool.hotel else '—',
                'place': r.pool.name if r.pool else '—',
                'user': r.user.full_name if r.user else '—',
                'updated': bool(r.updated_at),
                'edit_url': '/pools/edit/%d' % r.id,
            })
    if ftype in ('all', 'water'):
        sids = [s.id for s in WaterSystem.query.all() if s.hotel_id in hids]
        _wq = apply_period(WaterRecord.query.filter(WaterRecord.water_system_id.in_(sids or [-1])), WaterRecord.record_date)
        for r in (_wq.order_by(WaterRecord.recorded_at.desc()).limit(120).all()):
            ws = r.water_system
            items.append({
                'kind': 'water', 'when': r.recorded_at, 'date': r.record_date,
                'period': r.period,
                'hotel': ws.hotel.name if ws and ws.hotel else '—',
                'place': ws.name if ws else '—',
                'user': r.user.full_name if r.user else '—',
                'updated': bool(r.updated_at),
                'edit_url': '/edit/%d' % r.id,
            })
    items.sort(key=lambda x: x['when'] or datetime.min, reverse=True)
    return items[:150]

@app.route('/records')
def records_feed():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = current_user()
    ftype = request.args.get('type', 'all')          # all | pools | water
    items = _records_items(user, ftype)
    return render_template('records.html', items=items, ftype=ftype, user=user)

# v12.5 — Εξαγωγές Records (έτοιμες προς εκτύπωση)
@app.route('/records/export.pdf')
def records_export_pdf():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = current_user()
    ftype = request.args.get('type', 'all')
    items = _records_items(user, ftype)
    pdf_bytes = build_records_pdf(items, ftype)
    fname = 'estia-records-%s-%s.pdf' % (ftype, date.today().strftime('%Y-%m-%d'))
    return Response(pdf_bytes, mimetype='application/pdf',
                    headers={'Content-Disposition': 'attachment; filename=' + fname})

@app.route('/records/export.xlsx')
def records_export_xlsx():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = current_user()
    ftype = request.args.get('type', 'all')
    items = _records_items(user, ftype)
    xlsx_bytes = build_records_xlsx(items, ftype)
    fname = 'estia-records-%s-%s.xlsx' % (ftype, date.today().strftime('%Y-%m-%d'))
    return Response(xlsx_bytes,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename=' + fname})

@app.route('/pools/dashboard')
def pools_dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = current_user()
    hotels = allowed_hotels(user)
    hids = {h.id for h in hotels}
    pools = [p for p in Pool.query.filter_by(is_active=True).all() if p.hotel_id in hids]
    pids = {p.id for p in pools}
    # v12.5 — φίλτρα ανά ξενοδοχείο / πισίνα / χρήστη (mirror Πίνακα Νερών)
    f_hotel = request.args.get('hotel_id', type=int) or active_hotel_id()
    f_pool  = request.args.get('pool_id', type=int)
    f_staff = request.args.get('staff_id', type=int)
    def _scope(q):
        if f_pool:
            return q.filter(PoolRecord.pool_id == f_pool)
        if f_hotel:
            hpids = [p.id for p in pools if p.hotel_id == f_hotel]
            return q.filter(PoolRecord.pool_id.in_(hpids or [-1]))
        return q
    base = PoolRecord.query
    if role_rank(user.role) < ROLE_RANK['admin']:
        base = base.filter(PoolRecord.pool_id.in_(pids if pids else [-1]))
    q = _scope(base)
    if f_staff:
        q = q.filter(PoolRecord.user_id == f_staff)
    q = apply_period(q, PoolRecord.record_date)
    records = q.order_by(PoolRecord.record_date.desc(), PoolRecord.recorded_at.desc()).limit(80).all()
    users = User.query.filter_by(is_active=True, approved=True).all()
    today_records = (_scope(base.filter(PoolRecord.record_date == date.today()))
                     .order_by(PoolRecord.pool_id).all())
    today_m = [r for r in today_records if r.period == 'morning']
    today_a = [r for r in today_records if r.period == 'afternoon']
    return render_template('pools_dashboard.html', hotels=hotels, pools=pools,
                           records=records, limits=POOL_LIMITS, users=users,
                           today_morning_list=today_m, today_afternoon_list=today_a,
                           f_hotel=f_hotel, f_pool=f_pool, f_staff=f_staff,
                           is_admin=is_admin(), user=user)

@app.route('/dashboard/add-user', methods=['POST'])
def add_user():
    if not is_admin():
        return redirect(url_for('login'))
    data = request.form
    if not User.query.filter_by(username=data['username']).first():
        db.session.add(User(
            username=data['username'],
            password=generate_password_hash(data['password']),
            full_name=data['full_name'],
            email=data.get('email', '').strip(),
            phone=data.get('phone', '').strip(),
            role=data.get('role', 'staff'),
            language=data.get('language', 'el')
        ))
        db.session.commit()
        log_activity('user_add', data.get('username', ''))
    return redirect(url_for('users_admin') + '?success=user_added')

@app.route('/dashboard/delete-user/<int:user_id>', methods=['POST'])
def delete_user(user_id):
    if not is_admin():
        return redirect(url_for('login'))
    user = User.query.get(user_id)
    if user and role_rank(user.role) < ROLE_RANK['admin']:
        user.is_active = False
        db.session.commit()
        log_activity('user_delete', user.username)
    return redirect(url_for('users_admin'))

@app.route('/dashboard/approve-user/<int:user_id>', methods=['POST'])
def approve_user(user_id):
    if not is_admin():
        return redirect(url_for('login'))
    u = User.query.get(user_id)
    if u:
        u.approved = True; u.is_active = True
        db.session.commit()
        log_activity('user_approve', u.username)
        notify(u.id, 'Ο λογαριασμός σου εγκρίθηκε. Καλώς ήρθες!', '/pools')
    return redirect(url_for('users_admin') + '?success=user_approved')

def _protected(actor, target):
    # δεν επιτρέπεται σε χαμηλότερο ρόλο να πειράξει masteradmin
    return role_rank(target.role) >= ROLE_RANK['masteradmin'] and role_rank(actor.role) < ROLE_RANK['masteradmin']

@app.route('/dashboard/edit-user/<int:user_id>', methods=['POST'])
def edit_user(user_id):
    if not is_admin():
        return redirect(url_for('login'))
    actor = current_user(); u = User.query.get(user_id)
    if not u or _protected(actor, u):
        return redirect(url_for('users_admin'))
    fm = request.form
    u.full_name = (fm.get('full_name') or u.full_name).strip() or u.full_name
    u.email = fm.get('email', '').strip()
    u.phone = fm.get('phone', '').strip()
    # v12.12 — πλήρες edit: username (με έλεγχο μοναδικότητας) + γλώσσα
    new_un = (fm.get('username') or '').strip()
    if new_un and new_un != u.username and not User.query.filter_by(username=new_un).first():
        u.username = new_un
    if fm.get('language') in ('el', 'en', 'uk'):
        u.language = fm['language']
    new_role = fm.get('role', u.role)
    if new_role in ROLE_RANK and role_rank(new_role) <= role_rank(actor.role):
        u.role = new_role
    hids = [int(x) for x in fm.getlist('hotel_ids') if x.isdigit()]
    u.hotels = Hotel.query.filter(Hotel.id.in_(hids)).all() if hids else []
    db.session.commit()
    log_activity('user_edit', u.username)
    return redirect(url_for('users_admin') + '?success=user_edited')

@app.route('/dashboard/reset-password/<int:user_id>', methods=['POST'])
def reset_password(user_id):
    if not is_admin():
        return redirect(url_for('login'))
    actor = current_user(); u = User.query.get(user_id)
    if u and not _protected(actor, u):
        newpw = request.form.get('password', '').strip()
        if newpw:
            u.password = generate_password_hash(newpw); db.session.commit()
            log_activity('user_reset_password', u.username)
    return redirect(url_for('users_admin') + '?success=password_reset')

@app.route('/dashboard/toggle-user/<int:user_id>', methods=['POST'])
def toggle_user(user_id):
    if not is_admin():
        return redirect(url_for('login'))
    actor = current_user(); u = User.query.get(user_id)
    if u and u.id != actor.id and not _protected(actor, u):
        u.is_active = not u.is_active; db.session.commit()
        log_activity('user_toggle', u.username + (' -> active' if u.is_active else ' -> inactive'))
    return redirect(url_for('users_admin'))

# ── Εβδομαδιαία κάλυψη ──
@app.route('/pools/coverage')
def pools_coverage():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = current_user()
    hids = {h.id for h in allowed_hotels(user)}
    pools = [p for p in Pool.query.filter_by(is_active=True).order_by(Pool.hotel_id, Pool.name).all() if p.hotel_id in hids]
    days = [date.today() - timedelta(days=i) for i in range(6, -1, -1)]
    recs = PoolRecord.query.filter(
        PoolRecord.pool_id.in_([p.id for p in pools] or [-1]),
        PoolRecord.record_date.in_(days)).all()
    cov = {}
    for r in recs:
        cov.setdefault((r.pool_id, r.record_date), set()).add(r.period)
    grid = []
    for p in pools:
        cells = []
        for d in days:
            sset = cov.get((p.id, d), set())
            cells.append({'m': 'morning' in sset, 'a': 'afternoon' in sset})
        grid.append({'pool': p, 'cells': cells})
    return render_template('coverage.html', grid=grid, days=days)

# ── Αναφορά βλάβης: μεταφέρθηκε στο faults.py (Module Βλαβοληψία v12.14) ──

# ── In-app ειδοποιήσεις ──
@app.route('/notifications')
def notifications():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    items = Notification.query.filter_by(user_id=session['user_id']).order_by(Notification.id.desc()).limit(100).all()
    return render_template('notifications.html', items=items)

def _wants_json():
    return (request.is_json or request.args.get('ajax')
            or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            or 'application/json' in (request.headers.get('Accept') or ''))

@app.route('/notifications/read', methods=['POST'])
def notifications_read():
    if 'user_id' not in session:
        return ('', 401) if _wants_json() else redirect(url_for('login'))
    Notification.query.filter_by(user_id=session['user_id'], is_read=False).update({'is_read': True})
    db.session.commit()
    return ('', 204) if _wants_json() else redirect(url_for('notifications'))

@app.route('/notifications/read/<int:nid>', methods=['POST'])
def notifications_read_one(nid):
    if 'user_id' not in session:
        return ('', 401)
    n = Notification.query.filter_by(id=nid, user_id=session['user_id']).first()
    if n and not n.is_read:
        n.is_read = True
        db.session.commit()
    return ('', 204)

@app.route('/notifications/list')
def notifications_list():
    if 'user_id' not in session:
        return jsonify({'items': [], 'unread': 0}), 401
    items = Notification.query.filter_by(user_id=session['user_id']).order_by(Notification.id.desc()).limit(20).all()
    unread = Notification.query.filter_by(user_id=session['user_id'], is_read=False).count()
    def _ago(dt):
        if not dt: return ''
        sec = (datetime.utcnow() - dt).total_seconds()
        if sec < 60: return 'μόλις τώρα'
        if sec < 3600: return f'{int(sec//60)}\' πριν'
        if sec < 86400: return f'{int(sec//3600)}ω πριν'
        if sec < 604800: return f'{int(sec//86400)}η πριν'
        return dt.strftime('%d/%m %H:%M')
    return jsonify({'unread': unread, 'items': [
        {'id': n.id, 'text': n.text or '', 'link': (n.link or ''),
         'is_read': bool(n.is_read), 'ago': _ago(n.created_at)} for n in items]})

# ── v12.36 — «Τι νέο» (ιστορικό εκδόσεων, admin) ─────────────────────────────
@app.route('/dashboard/whatsnew')
def whatsnew():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('whatsnew.html', changelog=CHANGELOG,
                           cur_version=APP_VERSION, cur_build=APP_BUILD)

# ── v12.45 — ROADMAP ολόκληρης της Εστίας (status: done/progress/planned/idea) ──
ROADMAP = [
    {'area': 'Συντήρηση', 'items': [
        {'t': 'Module Βλάβες (workflow, SLA, αυτόματη ανάθεση, ιστορικό)', 's': 'done'},
        {'t': 'Πισίνες & Νερά Χρήσης — μετρήσεις, όρια, ενέργειες', 's': 'done'},
        {'t': 'Μητρώο Εξοπλισμού (assets, κόστος, QR) — δένει με Βλάβες', 's': 'planned'},
        {'t': 'Προληπτική συντήρηση / προγραμματισμένες εργασίες', 's': 'idea'},
    ]},
    {'area': 'Προσωπικό / HR', 'items': [
        {'t': 'Πρόγραμμα Εργασίας — βάρδιες, κανόνες, υποβολή λογιστηρίου', 's': 'done'},
        {'t': 'Οργανόγραμμα (εργαζόμενοι σε τμήματα) + εκκαθάριση διπλών', 's': 'done'},
        {'t': 'Multi-file importer ιστορικών (παλιά έτη/προσωπικό)', 's': 'progress'},
        {'t': 'Μισθοδοσία — θεμέλιο: μητρώο/εταιρείες/PII/συντελεστές (Φ1)', 's': 'progress'},
        {'t': 'Μισθοδοσία — μηχανή Management+Λογιστήριο + import Epsilon (Φ2)', 's': 'progress'},
        {'t': 'Μισθοδοσία — εκκαθαριστικά PDF, reporting, πρόβλεψη (Φ3+)', 's': 'planned'},
        {'t': 'Αξιολόγηση προσωπικού (πάνω στα ερωτηματολόγια)', 's': 'planned'},
    ]},
    {'area': 'Guest experience / Υποδοχή', 'items': [
        {'t': 'Ερωτηματολόγια (builder, δημόσιος σύνδεσμος, NPS)', 's': 'done'},
        {'t': 'Logbook (Παράπονα/Συμβάντα) — 1.762 ιστορικά', 's': 'progress'},
        {'t': 'Guest App (αιτήματα δωματίου)', 's': 'idea'},
    ]},
    {'area': 'Πλατφόρμα', 'items': [
        {'t': 'Ρόλοι/δικαιώματα, μενού ανά ρόλο, multi-hotel scope', 's': 'done'},
        {'t': 'Dashboard με widgets, dark mode, παλέτες', 's': 'done'},
        {'t': 'Backups → SharePoint, Κέντρο Εισαγωγής, Feedback χρηστών', 's': 'done'},
        {'t': 'Μεταφράσεις EN/UK', 's': 'planned'},
        {'t': 'Σύνδεση μισθοδοσίας με ΕΡΓΑΝΗ/Epsilon (μέσω προγράμματος)', 's': 'idea'},
    ]},
]
ROADMAP_STATUS = {
    'done':     ('✓ Ολοκληρώθηκε', '#dcfce7', '#166534'),
    'progress': ('● Σε εξέλιξη',    '#dbeafe', '#185FA5'),
    'planned':  ('◷ Σχεδιασμένο',   '#fef3c7', '#92400e'),
    'idea':     ('💡 Ιδέα',         '#f1f5f9', '#64748b'),
}

@app.route('/dashboard/roadmap')
def roadmap_page():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('roadmap.html', roadmap=ROADMAP, status=ROADMAP_STATUS,
                           cur_version=APP_VERSION, cur_build=APP_BUILD)

# ── v12.46 — Βοήθεια: FAQ + Επίλυση προβλημάτων (admin items κρυφά από non-admin) ──
FAQ_ITEMS = [
    {'c': 'Γενικά', 'q': 'Πώς αλλάζω σε σκούρο θέμα ή χρώματα;',
     'a': 'Πάνω δεξιά στο εικονίδιο προφίλ → «Θέμα» (Φωτεινό/Σκούρο/Σύστημα) και «Παλέτα» για χρώματα.'},
    {'c': 'Γενικά', 'q': 'Βλέπω μόνο το ξενοδοχείο μου — γιατί;',
     'a': 'Ο λογαριασμός σου είναι συνδεδεμένος με συγκεκριμένο ξενοδοχείο. Αν έχεις πρόσβαση σε πολλά, χρησιμοποίησε τον εναλλάκτη ξενοδοχείου πάνω-πάνω.'},
    {'c': 'Γενικά', 'q': 'Πού βλέπω τι άλλαξε ή τι έρχεται;',
     'a': 'Ενημέρωση → «Τι νέο» (αλλαγές ανά έκδοση) και «Roadmap» (η πορεία της πλατφόρμας).'},
    {'c': 'Γενικά', 'q': 'Πώς στέλνω σχόλιο, ιδέα ή πρόβλημα;',
     'a': 'Ενημέρωση → «Στείλε feedback». Φτάνει αμέσως στη διοίκηση και μας βοηθάει να προτεραιοποιήσουμε.'},
    {'c': 'Πρόγραμμα Εργασίας', 'q': 'Πώς καταχωρώ βάρδια;',
     'a': 'Προσωπικό → «Πρόγραμμα Εργασίας» → διάλεξε ξενοδοχείο/εβδομάδα → κλικ σε ένα κελί → διάλεξε κωδικό. Για «ΕΡΓ» βάζεις και ώρες (π.χ. 07:00–15:30).'},
    {'c': 'Πρόγραμμα Εργασίας', 'q': 'Τι σημαίνουν οι κωδικοί;',
     'a': 'ΕΡΓ=Εργασία (με ώρες), ΑΝ=Ρεπό, ΑΔ=Κανονική άδεια, ΑΣΘ=Αναρρωτική, ΑΠ=Αδικαιολόγητη απουσία, ΑΡΓ=Αργία, ΑΝΕΥ=Άνευ αποδοχών, Ειδ.Α=Ειδική άδεια. Υπάρχει υπόμνημα στο κάτω μέρος της σελίδας.'},
    {'c': 'Πρόγραμμα Εργασίας', 'q': 'Πώς δουλεύω σε πολλές εβδομάδες ή τμήματα μαζί;',
     'a': 'Με τον επιλογέα «Εβδομάδες» βλέπεις όσες εβδομάδες θες μπροστά. Με τα chips τμημάτων (π.χ. Housekeeping, Ρεσεψιόν) δείχνεις/κρύβεις τμήματα δυναμικά.'},
    {'c': 'Πρόγραμμα Εργασίας', 'q': 'Πώς στέλνω το πρόγραμμα στο λογιστήριο;',
     'a': 'Στην εβδομάδα που θες, κουμπί «Αποστολή στο λογιστήριο». Στέλνει όλα τα τμήματα του ξενοδοχείου μαζί. Πρέπει πρώτα να περνούν οι κανόνες (π.χ. κάθε εργαζόμενος ≥1 ρεπό/εβδομάδα).'},
    {'c': 'Πρόγραμμα Εργασίας', 'q': 'Γιατί δεν με αφήνει να αλλάξω την τρέχουσα εβδομάδα;',
     'a': 'Η εβδομάδα «κλειδώνει» στην προθεσμία (προεπιλογή: Πέμπτη της προηγούμενης εβδομάδας). Δουλεύεις τις επόμενες εβδομάδες· για έκτακτη αλλαγή χρειάζεται ο admin.'},
    {'c': 'Λογαριασμοί & Πρόσβαση', 'q': 'Δεν βλέπω κάποια επιλογή στο μενού.',
     'a': 'Κάθε ρόλος βλέπει διαφορετικά πράγματα. Ζήτα από τον διαχειριστή να σου το ενεργοποιήσει (Διαχείριση → «Μενού ανά ρόλο»).', 'admin': False},
    {'c': 'Εισαγωγές (Διαχείριση)', 'q': 'Πώς βάζω μαζικά εργαζόμενους & προγράμματα;', 'admin': True,
     'a': 'Διαχείριση → «Πρόγραμμα · Εισαγωγή». Ανέβασε πολλά .xlsx μαζί — η πλατφόρμα αναγνωρίζει αυτόματα μητρώο μισθοδοσίας ή πρόγραμμα. Πρόταση: πρώτα τα μητρώα (καθαρή λίστα), μετά τα προγράμματα (βάρδιες).'},
    {'c': 'Εισαγωγές (Διαχείριση)', 'q': 'Πώς ρυθμίζω προθεσμίες/κανόνες/αργίες;', 'admin': True,
     'a': 'Διαχείριση → «Πρόγραμμα · Ρυθμίσεις»: ημέρα/ώρα προθεσμίας, ορίζοντα εβδομάδων, αργίες, κανόνες υποβολής, κωδικοί βαρδιών.'},
]
TROUBLE_ITEMS = [
    {'p': 'Δεν μπορώ να συνδεθώ.',
     's': 'Έλεγξε username/κωδικό. Αν είσαι «εισαγμένο προσωπικό» δεν έχεις δικαίωμα σύνδεσης — ζήτα από τον διαχειριστή να ενεργοποιήσει τον λογαριασμό και να σου δώσει κωδικό.'},
    {'p': 'Η βάρδια δεν αποθηκεύεται / βγάζει «Κλειδωμένη εβδομάδα».',
     's': 'Πέρασε η προθεσμία αυτής της εβδομάδας. Δούλεψε σε επόμενη εβδομάδα, ή ζήτα από τον admin να ξεκλειδώσει.'},
    {'p': 'Η «Αποστολή στο λογιστήριο» μπλοκάρεται.',
     's': 'Κάποιος κανόνας δεν περνά (π.χ. εργαζόμενος χωρίς ρεπό). Δες τη λίστα ελέγχων κάτω από το πρόγραμμα, διόρθωσε και ξαναπροσπάθησε.'},
    {'p': 'Βλέπω διπλά ονόματα μετά από εισαγωγή.', 'admin': True,
     's': 'Διαχείριση → «Πρόγραμμα · Προσωπικό» → ενότητα «Πιθανές διπλοεγγραφές»: διάλεξε ποιον κρατάς και πάτησε «Συγχώνευση».'},
    {'p': 'Εργαζόμενος σε λάθος τμήμα ή ξενοδοχείο.', 'admin': True,
     's': 'Διαχείριση → «Πρόγραμμα · Προσωπικό»: άλλαξε τμήμα/ξενοδοχείο και «Αποθήκευση».'},
    {'p': 'Ανέβασα λάθος/«βρόμικο» αρχείο.', 'admin': True,
     's': 'Διαχείριση → «Πρόγραμμα · Εισαγωγή» → «Καθαρισμός εισαγμένου προσωπικού» (σβήνει μόνο εισαγμένους, όχι πραγματικούς λογαριασμούς). Μετά ξανακάνε εισαγωγή από καθαρή πηγή.'},
    {'p': 'Το αρχείο βγήκε «Άγνωστο» στην εισαγωγή.', 'admin': True,
     's': 'Δεν είχε αναγνωρίσιμη δομή. Χρειάζεται είτε γραμμή «ΤΜΗΜΑ» με στήλες ημερομηνιών (πρόγραμμα), είτε κεφαλίδα μητρώου με «Τμήμα»+«Επώνυμο/Ονοματεπώνυμο».'},
    {'p': 'Δεν λαμβάνω ειδοποιήσεις email.', 'admin': True,
     's': 'Διαχείριση → «Email»: έλεγξε τη ρύθμιση αποστολής. Επίσης βεβαιώσου ότι ο λογαριασμός έχει σωστό email.'},
    {'p': 'Η σελίδα δείχνει παλιά δεδομένα.',
     's': 'Πάτησε ανανέωση (refresh) στον browser. Αν συνεχίζει, αποσυνδέσου και ξανασυνδέσου.'},
]

@app.route('/dashboard/help')
def help_page():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    adm = is_admin()
    faqs = [x for x in FAQ_ITEMS if adm or not x.get('admin')]
    trbl = [x for x in TROUBLE_ITEMS if adm or not x.get('admin')]
    cats = []
    for x in faqs:
        if x['c'] not in cats:
            cats.append(x['c'])
    return render_template('help.html', faqs=faqs, trbl=trbl, cats=cats, is_admin=adm)

# ── Inbox/διαχείριση βλαβών: μεταφέρθηκε στο faults.py (Module Βλαβοληψία v12.14) ──

# ── Email admin (Graph/SMTP, test, manual, log) ──
@app.route('/dashboard/email')
def email_admin():
    if not is_admin():
        return redirect(url_for('login'))
    logs = EmailLog.query.order_by(EmailLog.id.desc()).limit(50).all()
    mode = 'Microsoft Graph (365)' if GRAPH_CLIENT_ID else ('SMTP (' + SMTP_SERVER + ')' if EMAIL_PASSWORD else 'Δεν έχει ρυθμιστεί')
    return render_template('email_admin.html', logs=logs, mode=mode)

@app.route('/dashboard/email/test', methods=['POST'])
def email_test():
    if not is_admin():
        return redirect(url_for('login'))
    to = request.form.get('to', '').strip()
    recips = [to] if to else EMAIL_TO_LIST
    send_email('Εστία — Δοκιμαστικό email', '<p>Δοκιμαστικό email από την Εστία (CONDIAN HOTELS). Αν το βλέπεις, η αποστολή λειτουργεί.</p>', recips)
    log_activity('email_test', ', '.join(recips))
    return redirect(url_for('email_admin'))

@app.route('/dashboard/email/reminder', methods=['POST'])
def email_reminder_now():
    if not is_admin():
        return redirect(url_for('login'))
    miss = missing_today()
    if miss:
        send_reminder_email(miss)
    log_activity('email_reminder_manual', str(len(miss)))
    return redirect(url_for('email_admin'))

# ── Theming (admin) ──
@app.route('/dashboard/theme', methods=['GET', 'POST'])
def theme_admin():
    if not is_admin():
        return redirect(url_for('login'))
    import re as _re
    if request.method == 'POST':
        def setk(k, v):
            row = Setting.query.get('theme_' + k)
            if row:
                row.value = v
            else:
                db.session.add(Setting(key='theme_' + k, value=v))
        for k in ['primary', 'accent']:
            v = request.form.get(k, '')
            if _re.match(r'^#[0-9a-fA-F]{6}$', v):
                setk(k, v)
        setk('app_title', (request.form.get('app_title', 'Εστία') or 'Εστία')[:60])
        if request.form.get('reset_logo'):
            setk('logo', '')
        else:
            logo = request.files.get('logo')
            if logo and logo.filename and logo.mimetype and logo.mimetype.startswith('image/'):
                raw = logo.read()
                if 0 < len(raw) <= 400 * 1024:
                    import base64
                    setk('logo', 'data:' + logo.mimetype + ';base64,' + base64.b64encode(raw).decode())
        # v12.11 — εξατομίκευση login/register + social
        lsz = (request.form.get('login_logo_size', '88') or '88').strip()
        setk('login_logo_size', lsz if (lsz.isdigit() and 24 <= int(lsz) <= 200) else '88')
        lf = request.form.get('login_font', 'Space Grotesk')
        if lf in LOGIN_FONTS:
            setk('login_font', lf)
        setk('login_tagline', (request.form.get('login_tagline', '') or '')[:120])
        setk('social_facebook', (request.form.get('social_facebook', '') or '').strip()[:200])
        setk('social_linkedin', (request.form.get('social_linkedin', '') or '').strip()[:200])
        if request.form.get('reset_login_logo'):
            setk('login_logo', '')
        else:
            llogo = request.files.get('login_logo')
            if llogo and llogo.filename and llogo.mimetype and llogo.mimetype.startswith('image/'):
                lraw = llogo.read()
                if 0 < len(lraw) <= 400 * 1024:
                    import base64
                    setk('login_logo', 'data:' + llogo.mimetype + ';base64,' + base64.b64encode(lraw).decode())
        db.session.commit(); log_activity('theme_update')
        return redirect(url_for('theme_admin') + '?saved=1')
    return render_template('theme_admin.html', theme=get_theme(), login_fonts=LOGIN_FONTS)

# ── AI σύνδεση (masteradmin) ──
@app.route('/dashboard/ai', methods=['GET', 'POST'])
def ai_admin():
    u = current_user()
    if u is None or u.role != 'masteradmin':
        return redirect(url_for('login'))
    def setk(k, v):
        row = Setting.query.get('ai_' + k)
        if row:
            row.value = v
        else:
            db.session.add(Setting(key='ai_' + k, value=v))
    if request.method == 'POST':
        setk('provider', request.form.get('provider', 'auto'))
        setk('model', request.form.get('model', '').strip())
        if request.form.get('clear_keys'):
            setk('anthropic', ''); setk('openai', '')
        else:
            ak = request.form.get('anthropic', '').strip()
            if ak and not ak.startswith('*'):
                setk('anthropic', ak)
            ok = request.form.get('openai', '').strip()
            if ok and not ok.startswith('*'):
                setk('openai', ok)
        db.session.commit(); log_activity('ai_config')
        return redirect(url_for('ai_admin') + '?saved=1')
    c = get_ai_config()
    masked = {'anthropic': ('*' * 8 + c['anthropic'][-4:]) if c['anthropic'] else '',
              'openai': ('*' * 8 + c['openai'][-4:]) if c['openai'] else ''}
    return render_template('ai_admin.html', cfg=c, masked=masked, configured=(resolve_provider() is not None))

# ── GENERIC AREAS (Πλατφόρμα Συντήρησης) ──
def can_access_area(u, area):
    if u is None or area is None:
        return False
    if role_rank(u.role) >= ROLE_RANK['admin']:
        return True
    assigned = {h.id for h in u.hotels if h.is_active}
    if not assigned:
        return True
    return area.hotel_id in assigned

def _templates_json():
    out = {}
    for t in MonitorTemplate.query.filter_by(is_active=True).order_by(MonitorTemplate.sort).all():
        out[t.key] = {'name': t.name, 'icon': t.icon, 'frequency': t.frequency,
                      'params': [{'pkey': p.pkey, 'label': p.label, 'unit': p.unit,
                                  'min': p.min_v, 'max': p.max_v, 'periodic': p.periodic} for p in t.params]}
    return out

def reading_cells(r):
    try:
        vals = json.loads(r.values or '{}')
    except Exception:
        vals = {}
    tpl = MonitorTemplate.query.get(r.template_key); cells = []
    if tpl:
        for p in tpl.params:
            v = vals.get(p.pkey); ok = True
            if v is not None:
                try:
                    fv = float(v); ok = (p.min_v is None or fv >= p.min_v) and (p.max_v is None or fv <= p.max_v)
                except Exception:
                    pass
            cells.append({'label': p.label, 'value': v, 'unit': p.unit, 'ok': ok})
    return cells

@app.route('/areas')
def areas_app():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if not can_log():
        return redirect(url_for('areas_dashboard'))
    user = current_user()
    hids = {h.id for h in allowed_hotels(user)}
    areas = [a for a in Area.query.filter_by(is_active=True).order_by(Area.template_key, Area.name).all() if a.hotel_id in hids]
    areas_json = [{'id': a.id, 'name': a.name, 'location': a.location or '',
                   'hotel': a.hotel.name if a.hotel else '', 'template': a.template_key} for a in areas]
    todays = Reading.query.filter_by(record_date=date.today()).all()
    done = {}
    for r in todays:
        done.setdefault(str(r.area_id), []).append(r.period)
    return render_template('areas.html', user=user, areas_json=areas_json,
                           templates=_templates_json(), done=done, freq_periods=FREQ_PERIODS)

@app.route('/areas/submit', methods=['POST'])
def areas_submit():
    if 'user_id' not in session or not can_log():
        return jsonify({'success': False, 'message': 'Μη εξουσιοδοτημενο'}), 401
    user = current_user(); data = request.form
    area = Area.query.filter_by(id=data.get('area_id'), is_active=True).first() if data.get('area_id') else None
    if not area:
        return jsonify({'success': False, 'message': 'Επιλεξτε τομεα'}), 400
    if not can_access_area(user, area):
        return jsonify({'success': False, 'message': 'Δεν εχεις προσβαση'}), 403
    period = data.get('period', 'day')
    tpl = MonitorTemplate.query.get(area.template_key)
    vals = {}
    for p in (tpl.params if tpl else []):
        v = data.get(p.pkey, '')
        if v not in (None, ''):
            try:
                vals[p.pkey] = float(v)
            except (ValueError, TypeError):
                pass
    rec = Reading.query.filter_by(area_id=area.id, record_date=date.today(), period=period).first()
    if rec:
        rec.values = json.dumps(vals); rec.notes = data.get('notes', '')
        rec.updated_at = datetime.utcnow(); rec.updated_by = user.id
    else:
        rec = Reading(area_id=area.id, template_key=area.template_key, user_id=user.id,
                      record_date=date.today(), period=period, values=json.dumps(vals), notes=data.get('notes', ''))
        db.session.add(rec)
    db.session.commit()
    log_activity('area_submit', area.name, hotel_id=area.hotel_id)
    acts = area_actions(rec)
    if acts:
        notify_admins('Εκτός ορίων: ' + area.name + ' (' + str(len(acts)) + ')', '/areas/dashboard')
    return jsonify({'success': True, 'message': 'Καταγραφη ' + area.name + ' αποθηκευτηκε!'})

@app.route('/areas/dashboard')
def areas_dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = current_user()
    hids = scoped_hotel_ids(user)
    areas = [a for a in Area.query.filter_by(is_active=True).all() if a.hotel_id in hids]
    aids = {a.id for a in areas}
    q = Reading.query
    if role_rank(user.role) < ROLE_RANK['admin']:
        q = q.filter(Reading.area_id.in_(aids if aids else [-1]))
    recs = q.order_by(Reading.record_date.desc(), Reading.recorded_at.desc()).limit(120).all()
    rows = [(r, reading_cells(r)) for r in recs]
    return render_template('areas_dashboard.html', areas=areas, rows=rows,
                           is_admin=is_admin(), templates=MonitorTemplate.query.order_by(MonitorTemplate.sort).all())

# ── v12.34 — Προσωπικό dashboard με tiles (drag&drop, ανά χρήστη) ─────────────
# Κατάλογος διαθέσιμων tiles. kind: 'kpi'|'widget'. scope: 'all'=όλοι · 'admin'=admin/master.
# v12.37 — κατηγορίες (ομάδες) + γεωμετρία πλέγματος (GridStack 12 στηλών)
CAT_LABELS = {'quality':'Ποιότητα','faults':'Βλάβες','records':'Καταγραφές','org':'Οργανισμός','people':'Χρήστες','pinned':'Καρφιτσωμένα'}
CATEGORIES = [('quality','Ποιότητα','ti-checkup-list'),('faults','Βλάβες','ti-tool'),
              ('records','Καταγραφές','ti-clipboard-check'),('org','Οργανισμός','ti-building'),
              ('people','Χρήστες','ti-users')]
TILE_CATALOG = [
    {'id':'compliance',   'title':'Συμμόρφωση σήμερα',          'icon':'ti-checkup-list',   'kind':'kpi',    'scope':'all',  'cat':'quality','w':3,'h':2,'minw':2,'minh':2},
    {'id':'alerts',       'title':'Εκτός ορίων',                'icon':'ti-alert-triangle', 'kind':'kpi',    'scope':'all',  'cat':'quality','w':3,'h':2,'minw':2,'minh':2},
    {'id':'open_faults',  'title':'Ανοιχτές βλάβες',            'icon':'ti-tool',           'kind':'kpi',    'scope':'all',  'cat':'faults', 'w':3,'h':2,'minw':2,'minh':2},
    {'id':'my_faults',    'title':'Δικές μου βλάβες',           'icon':'ti-user-cog',       'kind':'kpi',    'scope':'all',  'cat':'faults', 'w':3,'h':2,'minw':2,'minh':2},
    {'id':'pool_today',   'title':'Μετρήσεις πισινών σήμερα',   'icon':'ti-pool',           'kind':'kpi',    'scope':'all',  'cat':'records','w':3,'h':2,'minw':2,'minh':2},
    {'id':'water_today',  'title':'Μετρήσεις νερών σήμερα',     'icon':'ti-droplet',        'kind':'kpi',    'scope':'all',  'cat':'records','w':3,'h':2,'minw':2,'minh':2},
    {'id':'surveys_today','title':'Απαντήσεις ερωτημ. σήμερα',  'icon':'ti-clipboard-check','kind':'kpi',    'scope':'all',  'cat':'records','w':3,'h':2,'minw':2,'minh':2},
    {'id':'hotels',       'title':'Ξενοδοχεία',                 'icon':'ti-building',       'kind':'kpi',    'scope':'all',  'cat':'org',    'w':3,'h':2,'minw':2,'minh':2},
    {'id':'checkpoints',  'title':'Σημεία ελέγχου',             'icon':'ti-map-pin',        'kind':'kpi',    'scope':'all',  'cat':'org',    'w':3,'h':2,'minw':2,'minh':2},
    {'id':'users',        'title':'Χρήστες',                    'icon':'ti-users',          'kind':'kpi',    'scope':'admin','cat':'people', 'w':3,'h':2,'minw':2,'minh':2},
    {'id':'pending',      'title':'Εκκρεμείς εγγραφές',         'icon':'ti-user-plus',      'kind':'kpi',    'scope':'admin','cat':'people', 'w':3,'h':2,'minw':2,'minh':2},
    {'id':'coverage',     'title':'Κάλυψη ανά ξενοδοχείο',      'icon':'ti-checkup-list',   'kind':'widget', 'scope':'all',  'cat':'org',    'w':5,'h':4,'minw':3,'minh':3},
    {'id':'alerts_list',  'title':'Εκτός ορίων (πρόσφατα)',     'icon':'ti-alert-triangle', 'kind':'widget', 'scope':'all',  'cat':'quality','w':4,'h':4,'minw':3,'minh':3},
    {'id':'faults_list',  'title':'Ανοιχτές βλάβες (λίστα)',    'icon':'ti-tool',           'kind':'widget', 'scope':'all',  'cat':'faults', 'w':4,'h':4,'minw':3,'minh':3},
    {'id':'activity',     'title':'Πρόσφατη δραστηριότητα',     'icon':'ti-history',        'kind':'widget', 'scope':'all',  'cat':'people', 'w':4,'h':4,'minw':3,'minh':3},
    {'id':'pending_list', 'title':'Εκκρεμείς εγγραφές (λίστα)', 'icon':'ti-user-plus',      'kind':'widget', 'scope':'admin','cat':'people', 'w':4,'h':3,'minw':3,'minh':2},
]
DEFAULT_TILES_ADMIN = ['compliance','alerts','open_faults','pending','hotels','checkpoints',
                       'coverage','alerts_list','faults_list','pending_list']
DEFAULT_TILES_STAFF = ['my_faults','open_faults','pool_today','water_today','surveys_today',
                       'coverage','alerts_list','faults_list']

def _tile_allowed(t, admin):
    return admin or t.get('scope') != 'admin'

def _pin_dict(pw):
    return {'id': 'pin_%d' % pw.id, 'title': pw.title, 'icon': pw.icon or 'ti-pin',
            'kind': 'pinned', 'scope': 'all', 'cat': 'pinned',
            'w': 4, 'h': 3, 'minw': 3, 'minh': 2, 'link': pw.link or '', 'pid': pw.id}

def _user_pinned(user):
    try:
        return [_pin_dict(p) for p in PinnedWidget.query.filter_by(user_id=user.id).order_by(PinnedWidget.id).all()]
    except Exception:
        return []

def _allowed_catalog(user, admin):
    return [t for t in TILE_CATALOG if _tile_allowed(t, admin)] + _user_pinned(user)

def _user_layout(user, admin):
    """Λίστα {id,x,y,w,h} για τον χρήστη (saved geometry ή default auto-flow), φιλτραρισμένη."""
    valid = {t['id']: t for t in _allowed_catalog(user, admin)}
    items = None
    try:
        raw = getattr(user, 'dashboard', None)
        if raw:
            data = json.loads(raw)
            if isinstance(data, dict) and isinstance(data.get('items'), list):
                items = []
                for it in data['items']:
                    tid = it.get('id')
                    if tid in valid:
                        d = valid[tid]
                        items.append({'id': tid, 'x': it.get('x'), 'y': it.get('y'),
                                      'w': max(d['minw'], int(it.get('w') or d['w'])),
                                      'h': max(d['minh'], int(it.get('h') or d['h']))})
            elif isinstance(data, dict) and isinstance(data.get('tiles'), list):
                items = [{'id': i, 'x': None, 'y': None, 'w': valid[i]['w'], 'h': valid[i]['h']} for i in data['tiles'] if i in valid]
            elif isinstance(data, list):
                items = [{'id': i, 'x': None, 'y': None, 'w': valid[i]['w'], 'h': valid[i]['h']} for i in data if i in valid]
    except Exception:
        items = None
    if not items:
        deflist = DEFAULT_TILES_ADMIN if admin else DEFAULT_TILES_STAFF
        items = [{'id': i, 'x': None, 'y': None, 'w': valid[i]['w'], 'h': valid[i]['h']} for i in deflist if i in valid]
    seen, out = set(), []
    for it in items:
        if it['id'] not in seen:
            seen.add(it['id']); out.append(it)
    return out

@app.route('/overview')
def overview():
    user = current_user()
    if not user:
        return redirect(url_for('login'))
    admin = is_admin()
    today = date.today()
    hotels = scoped_hotels(user)
    hids = {h.id for h in hotels}
    F = __import__('faults').Fault

    # — Κάλυψη ανά ξενοδοχείο + συμμόρφωση —
    cov = []; total_exp = total_done = 0
    for h in hotels:
        exp = done = 0
        for p in [p for p in h.pools if p.is_active]:
            exp += 2
            done += min(2, PoolRecord.query.filter_by(pool_id=p.id, record_date=today).count())
        for a in Area.query.filter_by(hotel_id=h.id, is_active=True).all():
            tpl = MonitorTemplate.query.get(a.template_key)
            n = len(periods_for(tpl.frequency)) if tpl else 1
            exp += n
            done += min(n, Reading.query.filter_by(area_id=a.id, record_date=today).count())
        pct = int(100 * done / exp) if exp else 100
        cov.append({'hotel': h.name, 'exp': exp, 'done': done, 'pct': pct})
        total_exp += exp; total_done += done
    compliance = int(100 * total_done / total_exp) if total_exp else 100

    # — Εκτός ορίων (alerts) —
    rec_dates = [today, today - timedelta(days=1)]
    alerts = []
    for r in PoolRecord.query.filter(PoolRecord.record_date.in_(rec_dates)).order_by(PoolRecord.recorded_at.desc()).limit(80).all():
        if r.pool and r.pool.hotel_id not in hids: continue
        for a in compute_pool_actions(r):
            alerts.append({'where': (r.pool.hotel.name + ' / ' + r.pool.name) if (r.pool and r.pool.hotel) else (r.pool.name if r.pool else '—'),
                           'label': a['label'], 'urgent': a['urgent'], 'date': r.record_date})
    for r in Reading.query.filter(Reading.record_date.in_(rec_dates)).order_by(Reading.recorded_at.desc()).limit(80).all():
        if r.area and r.area.hotel_id not in hids: continue
        for a in area_actions(r):
            alerts.append({'where': (r.area.hotel.name + ' / ' + r.area.name) if (r.area and r.area.hotel) else '—',
                           'label': a['label'], 'urgent': False, 'date': r.record_date})
    alerts = sorted(alerts, key=lambda x: (not x['urgent'], x['date']))[:15]

    # — Βλάβες (νέο μοντέλο Fault) —
    _open = F.query.filter(F.hotel_id.in_(hids or [-1])).filter(~F.status.in_(('done','not_done','resubmitted')))
    open_faults_n = _open.count()
    faults = _open.order_by(F.submitted_at.desc()).limit(10).all()
    my_faults_n = F.query.filter(F.hotel_id.in_(hids or [-1])).filter(
        ~F.status.in_(('done','not_done','resubmitted'))).filter(
        (F.assigned_user_id == user.id) | (F.submitted_by == user.id)).count()

    # — Μετρήσεις σήμερα —
    pool_today = PoolRecord.query.join(Pool, PoolRecord.pool_id == Pool.id).filter(
        Pool.hotel_id.in_(hids or [-1]), PoolRecord.record_date == today).count()
    try:
        water_today = WaterRecord.query.filter(WaterRecord.record_date == today).count()
    except Exception:
        water_today = 0
    try:
        SR = __import__('surveys').SurveyResponse
        _t0 = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        surveys_today = SR.query.filter(SR.submitted_at >= _t0).count()
    except Exception:
        surveys_today = 0

    pending = User.query.filter_by(is_active=True, approved=False).all()
    activity = ActivityLog.query.order_by(ActivityLog.id.desc()).limit(12).all()
    checkpoints = sum(len([p for p in h.pools if p.is_active]) for h in hotels) + \
                  Area.query.filter(Area.hotel_id.in_(hids or [-1]), Area.is_active == True).count()

    tiledata = {
        'compliance': compliance, 'alerts': len(alerts), 'open_faults': open_faults_n,
        'my_faults': my_faults_n, 'pool_today': pool_today, 'water_today': water_today,
        'surveys_today': surveys_today, 'hotels': len(hotels), 'checkpoints': checkpoints,
        'users': User.query.filter_by(is_active=True, approved=True).count(), 'pending': len(pending),
        'cov': cov, 'alerts_list': alerts, 'faults_list': faults, 'pending_list': pending, 'activity': activity,
    }
    catalog = _allowed_catalog(user, admin)
    cmap = {t['id']: t for t in catalog}
    layout = _user_layout(user, admin)
    items = [{**cmap[it['id']], **it} for it in layout if it['id'] in cmap]
    predefined = [t for t in catalog if t.get('kind') != 'pinned']
    pinned = [t for t in catalog if t.get('kind') == 'pinned']
    return render_template('overview.html', tiledata=tiledata, items=items, catalog=catalog,
                           predefined=predefined, pinned=pinned, categories=CATEGORIES, cat_labels=CAT_LABELS,
                           is_admin=admin, areas_labels=AREA_LABELS, action_labels=ACTION_LABELS)

@app.route('/overview/layout', methods=['POST'])
def overview_layout():
    user = current_user()
    if not user:
        return ('', 401)
    admin = is_admin()
    valid = {t['id'] for t in TILE_CATALOG if _tile_allowed(t, admin)}
    try:
        payload = request.get_json(silent=True) or {}
        if payload.get('reset'):
            user.dashboard = None
            db.session.commit()
            return ('', 204)
        cmap = {t['id']: t for t in _allowed_catalog(user, admin)}
        out, seen = [], set()
        # νέα μορφή: items με γεωμετρία
        for it in (payload.get('items') or []):
            tid = it.get('id')
            if tid in cmap and tid not in seen:
                seen.add(tid); d = cmap[tid]
                def _ci(v, dv):
                    try: return int(v)
                    except Exception: return dv
                w = max(d['minw'], min(12, _ci(it.get('w'), d['w'])))
                h = max(d['minh'], _ci(it.get('h'), d['h']))
                rec = {'id': tid, 'w': w, 'h': h}
                if it.get('x') is not None: rec['x'] = max(0, _ci(it.get('x'), 0))
                if it.get('y') is not None: rec['y'] = max(0, _ci(it.get('y'), 0))
                out.append(rec)
        # συμβατότητα: παλιά μορφή tiles[]
        if not out and isinstance(payload.get('tiles'), list):
            for tid in payload['tiles']:
                if tid in cmap and tid not in seen:
                    seen.add(tid); d = cmap[tid]
                    out.append({'id': tid, 'w': d['w'], 'h': d['h']})
        user.dashboard = json.dumps({'items': out})
        db.session.commit()
        return ('', 204)
    except Exception as e:
        db.session.rollback()
        return (str(e), 400)

@app.route('/dashboard/pin', methods=['POST'])
def dashboard_pin():
    user = current_user()
    if not user:
        return ('', 401)
    p = request.get_json(silent=True) or {}
    title = (p.get('title') or '').strip()[:80]
    link = (p.get('link') or '').strip()[:300]
    icon = (p.get('icon') or 'ti-pin').strip()[:40]
    if not title:
        return jsonify({'error': 'title required'}), 400
    pw = PinnedWidget(user_id=user.id, title=title, link=link, icon=icon)
    db.session.add(pw); db.session.commit()
    return jsonify({'id': 'pin_%d' % pw.id, 'pid': pw.id})

@app.route('/dashboard/unpin/<int:pid>', methods=['POST'])
def dashboard_unpin(pid):
    user = current_user()
    if not user:
        return ('', 401)
    pw = PinnedWidget.query.filter_by(id=pid, user_id=user.id).first()
    if pw:
        db.session.delete(pw); db.session.commit()
    return ('', 204)

# ── v12.21 — Καταγραφή χρηστών (Activity Log) ────────────────────────────────
ACTION_LABELS = {
    'login': 'Σύνδεση', 'logout': 'Αποσύνδεση', 'profile_update': 'Ενημέρωση προφίλ',
    'water_submit': 'Καταγραφή νερών', 'pool_submit': 'Καταγραφή πισίνας', 'pool_edit': 'Διόρθωση πισίνας',
    'area_submit': 'Καταγραφή τομέα',
    'user_add': 'Νέος χρήστης', 'user_delete': 'Διαγραφή χρήστη', 'user_approve': 'Έγκριση χρήστη',
    'user_edit': 'Επεξεργασία χρήστη', 'user_reset_password': 'Επαναφορά κωδικού', 'user_toggle': 'Ενεργοπ./Απενεργ. χρήστη',
    'email_test': 'Δοκιμή email', 'email_reminder_manual': 'Χειροκίνητη υπενθύμιση',
    'theme_update': 'Αλλαγή εμφάνισης', 'ai_config': 'Ρύθμιση AI',
    'area_add': 'Νέος τομέας', 'template_new': 'Νέο template', 'template_edit': 'Επεξεργασία template',
    'seed_demo': 'Demo δεδομένα', 'water_system_add': 'Νέο δίκτυο νερού',
    'water_system_edit': 'Επεξεργασία δικτύου', 'water_system_delete': 'Διαγραφή δικτύου',
    'fault_report': 'Δήλωση βλάβης', 'faults_bulk': 'Μαζική ενέργεια βλαβών',
}
def _act_label(a):
    return ACTION_LABELS.get(a, a or '—')

ACTION_ICON = {
    'login': 'login', 'logout': 'logout', 'profile_update': 'user-edit',
    'water_submit': 'droplet', 'pool_submit': 'pool', 'pool_edit': 'edit', 'area_submit': 'checklist',
    'user_add': 'user-plus', 'user_delete': 'user-x', 'user_approve': 'user-check',
    'user_edit': 'user-edit', 'user_reset_password': 'key', 'user_toggle': 'toggle-right',
    'email_test': 'mail', 'email_reminder_manual': 'mail-forward', 'theme_update': 'palette', 'ai_config': 'plug',
    'area_add': 'map-pin-plus', 'template_new': 'template', 'template_edit': 'template', 'seed_demo': 'database',
    'water_system_add': 'plus', 'water_system_edit': 'edit', 'water_system_delete': 'trash',
    'fault_report': 'tool', 'faults_bulk': 'checks',
}
def _act_icon(a):
    return ACTION_ICON.get(a, 'point')

@app.route('/dashboard/activity')
def activity_log():
    if not is_admin():
        return redirect(url_for('login'))
    user = current_user()
    f_user   = request.args.get('user_id', type=int)
    f_action = request.args.get('action')
    search   = (request.args.get('q') or '').strip()
    chip     = request.args.get('chip')          # today | 7d | 30d | logins
    now = datetime.utcnow()
    midnight = now.replace(hour=0, minute=0, second=0, microsecond=0)
    aid = active_hotel_id()                       # v12.22 — scope ανά ξενοδοχείο
    def hscope(query):
        return query.filter(ActivityLog.hotel_id == aid) if aid else query
    base = apply_period(hscope(ActivityLog.query), ActivityLog.created_at, dt=True)
    q = apply_period(hscope(ActivityLog.query), ActivityLog.created_at, dt=True)
    if chip == 'today':  q = q.filter(ActivityLog.created_at >= midnight)
    elif chip == '7d':   q = q.filter(ActivityLog.created_at >= now - timedelta(days=7))
    elif chip == '30d':  q = q.filter(ActivityLog.created_at >= now - timedelta(days=30))
    elif chip == 'logins': q = q.filter(ActivityLog.action == 'login')
    if f_user:   q = q.filter(ActivityLog.user_id == f_user)
    if f_action: q = q.filter(ActivityLog.action == f_action)
    if search:
        like = '%%%s%%' % search
        q = q.filter((ActivityLog.detail.ilike(like)) | (ActivityLog.action.ilike(like)))
    logs = q.order_by(ActivityLog.created_at.desc()).limit(400).all()
    users = User.query.filter_by(is_active=True).order_by(User.full_name).all()
    actions = sorted({a[0] for a in db.session.query(ActivityLog.action).distinct().all() if a[0]})
    kpi = {
        'today':        hscope(ActivityLog.query).filter(ActivityLog.created_at >= midnight).count(),
        'active_today': hscope(db.session.query(ActivityLog.user_id)).filter(ActivityLog.created_at >= midnight).distinct().count(),
        'logins_today': hscope(ActivityLog.query).filter(ActivityLog.created_at >= midnight, ActivityLog.action == 'login').count(),
        'week':         hscope(ActivityLog.query).filter(ActivityLog.created_at >= now - timedelta(days=7)).count(),
    }
    # v12.22 — 7ήμερο mini γράφημα (scoped)
    spark = []
    for i in range(6, -1, -1):
        d0 = (midnight - timedelta(days=i))
        d1 = d0 + timedelta(days=1)
        cnt = hscope(ActivityLog.query).filter(ActivityLog.created_at >= d0, ActivityLog.created_at < d1).count()
        spark.append({'label': d0.strftime('%d/%m'), 'dow': ['Δε','Τρ','Τε','Πε','Πα','Σα','Κυ'][d0.weekday()], 'count': cnt})
    spark_max = max([x['count'] for x in spark] or [1]) or 1
    week_start = now - timedelta(days=7)
    per_user = []
    for u in users:
        last = hscope(ActivityLog.query).filter(ActivityLog.user_id == u.id).order_by(ActivityLog.created_at.desc()).first()
        cw = hscope(ActivityLog.query).filter(ActivityLog.user_id == u.id, ActivityLog.created_at >= week_start).count()
        if aid and not last and cw == 0:
            continue                              # σε scope ξενοδοχείου, κρύψε άσχετους χρήστες
        per_user.append({
            'u': u,
            'today': hscope(ActivityLog.query).filter(ActivityLog.user_id == u.id, ActivityLog.created_at >= midnight).count(),
            'week':  cw,
            'last':  last,
        })
    per_user.sort(key=lambda x: (x['last'].created_at if x['last'] else datetime.min), reverse=True)
    active_hotel = Hotel.query.get(aid) if aid else None
    return render_template('activity_log.html', logs=logs, users=users, actions=actions,
                           act_label=_act_label, act_icon=_act_icon, ACTION_LABELS=ACTION_LABELS,
                           f_user=f_user, f_action=f_action, search=search, chip=chip,
                           kpi=kpi, per_user=per_user, user=user, spark=spark, spark_max=spark_max,
                           active_hotel=active_hotel,
                           qs=request.query_string.decode('utf-8'))

# admin: areas management
@app.route('/dashboard/areas', methods=['GET', 'POST'])
def areas_admin():
    if not is_admin():
        return redirect(url_for('login'))
    if request.method == 'POST':
        hid = request.form.get('hotel_id'); tk = request.form.get('template_key'); nm = request.form.get('name', '').strip()
        if hid and tk and nm:
            db.session.add(Area(hotel_id=int(hid), template_key=tk, name=nm, location=request.form.get('location', '').strip()))
            db.session.commit(); log_activity('area_add', nm)
        return redirect(url_for('areas_admin') + '?saved=1')
    return render_template('areas_admin.html',
                           hotels=Hotel.query.filter_by(is_active=True).order_by(Hotel.name).all(),
                           templates=MonitorTemplate.query.filter_by(is_active=True).order_by(MonitorTemplate.sort).all(),
                           areas=Area.query.filter_by(is_active=True).all())

@app.route('/dashboard/area/<int:area_id>/delete', methods=['POST'])
def area_delete(area_id):
    if not is_admin():
        return redirect(url_for('login'))
    a = Area.query.get(area_id)
    if a:
        a.is_active = False; db.session.commit()
    return redirect(url_for('areas_admin'))

# admin: template editor (hybrid)
@app.route('/dashboard/templates')
def templates_admin():
    if not is_admin():
        return redirect(url_for('login'))
    return render_template('templates_admin.html',
                           templates=MonitorTemplate.query.order_by(MonitorTemplate.sort).all(),
                           freq_label=FREQ_LABEL)

@app.route('/dashboard/template/new', methods=['POST'])
def template_new():
    if not is_admin():
        return redirect(url_for('login'))
    key = (request.form.get('key', '').strip().lower() or '')
    key = ''.join(ch for ch in key if ch.isalnum() or ch == '_')[:30]
    if key and not MonitorTemplate.query.get(key):
        db.session.add(MonitorTemplate(key=key, name=request.form.get('name', key).strip(),
                       icon=request.form.get('icon', 'ti-checklist').strip() or 'ti-checklist',
                       frequency=request.form.get('frequency', 'daily'), sort=99))
        db.session.commit(); log_activity('template_new', key)
    return redirect(url_for('template_edit', key=key) if key else url_for('templates_admin'))

@app.route('/dashboard/template/<key>', methods=['GET', 'POST'])
def template_edit(key):
    if not is_admin():
        return redirect(url_for('login'))
    tpl = MonitorTemplate.query.get(key)
    if not tpl:
        return redirect(url_for('templates_admin'))
    if request.method == 'POST':
        tpl.name = request.form.get('name', tpl.name).strip()
        tpl.icon = request.form.get('icon', tpl.icon).strip() or tpl.icon
        tpl.frequency = request.form.get('frequency', tpl.frequency)
        def fnum(v):
            try:
                return float(v) if v not in (None, '') else None
            except (ValueError, TypeError):
                return None
        # update existing params
        for p in list(tpl.params):
            if request.form.get('del_%d' % p.id):
                db.session.delete(p); continue
            p.label = request.form.get('label_%d' % p.id, p.label).strip()
            p.unit = request.form.get('unit_%d' % p.id, p.unit).strip()
            p.min_v = fnum(request.form.get('min_%d' % p.id))
            p.max_v = fnum(request.form.get('max_%d' % p.id))
            p.action_low = request.form.get('alow_%d' % p.id, '').strip()
            p.action_high = request.form.get('ahigh_%d' % p.id, '').strip()
            p.periodic = bool(request.form.get('per_%d' % p.id))
        # new param
        npk = request.form.get('new_pkey', '').strip().lower()
        npk = ''.join(ch for ch in npk if ch.isalnum() or ch == '_')[:40]
        if npk and request.form.get('new_label', '').strip():
            db.session.add(MonitorParam(template_key=tpl.key, pkey=npk,
                           label=request.form.get('new_label').strip(),
                           unit=request.form.get('new_unit', '').strip(),
                           min_v=fnum(request.form.get('new_min')), max_v=fnum(request.form.get('new_max')),
                           action_low=request.form.get('new_alow', '').strip(),
                           action_high=request.form.get('new_ahigh', '').strip(),
                           periodic=bool(request.form.get('new_per')), sort=len(tpl.params) + 1))
        db.session.commit(); log_activity('template_edit', key)
        return redirect(url_for('template_edit', key=key) + '?saved=1')
    return render_template('template_edit.html', tpl=tpl, freq_label=FREQ_LABEL)

# ── Demo data seeder (admin) — τυχαίες καταγραφές 1-31/5 ──
@app.route('/admin/seed-demo')
def seed_demo():
    if not is_admin():
        return redirect(url_for('login'))
    import random
    user = current_user()
    pools = Pool.query.filter_by(is_active=True).all()
    if not pools:
        return 'Δεν υπάρχουν πισίνες. <a href="/pools/dashboard">Πίσω</a>'
    yr = int(request.args.get('year', date.today().year))
    created = 0
    for p in pools:
        d = date(yr, 5, 1)
        while d <= date(yr, 5, 31):
            for period in ('morning', 'afternoon'):
                if PoolRecord.query.filter_by(pool_id=p.id, record_date=d, period=period).first():
                    continue
                r = PoolRecord(pool_id=p.id, user_id=user.id, record_date=d, period=period)
                r.free_chlorine     = round(random.uniform(0.5, 1.3), 2) if random.random() > 0.10 else round(random.uniform(0.1, 0.39), 2)
                r.combined_chlorine = round(random.uniform(0.0, 0.4), 2) if random.random() > 0.10 else round(random.uniform(0.55, 0.9), 2)
                r.ph                = round(random.uniform(7.2, 7.8), 1) if random.random() > 0.10 else round(random.uniform(7.9, 8.3), 1)
                r.temp              = round(random.uniform(24, 30), 1)
                r.turbidity         = round(random.uniform(0.1, 0.8), 1) if random.random() > 0.10 else round(random.uniform(1.1, 1.8), 1)
                if period == 'morning':
                    r.cyanuric_acid    = round(random.uniform(20, 60))
                    r.total_alkalinity = round(random.uniform(80, 120)) if random.random() > 0.10 else round(random.uniform(60, 79))
                    r.orp              = round(random.uniform(650, 760)) if random.random() > 0.10 else round(random.uniform(580, 640))
                r.backwash_done = random.random() > 0.85
                r.recorded_at = datetime(yr, 5, d.day, 8 if period == 'morning' else 17, random.randint(0, 59))
                db.session.add(r); created += 1
            d += timedelta(days=1)
    db.session.commit()
    log_activity('seed_demo', f'{created} records {yr}-05')
    return ('<div style="font-family:Arial;padding:40px;text-align:center;">'
            '<h2 style="color:#193847;">Δημιουργήθηκαν ' + str(created) + ' demo καταγραφές (1-31/5/' + str(yr) + ')</h2>'
            '<p><a href="/pools/dashboard" style="color:#193847;">→ Pool dashboard</a> &nbsp; '
            '<a href="/pools/coverage" style="color:#193847;">→ Κάλυψη</a></p></div>')

# v12.7 — Διαχείριση Ξενοδοχείων & Πισινών (μεταφέρθηκε από τον Πίνακα Πισινών)
@app.route('/dashboard/hotels')
def hotels_admin():
    if not is_admin():
        return redirect(url_for('login'))
    hotels = Hotel.query.filter_by(is_active=True).order_by(Hotel.name).all()
    pools  = Pool.query.filter_by(is_active=True).all()
    return render_template('hotels_admin.html', hotels=hotels, pools=pools, is_admin=True)

@app.route('/dashboard/add-hotel', methods=['POST'])
def add_hotel():
    if not is_admin():
        return redirect(url_for('login'))
    name = request.form.get('name', '').strip()
    if name and not Hotel.query.filter_by(name=name).first():
        db.session.add(Hotel(name=name))
        db.session.commit()
    return redirect(url_for('hotels_admin') + '?success=hotel_added')

@app.route('/dashboard/delete-hotel/<int:hotel_id>', methods=['POST'])
def delete_hotel(hotel_id):
    if not is_admin():
        return redirect(url_for('login'))
    hotel = Hotel.query.get(hotel_id)
    if hotel:
        hotel.is_active = False
        for p in hotel.pools:
            p.is_active = False
        db.session.commit()
    return redirect(url_for('hotels_admin'))

@app.route('/dashboard/add-pool', methods=['POST'])
def add_pool():
    if not is_admin():
        return redirect(url_for('login'))
    data = request.form
    hotel_id = data.get('hotel_id')
    name = data.get('name', '').strip()
    if hotel_id and name:
        db.session.add(Pool(
            hotel_id=int(hotel_id),
            name=name,
            location=data.get('location', '').strip(),
            pool_type=data.get('pool_type', 'pool'),
            volume_m3=flt(data, 'volume_m3')
        ))
        db.session.commit()
    return redirect(url_for('hotels_admin') + '?success=pool_added')

@app.route('/dashboard/delete-pool/<int:pool_id>', methods=['POST'])
def delete_pool(pool_id):
    if not is_admin():
        return redirect(url_for('login'))
    pool = Pool.query.get(pool_id)
    if pool:
        pool.is_active = False
        db.session.commit()
    return redirect(url_for('hotels_admin'))


# ── v12.3 — Διαχείριση Δικτύων Νερού (mirror add/delete pool) ──
@app.route('/dashboard/add-water-system', methods=['POST'])
def add_water_system():
    if not is_admin():
        return redirect(url_for('login'))
    data = request.form
    hotel_id = data.get('hotel_id')
    name = (data.get('name') or '').strip()
    if hotel_id and name:
        db.session.add(WaterSystem(hotel_id=int(hotel_id), name=name,
                                   location=(data.get('location') or '').strip()))
        db.session.commit()
        log_activity('water_system_add', name)
    return redirect(url_for('dashboard') + '?success=system_added')

@app.route('/dashboard/edit-water-system/<int:system_id>', methods=['POST'])
def edit_water_system(system_id):
    if not is_admin():
        return redirect(url_for('login'))
    ws = WaterSystem.query.get(system_id)
    if ws:
        name = (request.form.get('name') or '').strip()
        if name:
            ws.name = name
        ws.location = (request.form.get('location') or '').strip()
        db.session.commit()
        log_activity('water_system_edit', ws.name)
    return redirect(url_for('dashboard') + '?success=system_edited')

@app.route('/dashboard/delete-water-system/<int:system_id>', methods=['POST'])
def delete_water_system(system_id):
    if not is_admin():
        return redirect(url_for('login'))
    ws = WaterSystem.query.get(system_id)
    if ws:
        ws.is_active = False   # soft delete — οι καταγραφές μένουν
        db.session.commit()
        log_activity('water_system_delete', ws.name)
    return redirect(url_for('dashboard'))


# ──────────────────────────────────────────────────────────────────────────
#  CHART APIs
# ──────────────────────────────────────────────────────────────────────────
@app.route('/api/history')
def api_history():
    if 'user_id' not in session:
        return jsonify([])
    q = WaterRecord.query.filter_by(period='morning')
    sid = request.args.get('system_id', type=int)   # v12.3 — προαιρετικό φίλτρο δικτύου
    if sid:
        q = q.filter_by(water_system_id=sid)
    records = q.order_by(WaterRecord.record_date.desc()).limit(14).all()
    return jsonify([{
        'date': r.record_date.strftime('%d/%m'),
        'clo2_tank': r.clo2_tank,
        'temp_dhw_out': r.temp_dhw_out,
        'temp_dhw_return': r.temp_dhw_return,
        'temp_tank': r.temp_tank,
    } for r in records])

@app.route('/api/pool-history/<int:pool_id>')
def api_pool_history(pool_id):
    if 'user_id' not in session:
        return jsonify([])
    records = (PoolRecord.query.filter_by(pool_id=pool_id, period='morning')
               .order_by(PoolRecord.record_date.desc()).limit(14).all())
    return jsonify([{
        'date': r.record_date.strftime('%d/%m'),
        'free_chlorine': r.free_chlorine,
        'ph': r.ph,
        'temp': r.temp,
    } for r in records])


# ──────────────────────────────────────────────────────────────────────────
#  INIT  (τρέχει και κάτω από gunicorn)
# ──────────────────────────────────────────────────────────────────────────
def _add_col(table, col, ddl):
    """v12.35 — Ανθεκτικό, idempotent & race-safe (gunicorn 2 workers).
    Postgres: ADD COLUMN IF NOT EXISTS. SQLite: inspect-check.
    Κάθε στήλη έχει ΔΙΚΟ της try/except — μία αποτυχία ΔΕΝ μπλοκάρει τις υπόλοιπες."""
    try:
        if db.engine.dialect.name == 'postgresql':
            db.session.execute(text(f'ALTER TABLE "{table}" ADD COLUMN IF NOT EXISTS {ddl}'))
            db.session.commit()
        else:
            insp = db.inspect(db.engine)
            if table in insp.get_table_names():
                cols = [c['name'] for c in insp.get_columns(table)]
                if col not in cols:
                    db.session.execute(text(f'ALTER TABLE "{table}" ADD COLUMN {ddl}'))
                    db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f'_add_col {table}.{col} skipped: {e}')

def ensure_columns():
    """Ελαφρύ migration: πρόσθεσε νέες στήλες σε ήδη υπάρχουσα βάση.
    v12.35: κάθε _add_col αυτόνομο (δικό του commit/rollback) → καμία στήλη δεν παρασύρει τις άλλες."""
    truth = 'true' if db.engine.dialect.name == 'postgresql' else '1'
    _add_col('pool', 'volume_m3', 'volume_m3 FLOAT')
    _add_col('user', 'email', 'email VARCHAR(120)')
    _add_col('user', 'phone', 'phone VARCHAR(40)')
    _add_col('user', 'approved', f'approved BOOLEAN DEFAULT {truth}')
    _add_col('user', 'avatar', 'avatar TEXT')
    _add_col('water_record', 'water_system_id', 'water_system_id INTEGER')  # v12.3
    _add_col('activity_log', 'hotel_id', 'hotel_id INTEGER')  # v12.22
    _add_col('survey_response', 'import_hash', 'import_hash VARCHAR(40)')  # v12.30
    _add_col('user', 'dashboard', 'dashboard TEXT')  # v12.34 — προσωπική διάταξη tiles (JSON)
    # v12.40 — Module Πρόγραμμα Εργασίας (User columns ΠΡΙΝ το init_db seed)
    _add_col('user', 'department_id', 'department_id INTEGER')
    _add_col('user', 'employer', 'employer VARCHAR(120)')
    _add_col('user', 'subunit', 'subunit VARCHAR(20)')
    _add_col('user', 'home_hotel_id', 'home_hotel_id INTEGER')
    _add_col('user', 'login_enabled', 'login_enabled BOOLEAN')
    _add_col('user', 'employment_active', 'employment_active BOOLEAN')

def init_db():
    with app.app_context():
        db.create_all()
        ensure_columns()
        try:
            # v12.25 — Bootstrap πρωτου masteradmin ΜΟΝΟ σε αδεια βαση, με credentials απο env.
            # Δεν υπαρχουν πλεον demo accounts με σταθερους κωδικους μεσα στον κωδικα.
            if User.query.count() == 0:
                _admin_user = os.environ.get('BOOTSTRAP_ADMIN_USER', 'admin')
                _admin_pw   = os.environ.get('BOOTSTRAP_ADMIN_PASSWORD')
                if not _admin_pw:
                    _admin_pw = secrets.token_urlsafe(12)
                    print(f'[ESTIA] Αρχικος masteradmin "{_admin_user}" με ΤΥΧΑΙΟ κωδικο: {_admin_pw}')
                    print('[ESTIA] Αλλαξε τον κωδικο αμεσως μετα την πρωτη συνδεση.')
                db.session.add(User(username=_admin_user, password=generate_password_hash(_admin_pw),
                                    full_name='Διαχειριστης', role='masteradmin', language='el',
                                    approved=True, is_active=True))
                db.session.commit()
                print('Βαση δεδομενων και αρχικος διαχειριστης δημιουργηθηκαν')
            if not Hotel.query.first():
                sergios = Hotel(name='Sergios Hotel')
                db.session.add(sergios)
                db.session.flush()
                db.session.add(Pool(hotel_id=sergios.id, name='Κύρια Πισίνα', location='Pool bar', pool_type='pool', volume_m3=200))
                db.session.add(Pool(hotel_id=sergios.id, name='Παιδική Πισίνα', location='Pool bar', pool_type='kids', volume_m3=20))
                db.session.commit()
                print('Δημιουργηθηκε δειγμα ξενοδοχειου & πισινων')

            # v12.3 — default Δίκτυο Νερού + backfill παλιών records (idempotent, ασφαλές για υπάρχουσα βάση)
            try:
                _h0 = Hotel.query.first()
                if _h0 and not WaterSystem.query.first():
                    db.session.add(WaterSystem(hotel_id=_h0.id, name='Κεντρικό Δίκτυο', location='Μηχανοστάσιο'))
                    db.session.commit()
                _ws0 = WaterSystem.query.first()
                if _ws0:
                    WaterRecord.query.filter_by(water_system_id=None).update({'water_system_id': _ws0.id})
                    db.session.commit()
            except Exception as _e:
                db.session.rollback()
                print(f'water_system seed skipped: {_e}')

            if not MonitorTemplate.query.first():
                tank = MonitorTemplate(key='tank', name='Στάθμες & Δεξαμενές', icon='ti-stack-2', frequency='daily', sort=10)
                db.session.add(tank); db.session.flush()
                db.session.add_all([
                    MonitorParam(template_key='tank', pkey='level_pct', label='Στάθμη', unit='%', min_v=20, max_v=100, action_low='Χαμηλή στάθμη — έλεγξε πλήρωση/διαρροή.', sort=1),
                    MonitorParam(template_key='tank', pkey='pressure', label='Πίεση', unit='bar', sort=2),
                    MonitorParam(template_key='tank', pkey='temp', label='Θερμοκρασία', unit='C', sort=3),
                ])
                energy = MonitorTemplate(key='energy', name='Ενέργεια & Μετρητές', icon='ti-bolt', frequency='daily', sort=20)
                db.session.add(energy); db.session.flush()
                db.session.add_all([
                    MonitorParam(template_key='energy', pkey='electricity_kwh', label='Ρεύμα (ένδειξη)', unit='kWh', sort=1),
                    MonitorParam(template_key='energy', pkey='water_m3', label='Νερό (ένδειξη)', unit='m3', sort=2),
                    MonitorParam(template_key='energy', pkey='gas_m3', label='Φυσικό αέριο (ένδειξη)', unit='m3', sort=3),
                ])
                db.session.commit()
                print('Δημιουργηθηκαν generic templates')

            if not Area.query.first():
                _h = Hotel.query.first()
                if _h:
                    db.session.add(Area(hotel_id=_h.id, template_key='tank', name='Κεντρική Δεξαμενή', location='Μηχανοστάσιο'))
                    db.session.add(Area(hotel_id=_h.id, template_key='energy', name='Γενικός Μετρητής', location='Πίνακας ΔΕΗ'))
                    db.session.commit()
                    print('Δημιουργηθηκαν δειγματικα areas')
        except Exception as e:
            db.session.rollback()
            print(f'init_db seed skipped: {e}')


def _athens_now():
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo('Europe/Athens'))
    except Exception:
        return datetime.utcnow() + timedelta(hours=3)

def missing_today():
    today = date.today()
    miss = []
    for p in Pool.query.filter_by(is_active=True).all():
        recs = {r.period for r in PoolRecord.query.filter_by(pool_id=p.id, record_date=today).all()}
        gaps = [per for per in ('morning', 'afternoon') if per not in recs]
        if gaps:
            hotel = p.hotel.name if p.hotel else ''
            miss.append(f"{hotel} — {p.name}: " + ', '.join('Πρωί' if g == 'morning' else 'Απόγευμα' for g in gaps))
    # v12.3 — έλεγχος ανά δίκτυο νερού (όχι πια ενιαίο/Sergios)
    for ws in WaterSystem.query.filter_by(is_active=True).all():
        recs = {r.period for r in WaterRecord.query.filter_by(water_system_id=ws.id, record_date=today).all()}
        gaps = [per for per in ('morning', 'afternoon') if per not in recs]
        if gaps:
            hotel = ws.hotel.name if ws.hotel else ''
            miss.append(f"Νερά Χρήσης — {hotel} · {ws.name}: " + ', '.join('Πρωί' if g == 'morning' else 'Απόγευμα' for g in gaps))
    return miss

def send_reminder_email(miss):
    if (not EMAIL_PASSWORD and not GRAPH_CLIENT_ID) or not miss:
        return False
    recips = list(EMAIL_TO_LIST)
    for u in User.query.filter_by(is_active=True, approved=True).all():
        if u.email and u.email not in recips:
            recips.append(u.email)
    items = ''.join(f'<li>{m}</li>' for m in miss)
    html = f'<div style="font-family:Arial,sans-serif"><h3 style="color:#193847">Εκκρεμείς καταγραφές σήμερα</h3><ul>{items}</ul><p style="color:#888;font-size:12px">Εστία — CONDIAN HOTELS · αυτόματη υπενθύμιση</p></div>'
    return send_email('Εστία — Υπενθύμιση καταγραφών - ' + date.today().strftime('%d/%m/%Y'), html, recips)

def reminder_tick():
    with app.app_context():
        now = _athens_now()
        if now.hour != REMINDER_HOUR:
            return
        today = now.strftime('%Y-%m-%d')
        if ReminderSent.query.get(today):
            return
        try:
            db.session.add(ReminderSent(day=today)); db.session.commit()
        except Exception:
            db.session.rollback(); return   # άλλος worker το ανέλαβε
        miss = missing_today()
        if miss:
            send_reminder_email(miss)
            print(f'[reminder] {today}: {len(miss)} εκκρεμή')

def reminder_loop():
    while True:
        try:
            reminder_tick()
        except Exception as e:
            print('[reminder] loop error:', e)
        time.sleep(1800)   # κάθε 30 λεπτά

def start_scheduler():
    if ENABLE_SCHEDULER:
        threading.Thread(target=reminder_loop, daemon=True).start()
        print('[scheduler] reminder loop started')


def seed_team():
    """Μία φορά: δημιουργεί την πραγματική ομάδα CONDIAN."""
    with app.app_context():
        try:
            if Setting.query.get('seeded_team_v1'):
                return
            team = [
                ('giakoumakis',  'Giakoumakis Giannis',    'masteradmin', 'g.giakoumakis@condianhotels.gr', '+306973728931'),
                ('giannoulakis', 'Γιαννουλάκης Δημήτρης',  'admin',       'dimitris@condianhotels.gr',      '+306936647778'),
                ('xypakis',      'Ξυπάκης Μάνος',          'manager',     'm.xypakis@condianhotels.gr',     '+306972238222'),
                ('smyrnakis',    'Σμυρνάκης Χριστόφορος',  'manager',     'c.smyrnakis@condianhotels.gr',   '+306992015939'),
                ('flouris',      'Φλουρής Στέφανος',       'manager',     's.flouris@condianhotels.gr',     '+306931549656'),
            ]
            # v12.25 — ο αρχικος κωδικος ομαδας ερχεται απο env (fallback για συμβατοτητα).
            # Συσταση: ορισε TEAM_DEFAULT_PASSWORD στο Railway και ζητα αλλαγη στην 1η συνδεση.
            _team_pw = os.environ.get('TEAM_DEFAULT_PASSWORD', 'condian2026')
            for un, fn, role, em, ph in team:
                u = User.query.filter_by(username=un).first()
                if u:
                    u.full_name = fn; u.role = role; u.email = em; u.phone = ph
                    u.approved = True; u.is_active = True
                else:
                    db.session.add(User(username=un, password=generate_password_hash(_team_pw),
                                        full_name=fn, role=role, email=em, phone=ph,
                                        approved=True, is_active=True, language='el'))
            db.session.add(Setting(key='seeded_team_v1', value='1'))
            db.session.commit()
            print('Team CONDIAN seeded')
        except Exception as e:
            db.session.rollback(); print('seed_team skipped:', e)


import faults          # v12.14 — Module Βλαβοληψία (μοντέλα + routes)· ΠΡΙΝ το create_all
import surveys         # v12.23 — Module Ερωτηματολόγια (μοντέλα + routes)· ΠΡΙΝ το create_all
import imports         # v12.29 — Κέντρο Εισαγωγής Δεδομένων (μοντέλο ImportUpload + routes)· ΠΡΙΝ το create_all
import backup          # v12.31 — Module Αντιγράφων Ασφαλείας (BackupLog + routes)· ΠΡΙΝ το create_all
import schedule        # v12.40 — Module Πρόγραμμα Εργασίας (μοντέλα + routes)· ΠΡΙΝ το create_all
import extras          # v12.43 — per-role μενού + Feedback· ΠΡΙΝ το create_all
import payroll         # v12.47 — Module Μισθοδοσία Φ1 (μοντέλα + routes)· ΠΡΙΝ το create_all
init_db()
backup.ensure_backup_columns()   # v12.33 — auto-migration στηλών backup_log + seed ρυθμίσεων
seed_team()
faults.seed_faults()   # seed κατηγορίες/ειδικότητες/SLA (idempotent)
surveys.seed_surveys() # seed δείγμα ερωτηματολογίου (idempotent)
schedule.ensure_schedule_columns()  # v12.40
schedule.seed_schedule()            # v12.40 (idempotent)
payroll.ensure_payroll_columns()    # v12.47 — Hotel.company_id
payroll.seed_payroll()              # v12.47 (idempotent)
start_scheduler()
backup.start_backup_scheduler()  # v12.31 — ημερήσιο backup -> SharePoint (αν BACKUP_ENABLED)

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
