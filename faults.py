# -*- coding: utf-8 -*-
"""
Εστία — Module Βλαβοληψία (v12.14, Φάση 1)
==========================================
Plug-in: γίνεται `import faults` από το ΤΕΛΟΣ του app.py (αφού οριστούν app/db/helpers,
πριν το init_db() ώστε το create_all να πιάσει τους νέους πίνακες).
Πηγή προδιαγραφών: 02_ΒΛΑΒΟΛΗΨΙΑ_MODULE/.
"""
import os, csv
from datetime import datetime
import base64, threading
from flask import request, redirect, url_for, render_template, session, jsonify
from app import (app, db, current_user, is_admin, allowed_hotels, notify, notify_admins,
                 log_activity, Hotel, User, Pool, Setting, ROLE_RANK, role_rank, BASE_DIR,
                 send_email, EMAIL_TO_LIST)

# ── Σταθερές ─────────────────────────────────────────────────────────────────
PRIORITIES = ('Υψηλή', 'Κανονική', 'Χαμηλή')
STATUSES   = ('pending_assign', 'auto_assigned', 'assigned', 'in_progress',
              'paused', 'done', 'not_done', 'winter', 'resubmitted')
STATUS_LABELS = {
    'pending_assign': 'Αναμένει ανάθεση', 'auto_assigned': 'Αυτόματη ανάθεση',
    'assigned': 'Ανατέθηκε', 'in_progress': 'Προς διεκπεραίωση', 'paused': 'Σε παύση',
    'done': 'Ολοκληρώθηκε', 'not_done': 'Δεν έγινε', 'winter': 'Για χειμώνα',
    'resubmitted': 'Υποβλήθηκε ξανά',
}
STATUS_COLOR = {  # bg / text (light· dark μέσω estia-theme badges)
    'pending_assign': ('#f1f5f9', '#64748b'), 'auto_assigned': ('#e0f2fe', '#0369a1'),
    'assigned': ('#dbeafe', '#185FA5'), 'in_progress': ('#fef3c7', '#b45309'),
    'paused': ('#e2e8f0', '#475569'), 'done': ('#dcfce7', '#16a34a'),
    'not_done': ('#fee2e2', '#dc2626'), 'winter': ('#cffafe', '#0e7490'),
    'resubmitted': ('#f3e8ff', '#7e22ce'),
}
PRIORITY_COLOR = {'Υψηλή': ('#fee2e2', '#dc2626'), 'Κανονική': ('#fef3c7', '#b45309'), 'Χαμηλή': ('#f1f5f9', '#64748b')}
TERMINAL = ('done', 'not_done', 'resubmitted')
SPECIALTIES = ('Υδραυλικός', 'Ηλεκτρολόγος', 'Ψυκτικός', 'Κηπουρός',
               'Ελαιοχρωματιστής', 'Συντηρητής', 'Εξωτερικός Συνεργάτης')
TRANSITIONS = {
    'pending_assign': ('auto_assigned', 'assigned'),
    'auto_assigned':  ('assigned', 'in_progress', 'done', 'not_done', 'winter'),
    'assigned':       ('in_progress', 'paused', 'done', 'not_done', 'winter'),
    'in_progress':    ('paused', 'done', 'not_done', 'winter', 'resubmitted'),
    'paused':         ('in_progress', 'done', 'not_done'),
    'winter':         ('pending_assign', 'in_progress'),
    'done':           ('resubmitted',),
    'not_done':       ('resubmitted',),
    'resubmitted':    (),
}
HOTEL_PREFIX = {
    'Asterias Village Resort': 'AST', 'Sergios Hotel': 'SRG',
    'Central Hersonissos Hotel': 'CNT', 'Piskopiano Village': 'PSV', 'Iro Hotel': 'IRO',
}
# χάρτης level-1 κλάδου → ειδικότητα (αρχικό· admin-editable στη Φάση 2)
ROOT_SPECIALTY = {
    'Α. ΣΥΝΤΗΡΗΣΗ': 'Συντηρητής', 'Β. ΗΛΕΚΤΡΟΛΟΓΙΚΑ': 'Ηλεκτρολόγος',
    'C. ΥΔΡΑΥΛΙΚΑ': 'Υδραυλικός', 'D. ΕΠΙΤΡΑΠΕΖΙΟΣ ΕΞΟΠΛΙΣΜΟΣ': 'Ψυκτικός',
    'E. ΕΞΩΤΕΡΙΚΟΙ ΣΥΝΕΡΓΑΤΕΣ': 'Εξωτερικός Συνεργάτης',
}
SLA_DEFAULTS = {'Υψηλή': 120, 'Κανονική': 360, 'Χαμηλή': 1440}

# ── Μοντέλα ──────────────────────────────────────────────────────────────────
class FaultCategory(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    parent_id = db.Column(db.Integer, db.ForeignKey('fault_category.id'), nullable=True)
    hotel_id  = db.Column(db.Integer, db.ForeignKey('hotel.id'), nullable=True)
    name      = db.Column(db.String(160), nullable=False)
    level     = db.Column(db.Integer, default=1)
    is_active = db.Column(db.Boolean, default=True)
    sort      = db.Column(db.Integer, default=0)
    children  = db.relationship('FaultCategory', backref=db.backref('parent', remote_side=[id]))

class FaultLocation(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    parent_id = db.Column(db.Integer, db.ForeignKey('fault_location.id'), nullable=True)
    hotel_id  = db.Column(db.Integer, db.ForeignKey('hotel.id'), nullable=False)
    name      = db.Column(db.String(120), nullable=False)
    kind      = db.Column(db.String(12), default='χώρος')
    is_active = db.Column(db.Boolean, default=True)
    children  = db.relationship('FaultLocation', backref=db.backref('parent', remote_side=[id]))

class Specialty(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    name      = db.Column(db.String(40), unique=True, nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    sort      = db.Column(db.Integer, default=0)

class CategorySpecialty(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    category_id = db.Column(db.Integer, db.ForeignKey('fault_category.id'), nullable=False)
    specialty   = db.Column(db.String(40), nullable=False)

class UserSpecialty(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    user_id   = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    specialty = db.Column(db.String(40), nullable=False)

class Fault(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    code          = db.Column(db.String(24), unique=True, index=True)
    type          = db.Column(db.String(12), default='βλάβη')
    hotel_id      = db.Column(db.Integer, db.ForeignKey('hotel.id'), nullable=False)
    pool_id       = db.Column(db.Integer, db.ForeignKey('pool.id'), nullable=True)
    location_id   = db.Column(db.Integer, db.ForeignKey('fault_location.id'), nullable=True)
    room_id       = db.Column(db.Integer, db.ForeignKey('fault_location.id'), nullable=True)
    geo_lat       = db.Column(db.Float, nullable=True)
    geo_lng       = db.Column(db.Float, nullable=True)
    category_id   = db.Column(db.Integer, db.ForeignKey('fault_category.id'), nullable=True)
    description   = db.Column(db.Text, nullable=False)
    priority      = db.Column(db.String(10), default='Κανονική')
    tag           = db.Column(db.String(40), nullable=True)
    due_at        = db.Column(db.DateTime, nullable=True)
    status        = db.Column(db.String(16), default='pending_assign')
    source        = db.Column(db.String(20), default='Ενδοξενοδοχειακά')
    submitted_by      = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    assigned_user_id  = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    completed_by      = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    cover_image   = db.Column(db.Text, nullable=True)
    submitted_at  = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at    = db.Column(db.DateTime, nullable=True)
    completed_at  = db.Column(db.DateTime, nullable=True)
    resolution_seconds = db.Column(db.Integer, nullable=True)
    imported_from = db.Column(db.String(30), nullable=True)
    legacy_from = db.Column(db.String(120)); legacy_assignee = db.Column(db.String(120))
    legacy_completed_by = db.Column(db.String(120)); legacy_category = db.Column(db.String(200))
    legacy_location = db.Column(db.String(200)); legacy_room = db.Column(db.String(80))
    category  = db.relationship('FaultCategory')
    hotel     = db.relationship('Hotel')
    submitter = db.relationship('User', foreign_keys=[submitted_by])
    assignee  = db.relationship('User', foreign_keys=[assigned_user_id])
    def can_transition(self, to):
        return to in TRANSITIONS.get(self.status, ())

class FaultCandidate(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    fault_id  = db.Column(db.Integer, db.ForeignKey('fault.id'), nullable=False)
    user_id   = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    specialty = db.Column(db.String(40), nullable=True)
    user      = db.relationship('User')

class FaultChangeLog(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    fault_id    = db.Column(db.Integer, db.ForeignKey('fault.id'), nullable=False)
    change_type = db.Column(db.String(12))
    field       = db.Column(db.String(40))
    from_value  = db.Column(db.String(200))
    to_value    = db.Column(db.String(200))
    by_user_id  = db.Column(db.Integer, db.ForeignKey('user.id'))
    at          = db.Column(db.DateTime, default=datetime.utcnow)
    by_user     = db.relationship('User')

class FaultComment(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    fault_id   = db.Column(db.Integer, db.ForeignKey('fault.id'), nullable=False)
    audience   = db.Column(db.String(12), default='διοίκηση')
    text       = db.Column(db.Text)
    file_url   = db.Column(db.Text, nullable=True)
    by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    at         = db.Column(db.DateTime, default=datetime.utcnow)
    by_user    = db.relationship('User')

class FaultAttachment(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    fault_id   = db.Column(db.Integer, db.ForeignKey('fault.id'), nullable=False)
    url        = db.Column(db.Text, nullable=False)
    by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    at         = db.Column(db.DateTime, default=datetime.utcnow)

class FaultTag(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    name      = db.Column(db.String(40), unique=True, nullable=False)
    color     = db.Column(db.String(10), default='#b91c1c')
    is_active = db.Column(db.Boolean, default=True)

class SLATarget(db.Model):
    id      = db.Column(db.Integer, primary_key=True)
    scope   = db.Column(db.String(10), default='priority')
    key     = db.Column(db.String(40), nullable=False)
    minutes = db.Column(db.Integer, nullable=False)

# ── Helpers ──────────────────────────────────────────────────────────────────
def gen_code(hotel):
    year = datetime.utcnow().year
    prefix = HOTEL_PREFIX.get(hotel.name, 'GEN')
    last = (Fault.query.filter(Fault.code.like('%s-%d-%%' % (prefix, year)))
                       .order_by(Fault.id.desc()).first())
    try:
        seq = (int(last.code.split('-')[2]) + 1) if last else 1
    except Exception:
        seq = (Fault.query.filter(Fault.code.like('%s-%d-%%' % (prefix, year))).count()) + 1
    return '%s-%d-%06d' % (prefix, year, seq)

def log_change(fault, change_type, field, frm, to, user_id):
    db.session.add(FaultChangeLog(fault_id=fault.id, change_type=change_type, field=field,
                                  from_value=str(frm)[:200], to_value=str(to)[:200], by_user_id=user_id))

def cat_root_name(cat):
    seen = 0
    while cat and cat.parent_id and seen < 6:
        cat = FaultCategory.query.get(cat.parent_id); seen += 1
    return cat.name if cat else None

def user_specialty_names(user):
    return [us.specialty for us in UserSpecialty.query.filter_by(user_id=user.id).all()]

def auto_assign(fault):
    """Γέμισε δεξαμενή υποψηφίων από κατηγορία→ειδικότητα (ίδιο ξενοδοχείο)."""
    specs = [cs.specialty for cs in CategorySpecialty.query.filter_by(category_id=fault.category_id).all()]
    if not specs and fault.category_id:                       # δοκίμασε τον level-1 κλάδο
        root = cat_root_name(FaultCategory.query.get(fault.category_id))
        sp = ROOT_SPECIALTY.get(root)
        if sp:
            specs = [sp]
    if not specs:
        fault.status = 'pending_assign'
        return
    fault.status = 'auto_assigned'
    seen = set()
    for sp in specs:
        if sp in seen:
            continue
        seen.add(sp)
        db.session.add(FaultCandidate(fault_id=fault.id, specialty=sp))
        for us in UserSpecialty.query.filter_by(specialty=sp).all():
            u = User.query.get(us.user_id)
            if u and u.is_active:
                db.session.add(FaultCandidate(fault_id=fault.id, user_id=u.id))

def visible_faults_query(user):
    hids = [h.id for h in allowed_hotels(user)]
    q = Fault.query.filter(Fault.hotel_id.in_(hids or [-1]))
    if not is_admin():
        my_specs = user_specialty_names(user)
        pool_ids = [c.fault_id for c in FaultCandidate.query.filter(
            (FaultCandidate.user_id == user.id) |
            (FaultCandidate.specialty.in_(my_specs or ['__none__']))).all()]
        q = q.filter((Fault.submitted_by == user.id) | (Fault.assigned_user_id == user.id) |
                     (Fault.id.in_(pool_ids or [-1])))
    return q

def can_view(user, f):
    if f.hotel_id not in [h.id for h in allowed_hotels(user)]:
        return False
    if is_admin():
        return True
    if f.submitted_by == user.id or f.assigned_user_id == user.id:
        return True
    my = set(user_specialty_names(user))
    for c in FaultCandidate.query.filter_by(fault_id=f.id).all():
        if c.user_id == user.id or (c.specialty and c.specialty in my):
            return True
    return False

# ── Seed (idempotent) ────────────────────────────────────────────────────────
def seed_faults():
    with app.app_context():
        try:
            if Setting.query.get('seeded_faults_v1'):
                return
            # ειδικότητες
            for i, sp in enumerate(SPECIALTIES):
                if not Specialty.query.filter_by(name=sp).first():
                    db.session.add(Specialty(name=sp, sort=i))
            # SLA
            for k, m in SLA_DEFAULTS.items():
                if not SLATarget.query.filter_by(scope='priority', key=k).first():
                    db.session.add(SLATarget(scope='priority', key=k, minutes=m))
            if not Setting.query.get('fault_stale_days'):
                db.session.add(Setting(key='fault_stale_days', value='60'))
            db.session.commit()
            # κατηγορίες από CSV (δέντρο 3 επιπέδων, dedup)
            path = os.path.join(BASE_DIR, 'seed', 'fault_categories.csv')
            if os.path.exists(path) and FaultCategory.query.count() == 0:
                roots = {}; mids = {}
                with open(path, encoding='utf-8') as fh:
                    rows = list(csv.DictReader(fh))
                for r in rows:
                    l1 = (r.get('level1') or '').strip()
                    l2 = (r.get('level2') or '').strip()
                    l3 = (r.get('level3') or '').strip()
                    try: occ = int(r.get('occurrences') or 0)
                    except Exception: occ = 0
                    if not l1:
                        continue
                    if l1 not in roots:
                        c = FaultCategory(name=l1, level=1, sort=0)
                        db.session.add(c); db.session.flush(); roots[l1] = c
                    if l2:
                        key2 = (l1, l2)
                        if key2 not in mids:
                            c2 = FaultCategory(name=l2, level=2, parent_id=roots[l1].id, sort=-occ)
                            db.session.add(c2); db.session.flush(); mids[key2] = c2
                        if l3:
                            c3 = FaultCategory(name=l3, level=3, parent_id=mids[key2].id, sort=-occ)
                            db.session.add(c3)
                db.session.commit()
                # χάρτης κατηγορία(level-1)→ειδικότητα
                for name, sp in ROOT_SPECIALTY.items():
                    root = FaultCategory.query.filter_by(name=name, level=1).first()
                    if root and not CategorySpecialty.query.filter_by(category_id=root.id).first():
                        db.session.add(CategorySpecialty(category_id=root.id, specialty=sp))
                db.session.commit()
            db.session.add(Setting(key='seeded_faults_v1', value='1'))
            db.session.commit()
            print('[faults] seeded specialties/SLA/categories (%d)' % FaultCategory.query.count())
        except Exception as e:
            db.session.rollback(); print('[faults] seed skipped:', e)

# ── Routes ───────────────────────────────────────────────────────────────────
def _ctx_lists(user):
    hotels = allowed_hotels(user)
    cats = FaultCategory.query.filter_by(is_active=True).order_by(FaultCategory.level, FaultCategory.sort, FaultCategory.name).all()
    return hotels, cats

@app.route('/fault', methods=['GET', 'POST'])
def fault_report():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = current_user()
    hotels, cats = _ctx_lists(user)
    if request.method == 'POST':
        desc = (request.form.get('message') or request.form.get('description') or '').strip()
        hid = request.form.get('hotel_id', type=int)
        hotel = Hotel.query.get(hid) if hid else (hotels[0] if hotels else None)
        if not desc or not hotel:
            return render_template('fault_submit.html', done=False, error='Συμπλήρωσε ξενοδοχείο & περιγραφή',
                                   hotels=hotels, cats=cats, priorities=PRIORITIES, user=user)
        f = Fault(code=gen_code(hotel), hotel_id=hotel.id,
                  category_id=request.form.get('category_id', type=int) or None,
                  description=desc, priority=request.form.get('priority', 'Κανονική'),
                  submitted_by=user.id, source='Ενδοξενοδοχειακά', status='pending_assign')
        db.session.add(f); db.session.flush()
        log_change(f, 'πεδίο', 'δημιουργία', '', f.code, user.id)
        _save_cover(f, request.files.get('cover'))
        auto_assign(f)
        db.session.commit()
        log_activity('fault_report', f.code)
        notify_admins('Νέα βλάβη: %s' % f.code, '/dashboard/fault/%d?embed=1' % f.id)
        threading.Thread(target=_bg_fault_email, args=(f.id,), daemon=True).start()
        return render_template('fault_submit.html', done=True, code=f.code, hotels=hotels, cats=cats,
                               priorities=PRIORITIES, user=user)
    return render_template('fault_submit.html', done=False, hotels=hotels, cats=cats,
                           priorities=PRIORITIES, user=user)

@app.route('/dashboard/faults')
def faults_inbox():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = current_user()
    q = visible_faults_query(user)
    f_hotel = request.args.get('hotel_id', type=int)
    f_status = request.args.get('status')
    f_priority = request.args.get('priority')
    f_assignee = request.args.get('assigned_user_id', type=int)
    search = (request.args.get('q') or '').strip()
    if f_hotel:    q = q.filter(Fault.hotel_id == f_hotel)
    if f_status:   q = q.filter(Fault.status == f_status)
    if f_priority: q = q.filter(Fault.priority == f_priority)
    if f_assignee: q = q.filter(Fault.assigned_user_id == f_assignee)
    if search:     q = q.filter((Fault.description.ilike('%%%s%%' % search)) | (Fault.code.ilike('%%%s%%' % search)))
    faults = q.order_by(Fault.id.desc()).limit(400).all()
    hotels = allowed_hotels(user)
    users = User.query.filter_by(is_active=True, approved=True).order_by(User.full_name).all()
    open_n = sum(1 for f in faults if f.status not in TERMINAL)
    return render_template('faults_list.html', faults=faults, hotels=hotels, users=users,
                           STATUS_LABELS=STATUS_LABELS, STATUS_COLOR=STATUS_COLOR, PRIORITY_COLOR=PRIORITY_COLOR,
                           PRIORITIES=PRIORITIES, STATUSES=STATUSES, is_admin=is_admin(),
                           f_hotel=f_hotel, f_status=f_status, f_priority=f_priority, f_assignee=f_assignee,
                           search=search, open_n=open_n, user=user, sla_state=sla_state,
                           qs=request.query_string.decode('utf-8'))

@app.route('/dashboard/fault/<int:fid>')
def fault_detail(fid):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = current_user()
    f = Fault.query.get_or_404(fid)
    if not can_view(user, f):
        return redirect(url_for('faults_inbox') + '?embed=1')
    logs = FaultChangeLog.query.filter_by(fault_id=fid).order_by(FaultChangeLog.at).all()
    comments = FaultComment.query.filter_by(fault_id=fid).order_by(FaultComment.at).all()
    if not is_admin() and f.submitted_by == user.id and f.assigned_user_id != user.id:
        comments = [c for c in comments if c.audience == 'υποβολέας']   # ο απλός υποβολέας δεν βλέπει «προς Διοίκηση»
    files = FaultAttachment.query.filter_by(fault_id=fid).all()
    cands = FaultCandidate.query.filter_by(fault_id=fid).all()
    users = User.query.filter_by(is_active=True, approved=True).order_by(User.full_name).all() if is_admin() else []
    allowed = TRANSITIONS.get(f.status, ())
    return render_template('fault_detail.html', f=f, logs=logs, comments=comments, files=files,
                           candidates=cands, users=users, allowed=allowed, is_admin=is_admin(),
                           STATUS_LABELS=STATUS_LABELS, STATUS_COLOR=STATUS_COLOR, PRIORITY_COLOR=PRIORITY_COLOR,
                           cat_path=_cat_path(f.category), user=user, me=user, sla_state=sla_state)

def _cat_path(cat):
    parts = []; seen = 0
    while cat and seen < 6:
        parts.append(cat.name); cat = FaultCategory.query.get(cat.parent_id) if cat.parent_id else None; seen += 1
    return ' › '.join(reversed(parts))

@app.route('/dashboard/fault/<int:fid>/status', methods=['POST'])
def fault_set_status(fid):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = current_user()
    f = Fault.query.get_or_404(fid)
    if not can_view(user, f):
        return redirect(url_for('faults_inbox') + '?embed=1')
    to = request.form.get('to')
    if not f.can_transition(to):
        return redirect(url_for('fault_detail', fid=fid) + '?embed=1&err=1')
    frm = f.status; f.status = to; f.updated_at = datetime.utcnow()
    if to == 'done':
        f.completed_at = datetime.utcnow(); f.completed_by = user.id
        if f.submitted_at:
            f.resolution_seconds = int((f.completed_at - f.submitted_at).total_seconds())
    log_change(f, 'κατάσταση', 'status', STATUS_LABELS.get(frm, frm), STATUS_LABELS.get(to, to), user.id)
    db.session.commit()
    if f.submitted_by:
        notify(f.submitted_by, 'Βλάβη %s: %s' % (f.code, STATUS_LABELS.get(to, to)), '/dashboard/fault/%d?embed=1' % f.id)
    return redirect(url_for('fault_detail', fid=fid) + '?embed=1')

@app.route('/dashboard/fault/<int:fid>/take', methods=['POST'])
def fault_take(fid):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = current_user()
    f = Fault.query.get_or_404(fid)
    if not can_view(user, f):
        return redirect(url_for('faults_inbox') + '?embed=1')
    f.assigned_user_id = user.id; f.updated_at = datetime.utcnow()
    log_change(f, 'σχέση', 'Ανάθεση σε', '', user.full_name, user.id)
    if f.can_transition('assigned'):
        f.status = 'assigned'
    elif f.can_transition('in_progress'):
        f.status = 'in_progress'
    db.session.commit()
    return redirect(url_for('fault_detail', fid=fid) + '?embed=1')

@app.route('/dashboard/fault/<int:fid>/assign', methods=['POST'])
def fault_assign(fid):
    if not is_admin():
        return redirect(url_for('login'))
    user = current_user()
    f = Fault.query.get_or_404(fid)
    uid = request.form.get('assignee_id', type=int)
    target = User.query.get(uid) if uid else None
    f.assigned_user_id = uid or None; f.updated_at = datetime.utcnow()
    log_change(f, 'σχέση', 'Ανάθεση σε', '', target.full_name if target else '—', user.id)
    if target and f.can_transition('assigned'):
        f.status = 'assigned'
    db.session.commit()
    if uid:
        notify(uid, 'Σου ανατέθηκε η βλάβη %s' % f.code, '/dashboard/fault/%d?embed=1' % f.id)
    return redirect(url_for('fault_detail', fid=fid) + '?embed=1')

@app.route('/dashboard/fault/<int:fid>/comment', methods=['POST'])
def fault_comment(fid):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = current_user()
    f = Fault.query.get_or_404(fid)
    if not can_view(user, f):
        return redirect(url_for('faults_inbox') + '?embed=1')
    audience = request.form.get('audience', 'διοίκηση')
    if audience not in ('διοίκηση', 'υποβολέας'):
        audience = 'διοίκηση'
    text = (request.form.get('text') or '').strip()
    if text:
        db.session.add(FaultComment(fault_id=fid, audience=audience, text=text, by_user_id=user.id))
        db.session.commit()
        if audience == 'υποβολέας' and f.submitted_by:
            notify(f.submitted_by, 'Σχόλιο στη βλάβη %s' % f.code, '/dashboard/fault/%d?embed=1' % f.id)
    return redirect(url_for('fault_detail', fid=fid) + '?embed=1')

@app.route('/dashboard/faults/bulk', methods=['POST'])
def faults_bulk():
    if not is_admin():
        return redirect(url_for('login'))
    user = current_user()
    action = request.form.get('action')
    ids = request.form.getlist('fault_ids')
    if ids == ['all'] or request.form.get('all') == '1':
        ids = [str(f.id) for f in visible_faults_query(user).all()]
    n = 0
    for sid in ids:
        try: f = Fault.query.get(int(sid))
        except Exception: f = None
        if not f:
            continue
        if action == 'complete' and f.status not in TERMINAL:
            frm = f.status; f.status = 'done'
            f.completed_at = datetime.utcnow(); f.completed_by = user.id
            if f.submitted_at:
                f.resolution_seconds = int((f.completed_at - f.submitted_at).total_seconds())
            log_change(f, 'κατάσταση', 'status', STATUS_LABELS.get(frm, frm), 'Ολοκληρώθηκε', user.id); n += 1
        elif action == 'assign':
            uid = request.form.get('assignee_id', type=int)
            f.assigned_user_id = uid or None
            if uid and f.can_transition('assigned'):
                f.status = 'assigned'
            log_change(f, 'σχέση', 'Ανάθεση σε', '', uid or '—', user.id); n += 1
    db.session.commit()
    log_activity('faults_bulk', '%s x%d' % (action, n))
    return redirect(url_for('faults_inbox') + '?embed=1')

# ── v12.16 (Φάση 2β) — συνημμένα/φωτό + email ────────────────────────────────
def _img_dataurl(file, limit=1500 * 1024):
    try:
        if file and file.filename and file.mimetype and file.mimetype.startswith('image/'):
            raw = file.read()
            if 0 < len(raw) <= limit:
                return 'data:' + file.mimetype + ';base64,' + base64.b64encode(raw).decode()
    except Exception:
        pass
    return None

def _save_cover(f, file):
    url = _img_dataurl(file)
    if url:
        f.cover_image = url
        db.session.add(FaultAttachment(fault_id=f.id, url=url, by_user_id=f.submitted_by))

def _bg_fault_email(fid):
    with app.app_context():
        f = Fault.query.get(fid)
        if not f:
            return
        try:
            cat = f.category.name if f.category else '—'
            who = f.submitter.full_name if f.submitter else '—'
            hotel = f.hotel.name if f.hotel else '—'
            html = ('<div style="font-family:Arial,sans-serif;max-width:600px;">'
                    '<div style="background:#193847;color:#fff;padding:16px;border-radius:8px 8px 0 0;">'
                    '<h2 style="margin:0;">Νέα βλάβη — %s</h2></div>'
                    '<div style="background:#f9f9f9;padding:16px;border:1px solid #eee;">'
                    '<p><b>Ξενοδοχείο:</b> %s</p><p><b>Κατηγορία:</b> %s</p>'
                    '<p><b>Προτεραιότητα:</b> %s</p><p><b>Από:</b> %s</p>'
                    '<p><b>Περιγραφή:</b><br>%s</p></div></div>'
                    % (f.code, hotel, cat, f.priority, who, (f.description or '')))
            send_email('Εστία — Νέα βλάβη %s' % f.code, html, EMAIL_TO_LIST)
        except Exception as e:
            print('[faults] email skipped:', e)

@app.route('/dashboard/fault/<int:fid>/attach', methods=['POST'])
def fault_attach(fid):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = current_user()
    f = Fault.query.get_or_404(fid)
    if not can_view(user, f):
        return redirect(url_for('faults_inbox') + '?embed=1')
    n = 0
    for file in request.files.getlist('files'):
        url = _img_dataurl(file)
        if url:
            db.session.add(FaultAttachment(fault_id=fid, url=url, by_user_id=user.id)); n += 1
            if not f.cover_image:
                f.cover_image = url
    if n:
        log_change(f, 'πεδίο', 'συνημμένα', '', '+%d' % n, user.id)
        db.session.commit()
    return redirect(url_for('fault_detail', fid=fid) + '?embed=1')

# ── v12.15 (Φάση 2) — SLA, admin ρυθμίσεις, export, ειδικότητες χρηστών ──────
def sla_minutes(f):
    t = SLATarget.query.filter_by(scope='tag', key=f.tag).first() if f.tag else None
    if not t:
        t = SLATarget.query.filter_by(scope='priority', key=f.priority).first()
    return t.minutes if t else None

def fault_stale_days():
    s = Setting.query.get('fault_stale_days')
    try:
        return int(s.value) if s and s.value else 60
    except Exception:
        return 60

def sla_state(f):
    """(code, label): '' | overdue 'Εκπρόθεσμη' | stale 'Χρονίζει'. Παγώνει σε paused/winter/τερματικές."""
    if (not f.submitted_at) or f.status in TERMINAL or f.status in ('paused', 'winter'):
        return ('', '')
    now = datetime.utcnow()
    if f.due_at and now > f.due_at:
        return ('overdue', 'Εκπρόθεσμη')
    elapsed_min = (now - f.submitted_at).total_seconds() / 60.0
    m = sla_minutes(f)
    if m and elapsed_min > m:
        return ('overdue', 'Εκπρόθεσμη')
    if elapsed_min > fault_stale_days() * 1440:
        return ('stale', 'Χρονίζει')
    return ('ok', '')

def _filtered_faults(user):
    q = visible_faults_query(user)
    a = request.args
    if a.get('hotel_id', type=int):   q = q.filter(Fault.hotel_id == a.get('hotel_id', type=int))
    if a.get('status'):               q = q.filter(Fault.status == a.get('status'))
    if a.get('priority'):             q = q.filter(Fault.priority == a.get('priority'))
    if a.get('assigned_user_id', type=int): q = q.filter(Fault.assigned_user_id == a.get('assigned_user_id', type=int))
    s = (a.get('q') or '').strip()
    if s:
        q = q.filter((Fault.description.ilike('%%%s%%' % s)) | (Fault.code.ilike('%%%s%%' % s)))
    return q.order_by(Fault.id.desc())

@app.route('/dashboard/faults/settings')
def faults_settings():
    if not is_admin():
        return redirect(url_for('login'))
    specs = Specialty.query.order_by(Specialty.sort, Specialty.name).all()
    slas = {t.key: t.minutes for t in SLATarget.query.filter_by(scope='priority').all()}
    tags = FaultTag.query.order_by(FaultTag.name).all()
    cmap = []
    for cs in CategorySpecialty.query.all():
        c = FaultCategory.query.get(cs.category_id)
        cmap.append({'id': cs.id, 'cat': c.name if c else '—', 'specialty': cs.specialty})
    allcats = FaultCategory.query.order_by(FaultCategory.level, FaultCategory.name).all()
    return render_template('faults_settings.html', specs=specs, slas=slas, tags=tags,
                           cmap=cmap, allcats=allcats, active_specs=[s.name for s in specs if s.is_active],
                           PRIORITIES=PRIORITIES, stale_days=fault_stale_days())

@app.route('/dashboard/faults/specialty/add', methods=['POST'])
def faults_specialty_add():
    if not is_admin(): return redirect(url_for('login'))
    name = (request.form.get('name') or '').strip()
    if name and not Specialty.query.filter_by(name=name).first():
        db.session.add(Specialty(name=name, sort=99)); db.session.commit()
    return redirect(url_for('faults_settings') + '?embed=1')

@app.route('/dashboard/faults/specialty/<int:sid>/toggle', methods=['POST'])
def faults_specialty_toggle(sid):
    if not is_admin(): return redirect(url_for('login'))
    sp = Specialty.query.get(sid)
    if sp: sp.is_active = not sp.is_active; db.session.commit()
    return redirect(url_for('faults_settings') + '?embed=1')

@app.route('/dashboard/faults/sla', methods=['POST'])
def faults_sla_save():
    if not is_admin(): return redirect(url_for('login'))
    for p in PRIORITIES:
        v = request.form.get('sla_' + p, type=int)
        if v and v > 0:
            t = SLATarget.query.filter_by(scope='priority', key=p).first()
            if t: t.minutes = v
            else: db.session.add(SLATarget(scope='priority', key=p, minutes=v))
    sd = request.form.get('stale_days', type=int)
    if sd and sd > 0:
        s = Setting.query.get('fault_stale_days')
        if s: s.value = str(sd)
        else: db.session.add(Setting(key='fault_stale_days', value=str(sd)))
    db.session.commit()
    return redirect(url_for('faults_settings') + '?embed=1&saved=1')

@app.route('/dashboard/faults/tag/add', methods=['POST'])
def faults_tag_add():
    if not is_admin(): return redirect(url_for('login'))
    name = (request.form.get('name') or '').strip()
    color = (request.form.get('color') or '#b91c1c').strip()
    if name and not FaultTag.query.filter_by(name=name).first():
        db.session.add(FaultTag(name=name, color=color)); db.session.commit()
    return redirect(url_for('faults_settings') + '?embed=1')

@app.route('/dashboard/faults/tag/<int:tid>/del', methods=['POST'])
def faults_tag_del(tid):
    if not is_admin(): return redirect(url_for('login'))
    t = FaultTag.query.get(tid)
    if t: db.session.delete(t); db.session.commit()
    return redirect(url_for('faults_settings') + '?embed=1')

@app.route('/dashboard/faults/map/add', methods=['POST'])
def faults_map_add():
    if not is_admin(): return redirect(url_for('login'))
    cid = request.form.get('category_id', type=int)
    sp = (request.form.get('specialty') or '').strip()
    if cid and sp and not CategorySpecialty.query.filter_by(category_id=cid, specialty=sp).first():
        db.session.add(CategorySpecialty(category_id=cid, specialty=sp)); db.session.commit()
    return redirect(url_for('faults_settings') + '?embed=1')

@app.route('/dashboard/faults/map/<int:mid>/del', methods=['POST'])
def faults_map_del(mid):
    if not is_admin(): return redirect(url_for('login'))
    m = CategorySpecialty.query.get(mid)
    if m: db.session.delete(m); db.session.commit()
    return redirect(url_for('faults_settings') + '?embed=1')

@app.route('/dashboard/faults/categories')
def faults_categories():
    if not is_admin(): return redirect(url_for('login'))
    roots = FaultCategory.query.filter_by(level=1).order_by(FaultCategory.name).all()
    def tree(node):
        kids = FaultCategory.query.filter_by(parent_id=node.id).order_by(FaultCategory.sort, FaultCategory.name).all()
        return {'node': node, 'children': [tree(k) for k in kids]}
    forest = [tree(r) for r in roots]
    allcats = FaultCategory.query.order_by(FaultCategory.level, FaultCategory.name).all()
    return render_template('faults_categories.html', forest=forest, allcats=allcats)

@app.route('/dashboard/faults/category/add', methods=['POST'])
def faults_category_add():
    if not is_admin(): return redirect(url_for('login'))
    name = (request.form.get('name') or '').strip()
    pid = request.form.get('parent_id', type=int)
    parent = FaultCategory.query.get(pid) if pid else None
    if name:
        lvl = (parent.level + 1) if parent else 1
        db.session.add(FaultCategory(name=name, parent_id=parent.id if parent else None, level=min(lvl, 3)))
        db.session.commit()
    return redirect(url_for('faults_categories') + '?embed=1')

@app.route('/dashboard/faults/category/<int:cid>/rename', methods=['POST'])
def faults_category_rename(cid):
    if not is_admin(): return redirect(url_for('login'))
    c = FaultCategory.query.get(cid); name = (request.form.get('name') or '').strip()
    if c and name: c.name = name; db.session.commit()
    return redirect(url_for('faults_categories') + '?embed=1')

@app.route('/dashboard/faults/category/<int:cid>/toggle', methods=['POST'])
def faults_category_toggle(cid):
    if not is_admin(): return redirect(url_for('login'))
    c = FaultCategory.query.get(cid)
    if c: c.is_active = not c.is_active; db.session.commit()
    return redirect(url_for('faults_categories') + '?embed=1')

@app.route('/dashboard/faults/user-specialties/<int:uid>', methods=['POST'])
def faults_user_specialties(uid):
    if not is_admin(): return redirect(url_for('login'))
    UserSpecialty.query.filter_by(user_id=uid).delete()
    for sp in request.form.getlist('specialties'):
        if Specialty.query.filter_by(name=sp).first():
            db.session.add(UserSpecialty(user_id=uid, specialty=sp))
    db.session.commit()
    return redirect(url_for('users_admin') + '?embed=1&success=user_edited')

@app.route('/dashboard/faults/export')
def faults_export():
    if 'user_id' not in session: return redirect(url_for('login'))
    from flask import Response
    user = current_user()
    faults = _filtered_faults(user).limit(5000).all()
    fmt = request.args.get('fmt', 'xlsx')
    headers = ['Κωδικός', 'Ξενοδοχείο', 'Κατηγορία', 'Προτεραιότητα', 'Κατάσταση', 'Ανάθεση', 'Υποβλήθηκε', 'Ολοκληρώθηκε', 'SLA']
    def row(f):
        return [f.code, f.hotel.name if f.hotel else '', f.category.name if f.category else '',
                f.priority, STATUS_LABELS.get(f.status, f.status),
                f.assignee.full_name if f.assignee else '',
                f.submitted_at.strftime('%d/%m/%Y %H:%M') if f.submitted_at else '',
                f.completed_at.strftime('%d/%m/%Y %H:%M') if f.completed_at else '',
                sla_state(f)[1]]
    fname = 'estia-faults-%s' % datetime.utcnow().strftime('%Y%m%d-%H%M')
    if fmt == 'csv':
        import io
        buf = io.StringIO(); w = csv.writer(buf); w.writerow(headers)
        for f in faults: w.writerow(row(f))
        return Response('﻿' + buf.getvalue(), mimetype='text/csv; charset=utf-8',
                        headers={'Content-Disposition': 'attachment; filename=%s.csv' % fname})
    if fmt == 'pdf':
        from fpdf import FPDF
        NAVY = (25, 56, 71)
        pdf = FPDF(orientation='L', unit='mm', format='A4'); pdf.set_auto_page_break(True, 12)
        pdf.add_font('dv', '', os.path.join(BASE_DIR, 'assets', 'fonts', 'DejaVuSans.ttf'))
        pdf.add_font('dv', 'B', os.path.join(BASE_DIR, 'assets', 'fonts', 'DejaVuSans-Bold.ttf'))
        pdf.add_page()
        pdf.set_font('dv', 'B', 14); pdf.set_text_color(*NAVY); pdf.cell(0, 9, 'Εστία — Βλάβες', ln=1)
        pdf.set_font('dv', '', 9); pdf.set_text_color(90, 90, 90)
        pdf.cell(0, 6, 'Σύνολο: %d · %s' % (len(faults), datetime.utcnow().strftime('%d/%m/%Y %H:%M')), ln=1); pdf.ln(2)
        widths = [30, 40, 50, 22, 34, 38, 30, 30]
        pdf.set_font('dv', 'B', 8); pdf.set_fill_color(*NAVY); pdf.set_text_color(255, 255, 255)
        for h, w in zip(headers[:8], widths): pdf.cell(w, 7, h, fill=True)
        pdf.ln(7); pdf.set_text_color(40, 40, 40)
        for f in faults:
            pdf.set_font('dv', '', 7.5)
            for val, w in zip(row(f)[:8], widths):
                s = str(val)
                while s and pdf.get_string_width(s) > w - 2 and len(s) > 3: s = s[:-2]
                pdf.cell(w, 6, s)
            pdf.ln(6)
        return Response(bytes(pdf.output()), mimetype='application/pdf',
                        headers={'Content-Disposition': 'attachment; filename=%s.pdf' % fname})
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = 'Βλάβες'
    navy = PatternFill('solid', fgColor='193847')
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h); c.font = Font(bold=True, color='FFFFFF'); c.fill = navy
    for i, f in enumerate(faults, 2):
        for j, v in enumerate(row(f), 1):
            ws.cell(row=i, column=j, value=v)
    for j, w in enumerate([16, 22, 30, 12, 18, 22, 17, 17, 12], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = 'A2'
    buf = io.BytesIO(); wb.save(buf)
    return Response(buf.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename=%s.xlsx' % fname})

print('[faults] module loaded')

