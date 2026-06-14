# -*- coding: utf-8 -*-
"""
Εστία — Module «Μισθοδοσία» (Payroll) — Φάση 1 (θεμέλιο & μητρώο)
================================================================
Plug-in: `import payroll` από το ΤΕΛΟΣ του app.py (αφού οριστούν app/db/helpers,
ΠΡΙΝ το init_db() ώστε το create_all να πιάσει τους νέους πίνακες).
Spec: 02_MODULES_ESTIA/ΜΙΣΘΟΔΟΣΙΑ/00_SPEC.md

Φ1: Company, Agreement, EmployeePII, PayrollRates (seed 2026) · Hotel.company_id ·
admin-only (P-08) · όψεις μητρώο/καρτέλα/εταιρείες/συντελεστές · διαβάζει schedule.py.

Αποφάσεις Giannis (14/06): Λογιστήριο όψη = Epsilon import (αλήθεια)· Management = τι
πληρώνεται· ωρομίσθιο = ημερομίσθιο÷8 (Α-02)· μενού νέα ομάδα «Οικονομικά» (Α-08).
ΔΕΝ αλλάζει layout (D-09). Μηχανή/Run/Line/Forecast/Outputs = Φ2+.
"""
import os, re, json, unicodedata
from datetime import datetime, date
from flask import request, redirect, url_for, render_template, session
from app import (app, db, current_user, is_admin, log_activity,
                 Hotel, User, Setting, role_rank, ROLE_RANK,
                 notify, notify_admins)

try:
    from schedule import EmploymentProfile, monthly_settlement  # L1/L2 source (S-13)
except Exception:
    EmploymentProfile = None
    monthly_settlement = None


# ── Χάρτης εταιρειών (SPEC §3) ────────────────────────────────────────────────
COMPANY_MAP = {
    'AST': ('Γ. ΓΙΑΝΝΟΥΛΑΚΗΣ – Α. ΓΙΑΝΝΟΥΛΑΚΗ Α.Ε.', '0002'),
    'CNT': ('Γ. ΓΙΑΝΝΟΥΛΑΚΗΣ – Α. ΓΙΑΝΝΟΥΛΑΚΗ Α.Ε.', None),
    'IRO': ('Γ. ΓΙΑΝΝΟΥΛΑΚΗΣ – Α. ΓΙΑΝΝΟΥΛΑΚΗ Α.Ε.', None),
    'SRG': ('ΑΦΟΙ ΣΕΡΓΙΟΥ Α.Ε.',                       None),
    'PSV': ('ΠΙΣΚΟΠΙΑΝΟ Α.Ε.',                         None),
    'PLM': ('ΦΥΤΩΡΙΑ ΚΡΗΤΗΣ Α.Ε.',                     None),
    'CND': ('CONDIAN HOTELS (HQ / Όμιλος)',            None),
}
COMPANY_CODE = {
    'AST': 'GIAN', 'CNT': 'GIAN', 'IRO': 'GIAN',
    'SRG': 'SERG', 'PSV': 'PISK', 'PLM': 'FYTO', 'CND': 'CND',
}
HOTEL_NAME_CODE = {
    'asterias': 'AST', 'central': 'CNT', 'iro': 'IRO', 'ηρω': 'IRO',
    'sergios': 'SRG', 'piskopiano': 'PSV', 'palm': 'PLM', 'condian': 'CND',
}


def _acc(s):
    return ''.join(c for c in unicodedata.normalize('NFD', str(s)) if unicodedata.category(c) != 'Mn')

def _norm(s):
    return re.sub(r'[^a-zα-ω0-9]', '', _acc(s).strip().lower()) if s else ''

def hotel_code(hotel):
    if not hotel:
        return None
    n = _norm(hotel.name)
    for key, code in HOTEL_NAME_CODE.items():
        if _norm(key) in n:
            return code
    return None


# ── ΜΟΝΤΕΛΑ ───────────────────────────────────────────────────────────────────
class Company(db.Model):
    id           = db.Column(db.Integer, primary_key=True)
    code         = db.Column(db.String(12), unique=True, index=True)
    legal_name   = db.Column(db.String(160), nullable=False)
    vat          = db.Column(db.String(20))
    subunit_code = db.Column(db.String(20))
    active       = db.Column(db.Boolean, default=True)


class Agreement(db.Model):
    """Συμφωνία με ιστορικό ισχύος (SPEC §4.2). Φ1: δομή + προαιρετική καταγραφή."""
    id                 = db.Column(db.Integer, primary_key=True)
    user_id            = db.Column(db.Integer, db.ForeignKey('user.id'), index=True, nullable=False)
    effective_from     = db.Column(db.Date, default=date.today)
    effective_to       = db.Column(db.Date)
    agreement_type     = db.Column(db.String(20), default='Μηνιαίος')
    agreed_amount      = db.Column(db.Float)
    folder_fixed       = db.Column(db.Float, default=0.0)
    hour_wage_override = db.Column(db.Float)
    channels_json      = db.Column(db.Text)
    note               = db.Column(db.String(200))
    created_at         = db.Column(db.DateTime, default=datetime.utcnow)
    created_by         = db.Column(db.Integer, db.ForeignKey('user.id'))


class EmployeePII(db.Model):
    """Ευαίσθητα στοιχεία — ξεχωριστός πίνακας, admin-gated (P-07/GDPR)."""
    id               = db.Column(db.Integer, primary_key=True)
    user_id          = db.Column(db.Integer, db.ForeignKey('user.id'), unique=True, nullable=False)
    afm              = db.Column(db.String(12))
    amka             = db.Column(db.String(12))
    ika_am           = db.Column(db.String(15))
    father_name      = db.Column(db.String(80))
    ergani_specialty = db.Column(db.String(120))
    contract_type    = db.Column(db.String(40))
    employment_kind  = db.Column(db.String(40))
    bank_name        = db.Column(db.String(60))
    bank_iban        = db.Column(db.String(34))
    hired_at         = db.Column(db.Date)
    left_at          = db.Column(db.Date)
    locked           = db.Column(db.Boolean, default=False)   # v12.56: κλειδωμένο μητρώο (πηγή=Epsilon)
    updated_at       = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    updated_by       = db.Column(db.Integer, db.ForeignKey('user.id'))


class PayrollRates(db.Model):
    """Συντελεστές ανά έτος (admin-editable). Seed 2026 ΓΙΑ ΕΚΤΙΜΗΣΗ/ΠΡΟΒΛΕΨΗ.
       Η αλήθεια των νόμιμων καθαρών έρχεται από Epsilon import."""
    id                    = db.Column(db.Integer, primary_key=True)
    year                  = db.Column(db.Integer, unique=True, index=True, nullable=False)
    efka_employee_pct     = db.Column(db.Float, default=13.87)
    efka_employer_pct     = db.Column(db.Float, default=22.29)
    efka_aux_employee_pct = db.Column(db.Float, default=3.25)
    efka_aux_employer_pct = db.Column(db.Float, default=3.25)
    tax_free_threshold    = db.Column(db.Float, default=8636.0)
    tax_brackets_json     = db.Column(db.Text)
    digital_fee           = db.Column(db.Float, default=0.0)
    note                  = db.Column(db.String(200))
    valid_from            = db.Column(db.Date)
    valid_to              = db.Column(db.Date)


TAX_BRACKETS_2026 = [[10000, 9], [20000, 22], [30000, 28], [40000, 36], [None, 44]]


# ── AUTH (admin-only — P-08) ──────────────────────────────────────────────────
def _padmin():
    return ('user_id' in session) and is_admin()


# ── MIGRATION (μη καταστροφικό) ───────────────────────────────────────────────
def ensure_payroll_columns():
    with app.app_context():
        try:
            from app import _add_col
            _add_col('hotel', 'company_id', 'company_id INTEGER')
            _add_col('legal_net_import', 'period_kind', "period_kind VARCHAR(24)")
            _add_col('payroll_line', 'extra_legal_net', 'extra_legal_net FLOAT')
            _add_col('payroll_line', 'extra_employer_cost', 'extra_employer_cost FLOAT')
            _add_col('payroll_line', 'extras_json', 'extras_json TEXT')
            _add_col('payroll_line', 'paid', 'paid FLOAT')
            _add_col('payroll_run', 'approved_by', 'approved_by INTEGER')
            _add_col('payroll_run', 'approved_at', 'approved_at DATETIME')
            _add_col('employee_pii', 'locked', 'locked BOOLEAN')
        except Exception as e:
            print('ensure_payroll_columns skipped:', e)


# ── SEED (idempotent) ─────────────────────────────────────────────────────────
def seed_payroll():
    with app.app_context():
        try:
            seen = {}
            for hcode, ccode in COMPANY_CODE.items():
                if ccode in seen:
                    continue
                legal, sub = COMPANY_MAP[hcode]
                if not Company.query.filter_by(code=ccode).first():
                    db.session.add(Company(code=ccode, legal_name=legal, subunit_code=sub))
                seen[ccode] = True
            db.session.commit()

            comp_by_code = {c.code: c for c in Company.query.all()}
            for h in Hotel.query.all():
                if getattr(h, 'company_id', None):
                    continue
                hc = hotel_code(h)
                cc = COMPANY_CODE.get(hc)
                if cc and cc in comp_by_code:
                    h.company_id = comp_by_code[cc].id
            db.session.commit()

            if not PayrollRates.query.filter_by(year=2026).first():
                db.session.add(PayrollRates(
                    year=2026,
                    tax_brackets_json=json.dumps(TAX_BRACKETS_2026, ensure_ascii=False),
                    note='Seed εκτίμησης (η αλήθεια από Epsilon import). Admin-editable.',
                    valid_from=date(2026, 1, 1), valid_to=date(2026, 12, 31)))
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print('seed_payroll skipped:', e)


# ── HELPERS ───────────────────────────────────────────────────────────────────
def _company_for_hotel(hotel_id):
    h = Hotel.query.get(hotel_id) if hotel_id else None
    if h and getattr(h, 'company_id', None):
        return Company.query.get(h.company_id)
    return None

def _employees():
    users = (User.query
             .filter((User.employment_active == True) | (User.employment_active.is_(None)))
             .all())
    out = []
    for u in users:
        prof = EmploymentProfile.query.filter_by(user_id=u.id).first() if EmploymentProfile else None
        pii = EmployeePII.query.filter_by(user_id=u.id).first()
        has_hr = bool(prof or pii or getattr(u, 'home_hotel_id', None) or getattr(u, 'department_id', None))
        if not has_hr:
            continue
        hid = getattr(u, 'home_hotel_id', None)
        comp = _company_for_hotel(hid)
        hotel = Hotel.query.get(hid) if hid else None
        out.append({'user': u, 'profile': prof, 'pii': pii, 'company': comp,
                    'hotel_name': (hotel.name if hotel else ''), 'hotel_id': hid,
                    'dept_id': getattr(u, 'department_id', None)})
    out.sort(key=lambda r: (r['company'].legal_name if r['company'] else 'ω', r['user'].full_name or ''))
    return out


# ── ROUTES ────────────────────────────────────────────────────────────────────
@app.route('/dashboard/payroll')
def payroll_home():
    if not _padmin():
        return redirect(url_for('login'))
    allrows = _employees()
    companies = Company.query.filter_by(active=True).order_by(Company.legal_name).all()
    # KPIs στο ΠΛΗΡΕΣ σύνολο
    n_emp = len(allrows)
    n_agree = sum(1 for r in allrows if r['profile'] and r['profile'].agreement_amount)
    n_pii = sum(1 for r in allrows if r['pii'] and r['pii'].afm)
    # φίλτρα
    company_id = request.args.get('company_id', type=int)
    hotel_id = request.args.get('hotel_id', type=int)
    flt = request.args.get('filter')
    rows = allrows
    if company_id: rows = [r for r in rows if r['company'] and r['company'].id == company_id]
    if hotel_id: rows = [r for r in rows if r.get('hotel_id') == hotel_id]
    if flt == 'agree':   rows = [r for r in rows if r['profile'] and r['profile'].agreement_amount]
    elif flt == 'noagree': rows = [r for r in rows if not (r['profile'] and r['profile'].agreement_amount)]
    elif flt == 'pii':   rows = [r for r in rows if r['pii'] and r['pii'].afm]
    elif flt == 'nopii': rows = [r for r in rows if not (r['pii'] and r['pii'].afm)]
    hotels = Hotel.query.order_by(Hotel.name).all()
    rates = PayrollRates.query.filter_by(year=2026).first()
    log_activity('payroll_view', 'μητρώο')
    return render_template('payroll_home.html',
        rows=rows, companies=companies, hotels=hotels, n_emp=n_emp, n_agree=n_agree, n_pii=n_pii,
        company_id=company_id, hotel_id=hotel_id, flt=flt, total_shown=len(rows),
        rates=rates, is_admin=is_admin())


@app.route('/dashboard/payroll/companies', methods=['GET', 'POST'])
def payroll_companies():
    if not _padmin():
        return redirect(url_for('login'))
    if request.method == 'POST':
        cid = request.form.get('id', type=int)
        c = Company.query.get(cid) if cid else None
        if c:
            c.legal_name   = (request.form.get('legal_name') or c.legal_name).strip()
            c.vat          = (request.form.get('vat') or '').strip() or None
            c.subunit_code = (request.form.get('subunit_code') or '').strip() or None
            c.active       = request.form.get('active') == '1'
            db.session.commit()
            log_activity('payroll_company_edit', c.legal_name)
        return redirect(url_for('payroll_companies'))
    companies = Company.query.order_by(Company.legal_name).all()
    hcount = {}
    for h in Hotel.query.all():
        cid = getattr(h, 'company_id', None)
        if cid:
            hcount.setdefault(cid, []).append(h.name)
    return render_template('payroll_companies.html', companies=companies, hcount=hcount, is_admin=is_admin())


@app.route('/dashboard/payroll/rates', methods=['GET', 'POST'])
def payroll_rates():
    if not _padmin():
        return redirect(url_for('login'))
    r = PayrollRates.query.filter_by(year=2026).first()
    if not r:
        r = PayrollRates(year=2026, tax_brackets_json=json.dumps(TAX_BRACKETS_2026, ensure_ascii=False))
        db.session.add(r); db.session.commit()
    if request.method == 'POST':
        def _f(name, cur):
            v = request.form.get(name)
            try:
                return float(v) if v not in (None, '') else cur
            except Exception:
                return cur
        r.efka_employee_pct     = _f('efka_employee_pct', r.efka_employee_pct)
        r.efka_employer_pct     = _f('efka_employer_pct', r.efka_employer_pct)
        r.efka_aux_employee_pct = _f('efka_aux_employee_pct', r.efka_aux_employee_pct)
        r.efka_aux_employer_pct = _f('efka_aux_employer_pct', r.efka_aux_employer_pct)
        r.tax_free_threshold    = _f('tax_free_threshold', r.tax_free_threshold)
        r.digital_fee           = _f('digital_fee', r.digital_fee)
        r.note                  = (request.form.get('note') or '').strip() or r.note
        db.session.commit()
        log_activity('payroll_rates_edit', '2026')
        return redirect(url_for('payroll_rates'))
    try:
        brackets = json.loads(r.tax_brackets_json or '[]')
    except Exception:
        brackets = []
    return render_template('payroll_rates.html', r=r, brackets=brackets, is_admin=is_admin())


@app.route('/dashboard/payroll/employee/<int:uid>', methods=['GET', 'POST'])
def payroll_employee(uid):
    if not _padmin():
        return redirect(url_for('login'))
    u = User.query.get_or_404(uid)
    prof = EmploymentProfile.query.filter_by(user_id=uid).first() if EmploymentProfile else None
    pii = EmployeePII.query.filter_by(user_id=uid).first()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'pii':
            if not pii:
                pii = EmployeePII(user_id=uid); db.session.add(pii)
            for f in ('afm', 'amka', 'ika_am', 'father_name', 'ergani_specialty',
                      'contract_type', 'employment_kind', 'bank_name', 'bank_iban'):
                setattr(pii, f, (request.form.get(f) or '').strip() or None)
            for f in ('hired_at', 'left_at'):
                v = request.form.get(f)
                try:
                    setattr(pii, f, datetime.strptime(v, '%Y-%m-%d').date() if v else None)
                except Exception:
                    pass
            cu = current_user()
            pii.updated_by = cu.id if cu else None
            db.session.commit()
            log_activity('payroll_pii_edit', u.full_name or u.username)
        elif action == 'agreement' and prof is not None:
            amt = request.form.get('agreement_amount')
            try:
                prof.agreement_amount = float(amt) if amt not in (None, '') else prof.agreement_amount
            except Exception:
                pass
            prof.agreement_type = request.form.get('agreement_type') or prof.agreement_type
            db.session.commit()
            log_activity('payroll_agreement_edit', u.full_name or u.username)
        return redirect(url_for('payroll_employee', uid=uid))
    comp = _company_for_hotel(getattr(u, 'home_hotel_id', None))
    hotel = Hotel.query.get(u.home_hotel_id) if getattr(u, 'home_hotel_id', None) else None
    agreements = (Agreement.query.filter_by(user_id=uid)
                  .order_by(Agreement.effective_from.desc()).all())
    return render_template('payroll_employee.html',
        u=u, prof=prof, pii=pii, comp=comp, hotel=hotel, agreements=agreements, is_admin=is_admin())


# ══════════════════════════════════════════════════════════════════════════════
# ΦΑΣΗ 2 — Μηχανή υπολογισμού + δύο όψεις + Epsilon import (Λογιστήριο = αλήθεια)
# ══════════════════════════════════════════════════════════════════════════════
import hashlib, io, zipfile
try:
    import openpyxl
except Exception:
    openpyxl = None

def _safe_load_xlsx(raw):
    """Ανθεκτικό άνοιγμα .xlsx — διορθώνει Epsilon exports με cellStyle χωρίς όνομα
    (openpyxl TypeError: name should be str but value is None)."""
    try:
        return openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except TypeError:
        zin = zipfile.ZipFile(io.BytesIO(raw)); buf = io.BytesIO()
        zout = zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED)
        for it in zin.namelist():
            data = zin.read(it)
            if it == 'xl/styles.xml':
                txt = data.decode('utf-8', 'replace')
                def _fix(m):
                    tag = m.group(0)
                    return tag if 'name=' in tag else tag[:-2] + ' name="Normal_x"/>'
                txt = re.sub(r'<cellStyle [^>]*/>', _fix, txt)
                data = txt.encode('utf-8')
            zout.writestr(it, data)
        zout.close(); buf.seek(0)
        return openpyxl.load_workbook(buf, data_only=True)

MONTHS_EL2 = ['', 'Ιανουάριος','Φεβρουάριος','Μάρτιος','Απρίλιος','Μάιος','Ιούνιος',
              'Ιούλιος','Αύγουστος','Σεπτέμβριος','Οκτώβριος','Νοέμβριος','Δεκέμβριος']
_EPSILON_MONTH = {}
for _i, _n in enumerate(MONTHS_EL2):
    if _i: _EPSILON_MONTH[_norm(_n)] = _i
_EPSILON_MONTH.update({_norm(k): v for k, v in {
    'ΙΑΝ':1,'ΦΕΒ':2,'ΜΑΡ':3,'ΑΠΡ':4,'ΜΑΙ':5,'ΜΑΪΟΣ':5,'ΙΟΥΝ':6,'ΙΟΥΛ':7,'ΑΥΓ':8,
    'ΣΕΠ':9,'ΟΚΤ':10,'ΝΟΕ':11,'ΔΕΚ':12,'JANUARY':1,'FEBRUARY':2,'MARCH':3,'APRIL':4,
    'MAY':5,'JUNE':6,'JULY':7,'AUGUST':8,'SEPTEMBER':9,'OCTOBER':10,'NOVEMBER':11,'DECEMBER':12,
}.items()})

RUN_STATUS = ('draft', 'calculated', 'verified', 'locked', 'paid')
RUN_LABELS = {'draft':'Πρόχειρο','calculated':'Υπολογίστηκε','verified':'Εγκρίθηκε (ΓΔ)',
              'locked':'Κλειδωμένο','paid':'Πληρωμένο'}


class PayrollRun(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    company_id    = db.Column(db.Integer, db.ForeignKey('company.id'), index=True, nullable=False)
    year          = db.Column(db.Integer, nullable=False)
    month         = db.Column(db.Integer, nullable=False)
    status        = db.Column(db.String(12), default='draft')
    rates_version = db.Column(db.Integer)
    note          = db.Column(db.String(200))
    created_by    = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)
    locked_at     = db.Column(db.DateTime)
    approved_by   = db.Column(db.Integer, db.ForeignKey('user.id'))
    approved_at   = db.Column(db.DateTime)
    __table_args__ = (db.UniqueConstraint('company_id', 'year', 'month', name='uq_payrollrun'),)


class PayrollLine(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    run_id        = db.Column(db.Integer, db.ForeignKey('payroll_run.id'), index=True, nullable=False)
    user_id       = db.Column(db.Integer, db.ForeignKey('user.id'), index=True, nullable=False)
    # ώρες/ημέρες (από monthly_settlement)
    work_days     = db.Column(db.Integer, default=0)
    sundays       = db.Column(db.Integer, default=0)
    holidays_worked = db.Column(db.Integer, default=0)
    extra_hours   = db.Column(db.Float, default=0.0)
    repo          = db.Column(db.Integer, default=0)
    total_days    = db.Column(db.Integer, default=0)
    elsewhere_days= db.Column(db.Integer, default=0)
    # Management όψη
    gross_agreement = db.Column(db.Float, default=0.0)
    extra_pay     = db.Column(db.Float, default=0.0)
    gross_total   = db.Column(db.Float, default=0.0)
    bank_target   = db.Column(db.Float)
    bonus         = db.Column(db.Float, default=0.0)
    bank_total    = db.Column(db.Float, default=0.0)
    folder_total  = db.Column(db.Float, default=0.0)
    in_hand       = db.Column(db.Float, default=0.0)
    # Λογιστήριο όψη (εκτίμηση) + αλήθεια Epsilon
    efka_employee = db.Column(db.Float, default=0.0)
    fmy           = db.Column(db.Float, default=0.0)
    net_calc      = db.Column(db.Float, default=0.0)
    employer_cost = db.Column(db.Float, default=0.0)
    net_legal     = db.Column(db.Float)        # από Epsilon
    employer_cost_legal = db.Column(db.Float)
    net_diff      = db.Column(db.Float)
    extra_legal_net = db.Column(db.Float, default=0.0)      # δώρα/άδεια καθαρά (Epsilon)
    extra_employer_cost = db.Column(db.Float, default=0.0)  # κόστος εργοδ. δώρων/άδειας
    extras_json   = db.Column(db.Text)                      # JSON [{kind,net,cost}]
    paid          = db.Column(db.Float, default=0.0)        # πληρωμένο (παρακολούθηση)
    note          = db.Column(db.String(200))


class LegalNetImport(db.Model):
    """Νόμιμα καθαρά από Epsilon (η αλήθεια του Λογιστηρίου). Idempotent ανά εταιρεία/μήνα/ΑΦΜ."""
    id            = db.Column(db.Integer, primary_key=True)
    company_id    = db.Column(db.Integer, db.ForeignKey('company.id'), index=True)
    user_id       = db.Column(db.Integer, db.ForeignKey('user.id'), index=True)
    year          = db.Column(db.Integer, index=True)
    month         = db.Column(db.Integer, index=True)
    afm           = db.Column(db.String(12), index=True)
    emp_name      = db.Column(db.String(120))
    gross_legal   = db.Column(db.Float)
    efka_employee_legal = db.Column(db.Float)
    fmy_legal     = db.Column(db.Float)
    net_legal     = db.Column(db.Float)
    employer_cost_legal = db.Column(db.Float)
    period_kind   = db.Column(db.String(24), default='monthly')  # monthly/Δώρο Πάσχα/Επίδομα Αδείας...
    import_hash   = db.Column(db.String(40), unique=True, index=True)
    source_file   = db.Column(db.String(160))
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)


# ── EPSILON PARSER ────────────────────────────────────────────────────────────
def _epsilon_headers(ws):
    """Χάρτης normalized-header -> column index από την 1η γραμμή."""
    h = {}
    for c in range(1, ws.max_column + 1):
        h[_norm(ws.cell(1, c).value)] = c
    return h

def parse_epsilon(wb):
    """Επιστρέφει λίστα dict ανά εργαζόμενο από workbook Epsilon (44 στήλες)."""
    ws = wb.worksheets[0]
    H = _epsilon_headers(ws)
    def col(*names, default=None):
        for n in names:
            c = H.get(_norm(n))
            if c: return c
        return default
    c_epon=col('Επώνυμο'); c_onoma=col('Όνομα'); c_afm=col('ΑΦΜ')
    c_amka=col('ΑΜΚΑ'); c_ika=col('Α.Μ. ΙΚΑ','ΑΜ ΙΚΑ'); c_pat=col('Όνομα Πατρός')
    c_spec=col('Ειδικότητα'); c_kind=col('Είδος Εργάζ','Είδος Εργαζ'); c_contract=col('Διάρκεια Σύμβασης')
    c_sub=col('Περιγραφή Υποκαταστήματος'); c_dept=col('Περιγραφή Τμήματος')
    c_bank=col('Τράπεζα'); c_period=col('Περίοδος'); c_year=col('Έτος')
    c_gross=col('Συν.Αποδ.','Συν.Αποδ'); c_emain=col('Εισφ. Εργάζ. Κύριου Ταμείου')
    c_eaux=col('Εισφ. Εργάζ. Επικ. Ταμείου'); c_fmy=col('Φ.Μ.Υ','ΦΜΥ')
    c_net=col('Καθαρές Αποδοχές'); c_cost=col('Συνολικό Κόστος')
    rows=[]
    for r in range(2, ws.max_row + 1):
        epon = ws.cell(r, c_epon).value if c_epon else None
        if not epon: continue
        onoma = ws.cell(r, c_onoma).value if c_onoma else ''
        per = ws.cell(r, c_period).value if c_period else None
        yr  = ws.cell(r, c_year).value if c_year else None
        month = _EPSILON_MONTH.get(_norm(per)) if per else None
        period_raw = str(per).strip() if per else ''
        try: yr = int(yr) if yr else None
        except Exception: yr = None
        def num(c):
            v = ws.cell(r, c).value if c else None
            try: return float(v) if v not in (None,'') else 0.0
            except Exception: return 0.0
        emain=num(c_emain); eaux=num(c_eaux)
        rows.append({
            'epon': str(epon).strip(), 'onoma': str(onoma).strip() if onoma else '',
            'afm': str(ws.cell(r,c_afm).value).strip() if c_afm and ws.cell(r,c_afm).value else None,
            'amka': str(ws.cell(r,c_amka).value).strip() if c_amka and ws.cell(r,c_amka).value else None,
            'ika': str(ws.cell(r,c_ika).value).strip() if c_ika and ws.cell(r,c_ika).value else None,
            'father': (ws.cell(r,c_pat).value if c_pat else None),
            'specialty': (ws.cell(r,c_spec).value if c_spec else None),
            'kind': (ws.cell(r,c_kind).value if c_kind else None),
            'contract': (ws.cell(r,c_contract).value if c_contract else None),
            'subunit_desc': (ws.cell(r,c_sub).value if c_sub else None),
            'dept_desc': (ws.cell(r,c_dept).value if c_dept else None),
            'bank': (ws.cell(r,c_bank).value if c_bank else None),
            'year': yr, 'month': month, 'period_raw': period_raw,
            'gross': num(c_gross), 'efka_employee': round(emain+eaux,2),
            'fmy': num(c_fmy), 'net': num(c_net), 'employer_cost': num(c_cost),
        })
    return rows

def _match_user_by_afm_or_name(afm, epon, onoma):
    if afm:
        pii = EmployeePII.query.filter_by(afm=str(afm)).first()
        if pii: return User.query.get(pii.user_id)
    full = _norm((str(epon) + str(onoma or '')))
    for u in User.query.all():
        if _norm(u.full_name) == full:
            return u
    # χαλαρό: επώνυμο match μοναδικό
    cand = [u for u in User.query.all() if _norm(u.full_name).startswith(_norm(epon))]
    return cand[0] if len(cand) == 1 else None

_KIND_MAP = {
    'δωροπασχα':'Δώρο Πάσχα', 'δωροχριστουγεννων':'Δώρο Χριστουγέννων',
    'επιδομααδειας':'Επίδομα Αδείας', 'αποζημιωσηαδειας':'Αποζημίωση Αδείας',
    'αποδοχεςαδειας':'Αποδοχές Αδείας',
}
def _period_kind(period_raw):
    n = _norm(period_raw)
    if n in _EPSILON_MONTH: return 'monthly'
    return _KIND_MAP.get(n, period_raw or 'extra')

_HOTEL_KW = {
    'AST': ['asterias', 'αστεριας'], 'CNT': ['central', 'χερσονησ', 'hersoniss'],  # ΟΧΙ 'κεντρικο' (=έδρα)
    'IRO': ['iro', 'ηρω'], 'SRG': ['sergios', 'σεργιος', 'σεργιου'],
    'PSV': ['piskopiano', 'πισκοπιανο'], 'PLM': ['palm', 'παλμ'], 'CND': ['condian', 'κονντιαν'],
}
def _hotel_from_epsilon(row, comp=None):
    """Ξενοδοχείο: ΕΤΑΙΡΕΙΑ-ΠΡΩΤΑ. Μονοξενοδοχειακή εταιρεία -> το ξεν. της.
    Πολυξενοδοχειακή (Γιαννουλάκης: AST/CNT/IRO) -> disambiguate από ΥΠΟΚ.
    («Κεντρικό» = έδρα/λογιστικό υποκατάστημα, ΟΧΙ ξενοδοχείο)."""
    blob = _norm(str(row.get('subunit_desc') or '') + ' ' + str(row.get('dept_desc') or ''))
    if comp:
        hs = Hotel.query.filter_by(company_id=comp.id).all()
        if len(hs) == 1:
            return hs[0]
        for h in hs:                       # πολλά ξεν. -> διάλεξε με βάση το ΥΠΟΚ
            code = hotel_code(h)
            if any(_norm(k) in blob for k in _HOTEL_KW.get(code, [])):
                return h
        return None                        # έδρα/«Κεντρικό» -> χωρίς ξεν.
    # χωρίς εταιρεία: καθαρό keyword
    for code, kws in _HOTEL_KW.items():
        if any(_norm(k) in blob for k in kws):
            for h in Hotel.query.all():
                if hotel_code(h) == code:
                    return h
    return None

def _create_locked_employee(row):
    """Δημιουργεί χρήστη-εργαζόμενο (login-off) από Epsilon (master)."""
    from werkzeug.security import generate_password_hash
    full = (str(row['epon']).strip() + ' ' + (str(row['onoma']).strip() if row['onoma'] else '')).strip()
    base = re.sub(r'[^a-z0-9.]', '', _acc(full).lower().replace(' ', '.')) or ('emp' + os.urandom(3).hex())
    uname = base[:40]
    if User.query.filter_by(username=uname).first():
        uname = (base[:32] + '.' + os.urandom(2).hex())[:46]
    u = User(username=uname, password=generate_password_hash(os.urandom(8).hex()),
             full_name=full[:100], role='staff', approved=True, is_active=True)
    for attr, val in [('login_enabled', False), ('employment_active', True)]:
        if hasattr(u, attr):
            setattr(u, attr, val)
    db.session.add(u); db.session.flush()
    return u

def _apply_epsilon_identity(u, row, comp=None):
    """Ταυτότητα από Epsilon (αλήθεια): ξενοδοχείο + PII + κλείδωμα."""
    hotel = _hotel_from_epsilon(row, comp)
    if hotel:
        u.home_hotel_id = hotel.id
    pii = EmployeePII.query.filter_by(user_id=u.id).first()
    if not pii:
        pii = EmployeePII(user_id=u.id); db.session.add(pii)
    filled = False
    for fld, val in [('afm',row['afm']),('amka',row['amka']),('ika_am',row['ika']),
                     ('father_name',row['father']),('ergani_specialty',row['specialty']),
                     ('employment_kind',row['kind']),('contract_type',row['contract']),
                     ('bank_name',row['bank'])]:
        if val and not getattr(pii, fld, None):
            setattr(pii, fld, str(val)[:120]); filled = True
    pii.locked = True
    db.session.flush()
    return filled

def import_epsilon_bytes(raw, filename='', company_id=None):
    """Εισάγει Epsilon workbook -> LegalNetImport (κανονικός + δώρα/άδεια). Idempotent."""
    if openpyxl is None:
        return {'error': 'openpyxl μη διαθέσιμο'}
    if filename and filename.lower().endswith('.xls'):
        return {'error': 'Παλιό format .xls — άνοιξέ το και αποθήκευσέ το ως .xlsx και ξαναδοκίμασε.'}
    wb = _safe_load_xlsx(raw)
    rows = parse_epsilon(wb); wb.close()
    comp = Company.query.get(company_id) if company_id else None
    if not comp and filename:
        pref = re.split(r'[ _]', filename)[0].upper()
        cc = COMPANY_CODE.get(pref)
        if cc: comp = Company.query.filter_by(code=cc).first()
    # κυρίαρχος μήνας αρχείου (από τις γραμμές που ΕΧΟΥΝ μήνα)
    from collections import Counter
    mc = Counter((r['year'], r['month']) for r in rows if r['year'] and r['month'])
    dom = mc.most_common(1)[0][0] if mc else (None, None)
    added = updated = matched = pii_filled = extras = created = 0
    period = set()
    for row in rows:
        kind = _period_kind(row['period_raw'])
        if kind == 'monthly' and row['year'] and row['month']:
            yr, mo = row['year'], row['month']
        else:
            yr, mo = dom            # δώρα/άδεια -> κυρίαρχος μήνας αρχείου
            if kind != 'monthly': extras += 1
        if not yr or not mo: continue
        period.add((yr, mo))
        u = _match_user_by_afm_or_name(row['afm'], row['epon'], row['onoma'])
        if u is None and kind == 'monthly':
            u = _create_locked_employee(row); created += 1
        if u is not None:
            matched += 1
            if _apply_epsilon_identity(u, row, comp):
                pii_filled += 1
        key = '%s|%s|%s|%s|%s' % (comp.id if comp else 'x', yr, mo, kind, row['afm'] or (row['epon']+row['onoma']))
        h = hashlib.sha1(key.encode('utf-8')).hexdigest()[:40]
        rec = LegalNetImport.query.filter_by(import_hash=h).first()
        if not rec:
            rec = LegalNetImport(import_hash=h); db.session.add(rec); added += 1
        else:
            updated += 1
        rec.company_id = comp.id if comp else None
        rec.user_id = u.id if u else None
        rec.year = yr; rec.month = mo; rec.afm = row['afm']; rec.period_kind = kind
        rec.emp_name = (row['epon'] + ' ' + (row['onoma'] or '')).strip()
        rec.gross_legal = row['gross']; rec.efka_employee_legal = row['efka_employee']
        rec.fmy_legal = row['fmy']; rec.net_legal = row['net']; rec.employer_cost_legal = row['employer_cost']
        rec.source_file = filename[:160]
    db.session.commit()
    return {'added': added, 'updated': updated, 'matched': matched, 'rows': len(rows),
            'created': created, 'pii_filled': pii_filled, 'extras': extras,
            'company': comp.legal_name if comp else None, 'periods': sorted(period)}


# ── ΜΗΧΑΝΗ: build_run ─────────────────────────────────────────────────────────
def _hotels_of_company(company_id):
    return [h.id for h in Hotel.query.filter_by(company_id=company_id).all()]

def build_run(company_id, year, month, created_by=None):
    """Δημιουργεί/ξαναϋπολογίζει PayrollRun + PayrollLines για εταιρεία×μήνα."""
    if monthly_settlement is None:
        return {'error': 'schedule module μη διαθέσιμο'}
    run = PayrollRun.query.filter_by(company_id=company_id, year=year, month=month).first()
    if run and run.status in ('verified', 'locked', 'paid'):
        return {'run_id': run.id, 'lines': PayrollLine.query.filter_by(run_id=run.id).count(),
                'locked': True}   # εγκεκριμένη — δεν ξαναϋπολογίζεται
    if not run:
        run = PayrollRun(company_id=company_id, year=year, month=month, created_by=created_by)
        db.session.add(run); db.session.flush()
    rates = PayrollRates.query.filter_by(year=year).first()
    run.rates_version = rates.id if rates else None
    prev_paid = {l.user_id: (l.paid or 0.0) for l in PayrollLine.query.filter_by(run_id=run.id).all()}
    PayrollLine.query.filter_by(run_id=run.id).delete()
    hotel_ids = _hotels_of_company(company_id)
    seen = set()
    legal_by_user = {}; extras_by_user = {}
    for li in LegalNetImport.query.filter_by(company_id=company_id, year=year, month=month).all():
        if not li.user_id: continue
        if (li.period_kind or 'monthly') == 'monthly':
            legal_by_user[li.user_id] = li
        else:
            extras_by_user.setdefault(li.user_id, []).append(li)
    n = 0
    for hid in hotel_ids:
        for row in monthly_settlement(year, month, hid):
            u = row['user']; agg = row['agg']; prof = row['profile']
            if u.id in seen: continue
            seen.add(u.id)
            day_wage = prof.day_wage if prof else 0.0
            hour_wage = (prof.hour_wage if prof else 0.0)
            if prof and getattr(prof, 'agreement_type', '') == 'Management' and prof.agreement_amount:
                gross_agreement = round(prof.agreement_amount, 2)
            else:
                gross_agreement = round(day_wage * agg['total_days'], 2)
            extra_pay = round(hour_wage * agg['extra_hours'], 2)
            gross_total = round(gross_agreement + extra_pay, 2)
            # Λογιστήριο: αλήθεια Epsilon αν υπάρχει, αλλιώς εκτίμηση
            li = legal_by_user.get(u.id)
            if li and li.net_legal:
                net_ref = li.net_legal; efka = li.efka_employee_legal or 0.0
                fmy = li.fmy_legal or 0.0; emp_cost = li.employer_cost_legal or 0.0
                net_legal = li.net_legal; net_calc = net_ref
            else:
                erate = ((rates.efka_employee_pct or 0) + (rates.efka_aux_employee_pct or 0)) / 100.0 if rates else 0.1387
                efka = round(gross_total * erate, 2)
                fmy = 0.0
                net_calc = round(gross_total - efka - fmy, 2)
                emp_rate = ((rates.efka_employer_pct or 0) + (rates.efka_aux_employer_pct or 0)) / 100.0 if rates else 0.2229
                emp_cost = round(gross_total * (1 + emp_rate), 2)
                net_ref = net_calc; net_legal = None
            ag = Agreement.query.filter_by(user_id=u.id, effective_to=None).first()
            # bank_target: στόχος τράπεζας από συμφωνία (Φ2: αν δηλωθεί στο channels_json)· αλλιώς None
            bank_target = None
            if ag and ag.channels_json:
                try:
                    bank_target = (json.loads(ag.channels_json) or {}).get('bank_target')
                except Exception:
                    bank_target = None
            bonus = round(max(0.0, (bank_target - net_ref)), 2) if bank_target else 0.0
            bank_total = round(net_ref + bonus, 2)
            folder_fixed = (ag.folder_fixed if ag else 0.0) or 0.0
            folder_total = round(folder_fixed + extra_pay, 2)
            in_hand = round(bank_total + folder_total, 2)
            ex_list = extras_by_user.get(u.id, [])
            extra_legal_net = round(sum((e.net_legal or 0) for e in ex_list), 2)
            extra_emp_cost = round(sum((e.employer_cost_legal or 0) for e in ex_list), 2)
            extras_json = json.dumps([{'kind': e.period_kind, 'net': round(e.net_legal or 0, 2),
                                       'cost': round(e.employer_cost_legal or 0, 2)} for e in ex_list],
                                     ensure_ascii=False) if ex_list else None
            line = PayrollLine(run_id=run.id, user_id=u.id,
                extra_legal_net=extra_legal_net, extra_employer_cost=extra_emp_cost, extras_json=extras_json,
                work_days=agg['work_days'], sundays=agg['sundays'],
                holidays_worked=agg['holidays_worked'], extra_hours=agg['extra_hours'],
                repo=agg['repo'], total_days=agg['total_days'], elsewhere_days=agg['elsewhere_days'],
                gross_agreement=gross_agreement, extra_pay=extra_pay, gross_total=gross_total,
                bank_target=bank_target, bonus=bonus, bank_total=bank_total,
                folder_total=folder_total, in_hand=in_hand,
                efka_employee=efka, fmy=fmy, net_calc=net_calc, employer_cost=emp_cost,
                net_legal=net_legal,
                employer_cost_legal=(li.employer_cost_legal if li else None),
                net_diff=(round((net_legal - net_calc), 2) if net_legal is not None else None))
            line.paid = prev_paid.get(u.id, 0.0)
            db.session.add(line); n += 1
    run.status = 'calculated'
    db.session.commit()
    return {'run_id': run.id, 'lines': n, 'legal_matched': len(legal_by_user)}

def run_totals(run_id):
    lines = PayrollLine.query.filter_by(run_id=run_id).all()
    t = {'gross_total':0.0,'extra_pay':0.0,'bank_total':0.0,'folder_total':0.0,'in_hand':0.0,
         'bonus':0.0,'net_legal':0.0,'employer_cost':0.0,'extra_legal':0.0,
         'in_hand_total':0.0,'employer_cost_total':0.0,'n':len(lines),'n_legal':0,'n_extras':0}
    for l in lines:
        t['gross_total']+=l.gross_total or 0; t['extra_pay']+=l.extra_pay or 0
        t['bank_total']+=l.bank_total or 0; t['folder_total']+=l.folder_total or 0
        t['in_hand']+=l.in_hand or 0; t['bonus']+=l.bonus or 0
        ec = (l.employer_cost_legal or l.employer_cost or 0)
        t['employer_cost']+= ec
        eln = l.extra_legal_net or 0; eec = l.extra_employer_cost or 0
        t['extra_legal']+= eln
        t['in_hand_total']+= (l.in_hand or 0) + eln
        t['employer_cost_total']+= ec + eec
        if eln: t['n_extras']+=1
        if l.net_legal is not None: t['net_legal']+=l.net_legal; t['n_legal']+=1
    for k in t:
        if isinstance(t[k], float): t[k]=round(t[k],2)
    return t


# ── ROUTES Φ2 ─────────────────────────────────────────────────────────────────
@app.route('/dashboard/payroll/runs', methods=['GET', 'POST'])
def payroll_runs():
    if not _padmin():
        return redirect(url_for('login'))
    if request.method == 'POST':
        cid = request.form.get('company_id', type=int)
        yr = request.form.get('year', type=int); mo = request.form.get('month', type=int)
        if cid and yr and mo:
            res = build_run(cid, yr, mo, created_by=(current_user().id if current_user() else None))
            log_activity('payroll_run_build', '%s/%s/%s -> %s' % (cid, yr, mo, res))
            run = PayrollRun.query.filter_by(company_id=cid, year=yr, month=mo).first()
            if run: return redirect(url_for('payroll_run_view', rid=run.id))
        return redirect(url_for('payroll_runs'))
    runs = PayrollRun.query.order_by(PayrollRun.year.desc(), PayrollRun.month.desc()).all()
    companies = Company.query.filter_by(active=True).order_by(Company.legal_name).all()
    comp_by_id = {c.id: c for c in Company.query.all()}
    tot_by_run = {r.id: run_totals(r.id) for r in runs}
    return render_template('payroll_runs.html', runs=runs, companies=companies,
        comp_by_id=comp_by_id, tot_by_run=tot_by_run, months=MONTHS_EL2,
        run_labels=RUN_LABELS, is_admin=is_admin())

@app.route('/dashboard/payroll/run/<int:rid>')
def payroll_run_view(rid):
    if not _padmin():
        return redirect(url_for('login'))
    run = PayrollRun.query.get_or_404(rid)
    comp = Company.query.get(run.company_id)
    lines = PayrollLine.query.filter_by(run_id=rid).all()
    umap = {u.id: u for u in User.query.all()}
    lines.sort(key=lambda l: (umap.get(l.user_id).full_name if umap.get(l.user_id) else ''))
    view = request.args.get('view', 'management')
    return render_template('payroll_run.html', run=run, comp=comp, lines=lines, umap=umap,
        totals=run_totals(rid), months=MONTHS_EL2, run_labels=RUN_LABELS, view=view, is_admin=is_admin())

@app.route('/dashboard/payroll/import', methods=['GET', 'POST'])
def payroll_import():
    if not _padmin():
        return redirect(url_for('login'))
    results = []
    if request.method == 'POST':
        files = request.files.getlist('file')
        cid = request.form.get('company_id', type=int)
        for f in files:
            if not (f and f.filename):
                continue
            try:
                r = import_epsilon_bytes(f.read(), filename=f.filename, company_id=cid)
            except Exception as e:
                r = {'error': str(e)}
            r['_file'] = f.filename
            results.append(r)
        log_activity('payroll_epsilon_import', '%d αρχεία' % len(results))
    companies = Company.query.filter_by(active=True).order_by(Company.legal_name).all()
    n_legal = LegalNetImport.query.count()
    return render_template('payroll_import.html', companies=companies, results=results,
        n_legal=n_legal, is_admin=is_admin())

# ══════════════════════════════════════════════════════════════════════════════
# ΦΑΣΗ 2.2 — Σελίδα «Έλεγχος & Έγκριση» (πίνακας Γενικού Διευθυντή)
# ══════════════════════════════════════════════════════════════════════════════
try:
    from schedule import Department as _Dept
except Exception:
    _Dept = None

def _control_rows(year, month, company_id=None, hotel_id=None, dept_id=None):
    runs_q = PayrollRun.query.filter_by(year=year, month=month)
    if company_id:
        runs_q = runs_q.filter_by(company_id=company_id)
    runs = runs_q.all()
    comp_by_id = {c.id: c for c in Company.query.all()}
    umap = {u.id: u for u in User.query.all()}
    dmap = {d.id: d.name for d in _Dept.query.all()} if _Dept else {}
    rows = []
    tot = {'work_days':0,'repo':0,'extra_hours':0.0,'mgmt':0.0,'legal':0.0,
           'ektos':0.0,'payable':0.0,'paid':0.0,'remaining':0.0}
    for run in runs:
        comp = comp_by_id.get(run.company_id)
        for l in PayrollLine.query.filter_by(run_id=run.id).all():
            u = umap.get(l.user_id)
            if not u:
                continue
            hid = getattr(u, 'home_hotel_id', None)
            if hotel_id and hid != hotel_id:
                continue
            did = getattr(u, 'department_id', None)
            if dept_id and did != dept_id:
                continue
            hotel = Hotel.query.get(hid) if hid else None
            mgmt = round((l.in_hand or 0) + (l.extra_legal_net or 0), 2)         # στο χέρι σύνολο
            legal = (round((l.net_legal or 0), 2) if l.net_legal is not None else None)
            ektos = round(mgmt - (l.net_legal or 0), 2)                          # εκτός λογιστηρίου
            paid = round(l.paid or 0.0, 2)
            remaining = round(mgmt - paid, 2)
            diff = (round((l.net_legal or 0) - (l.net_calc or 0), 2) if l.net_legal is not None else None)
            flags = []
            if not (l.gross_agreement or 0): flags.append('χωρίς συμφωνία')
            if (l.repo or 0) == 0 and (l.work_days or 0) > 0: flags.append('0 ρεπό')
            if (l.extra_hours or 0) > 40: flags.append('πολλές έξτρα')
            if legal is None and (l.work_days or 0) > 0: flags.append('χωρίς Epsilon')
            rows.append({'user': u, 'company': comp, 'run': run,
                'hotel': hotel.name if hotel else '', 'dept': dmap.get(did, ''),
                'work_days': l.work_days or 0, 'repo': l.repo or 0,
                'extra_hours': l.extra_hours or 0, 'mgmt': mgmt, 'legal': legal,
                'diff': diff, 'ektos': ektos, 'payable': mgmt, 'paid': paid,
                'remaining': remaining, 'flags': flags})
            tot['work_days'] += l.work_days or 0; tot['repo'] += l.repo or 0
            tot['extra_hours'] += l.extra_hours or 0; tot['mgmt'] += mgmt
            tot['legal'] += (l.net_legal or 0); tot['ektos'] += ektos
            tot['payable'] += mgmt; tot['paid'] += paid; tot['remaining'] += remaining
    for k in tot:
        if isinstance(tot[k], float): tot[k] = round(tot[k], 2)
    rows.sort(key=lambda r: (r['company'].legal_name if r['company'] else '', r['hotel'], r['user'].full_name or ''))
    return rows, runs, tot


@app.route('/dashboard/payroll/control', methods=['GET', 'POST'])
def payroll_control():
    if not _padmin():
        return redirect(url_for('login'))
    year = request.values.get('year', type=int) or 2026
    month = request.values.get('month', type=int) or 5
    if request.method == 'POST' and request.form.get('action') == 'calc':
        # υπολόγισε όλες τις εταιρείες για τον μήνα
        for c in Company.query.filter_by(active=True).all():
            build_run(c.id, year, month, created_by=(current_user().id if current_user() else None))
        log_activity('payroll_calc_month', '%s/%s' % (year, month))
        return redirect(url_for('payroll_control', year=year, month=month))
    company_id = request.values.get('company_id', type=int)
    hotel_id = request.values.get('hotel_id', type=int)
    dept_id = request.values.get('dept_id', type=int)
    rows, runs, tot = _control_rows(year, month, company_id, hotel_id, dept_id)
    companies = Company.query.filter_by(active=True).order_by(Company.legal_name).all()
    hotels = Hotel.query.order_by(Hotel.name).all()
    depts = _Dept.query.order_by(_Dept.name).all() if _Dept else []
    # per-company run status (για κουμπιά έγκρισης)
    run_map = {r.company_id: r for r in runs}
    return render_template('payroll_control.html', rows=rows, tot=tot, runs=runs, run_map=run_map,
        companies=companies, hotels=hotels, depts=depts, months=MONTHS_EL2,
        year=year, month=month, company_id=company_id, hotel_id=hotel_id, dept_id=dept_id,
        run_labels=RUN_LABELS, is_admin=is_admin())


@app.route('/dashboard/payroll/approve', methods=['POST'])
def payroll_approve():
    if not _padmin():
        return ('', 403)
    rid = request.form.get('run_id', type=int)
    run = PayrollRun.query.get(rid) if rid else None
    if run and run.status not in ('verified', 'locked', 'paid'):
        cu = current_user()
        run.status = 'verified'
        run.approved_by = cu.id if cu else None
        run.approved_at = datetime.utcnow()
        run.locked_at = datetime.utcnow()
        db.session.commit()
        comp = Company.query.get(run.company_id)
        period = '%s %s' % (MONTHS_EL2[run.month], run.year)
        link = '/dashboard/payroll/run/%d?embed=1' % run.id
        msg = 'Εγκρίθηκε μισθοδοσία: %s — %s (προς πληρωμή/λογιστήριο)' % (comp.legal_name if comp else '', period)
        # ειδοποίηση: προς το παρόν ΜΟΝΟ masteradmin (πλατφόρμα + email) — απόφαση Giannis
        masters = User.query.filter_by(role='masteradmin').all()
        for u in masters:
            notify(u.id, msg, link)
        emails = [u.email for u in masters if u.email]
        try:
            from app import send_email
            if emails:
                send_email('Εστία — Έγκριση μισθοδοσίας: %s %s' % (comp.legal_name if comp else '', period),
                           '<p>%s</p><p>Πίνακας ανά ξενοδοχείο: εξαγωγή από τη σελίδα εκτέλεσης.</p>' % msg,
                           emails)
        except Exception as e:
            print('approve email skipped:', e)
        log_activity('payroll_approve', msg)
    return redirect(request.referrer or url_for('payroll_control'))


@app.route('/dashboard/payroll/run/<int:rid>/markpaid', methods=['POST'])
def payroll_markpaid(rid):
    if not _padmin():
        return ('', 403)
    run = PayrollRun.query.get_or_404(rid)
    for l in PayrollLine.query.filter_by(run_id=rid).all():
        l.paid = round((l.in_hand or 0) + (l.extra_legal_net or 0), 2)
    run.status = 'paid'
    db.session.commit()
    log_activity('payroll_markpaid', str(rid))
    return redirect(request.referrer or url_for('payroll_run_view', rid=rid))


@app.route('/dashboard/payroll/run/<int:rid>/export.xlsx')
def payroll_run_export(rid):
    if not _padmin():
        return redirect(url_for('login'))
    if openpyxl is None:
        return ('openpyxl μη διαθέσιμο', 500)
    run = PayrollRun.query.get_or_404(rid)
    comp = Company.query.get(run.company_id)
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = 'Μισθοδοσία'
    umap = {u.id: u for u in User.query.all()}
    hmap = {h.id: h.name for h in Hotel.query.all()}
    hdr = ['Ξενοδοχείο', 'Εργαζόμενος', 'Εργάσιμες', 'Ρεπό', 'Έξτρα ώρες',
           'Management (στο χέρι)', 'Λογιστήριο (καθαρά)', 'Δώρα/άδεια',
           'Εκτός λογιστηρίου', 'Πληρωτέο', 'Πληρωμένο', 'Υπόλοιπο']
    ws.append(hdr)
    lines = PayrollLine.query.filter_by(run_id=rid).all()
    def hof(u): return hmap.get(getattr(u, 'home_hotel_id', None), '')
    lines.sort(key=lambda l: (hof(umap.get(l.user_id)), (umap.get(l.user_id).full_name if umap.get(l.user_id) else '')))
    for l in lines:
        u = umap.get(l.user_id)
        mgmt = round((l.in_hand or 0) + (l.extra_legal_net or 0), 2)
        ws.append([hof(u), (u.full_name if u else l.user_id), l.work_days or 0, l.repo or 0,
                   round(l.extra_hours or 0, 2), mgmt, (l.net_legal if l.net_legal is not None else ''),
                   round(l.extra_legal_net or 0, 2), round(mgmt - (l.net_legal or 0), 2),
                   mgmt, round(l.paid or 0, 2), round(mgmt - (l.paid or 0), 2)])
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    from flask import send_file
    fn = 'misthodosia_%s_%s_%s.xlsx' % ((comp.code if comp else 'X'), run.year, run.month)
    return send_file(bio, as_attachment=True, download_name=fn,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

print('payroll module loaded (Φ2.2)')
